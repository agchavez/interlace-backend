"""
Genera un archivo Excel de prueba con pautas realistas para CD LA GRANJA.

Uso:
    python scripts/generate_test_pautas.py

Columnas: Viaje, Transporte, Camión, Ruta, Cajas, SKUs,
          Pallets Completos, Fracciones Armadas, Complejidad

Mezcla de cargas (viaje 1) y recargas (viaje 2+). Camiones del catálogo real.
"""
import os
import sys
import random
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Camiones reales del catálogo (CD LA GRANJA) con pallet_spaces.
TRUCKS = [
    ('C-101', 22),
    ('C-102', 22),
    ('C-103', 18),
    ('C-104', 12),
    ('C-105', 20),
    ('C-106', 10),
    ('C-107', 24),
    ('C-108', 22),
]

ROUTES = [
    'R-TGU-01', 'R-TGU-02', 'R-TGU-03', 'R-TGU-04', 'R-TGU-05',
    'R-CMY-01', 'R-CMY-02', 'R-VLE-01',
]


def build_row(transport: str, trip: int, truck_code: str, pallet_spaces: int, route: str):
    """Genera valores realistas según el tamaño del camión."""
    # Para recargas (trip > 1) solemos llevar menos carga.
    load_factor = 1.0 if trip == 1 else random.uniform(0.35, 0.75)
    full_pallets = max(1, int(pallet_spaces * load_factor))
    assembled = random.randint(1, 4)
    boxes_per_full = random.randint(32, 50)
    boxes_per_fraction = random.randint(8, 20)
    total_boxes = full_pallets * boxes_per_full + assembled * boxes_per_fraction
    total_skus = random.randint(max(8, int(total_boxes / 22)), max(14, int(total_boxes / 12)))
    complexity = round(total_skus / max(1, full_pallets) + random.uniform(0.1, 1.2), 2)
    return [
        trip,
        transport,
        truck_code,
        route,
        total_boxes,
        total_skus,
        full_pallets,
        assembled,
        complexity,
    ]


def main():
    random.seed(42)  # reproducible

    rows = []
    base_transport = 4000825100

    # 6 cargas (viaje 1), una por camión (los primeros 6).
    for i, (code, spaces) in enumerate(TRUCKS[:6]):
        rows.append(build_row(
            transport=str(base_transport + i),
            trip=1,
            truck_code=code,
            pallet_spaces=spaces,
            route=ROUTES[i % len(ROUTES)],
        ))

    # 4 recargas (viaje 2) sobre transportes nuevos.
    for i in range(4):
        truck_idx = i  # usa los primeros 4 camiones
        code, spaces = TRUCKS[truck_idx]
        rows.append(build_row(
            transport=str(base_transport + 100 + i),
            trip=2,
            truck_code=code,
            pallet_spaces=spaces,
            route=ROUTES[(i + 3) % len(ROUTES)],
        ))

    # 2 terceras recargas (viaje 3).
    for i in range(2):
        code, spaces = TRUCKS[i]
        rows.append(build_row(
            transport=str(base_transport + 200 + i),
            trip=3,
            truck_code=code,
            pallet_spaces=spaces,
            route=ROUTES[(i + 5) % len(ROUTES)],
        ))

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pautas'

    columns = [
        ('Viaje', 10),
        ('Transporte', 18),
        ('Camión', 12),
        ('Ruta', 14),
        ('Cajas', 10),
        ('SKUs', 10),
        ('Pallets Completos', 18),
        ('Fracciones Armadas', 20),
        ('Complejidad', 14),
    ]

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, (col_name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    ws.freeze_panes = 'A2'

    out_path = Path(__file__).resolve().parents[2] / 'test_pautas.xlsx'
    wb.save(out_path)
    print(f'Generado: {out_path}')
    print(f'Filas: {len(rows)} ({sum(1 for r in rows if r[0] == 1)} cargas, '
          f'{sum(1 for r in rows if r[0] > 1)} recargas)')


if __name__ == '__main__':
    # Asegurar que el script corre desde la raíz del backend
    backend_root = Path(__file__).resolve().parents[1]
    os.chdir(backend_root)
    sys.path.insert(0, str(backend_root))
    main()
