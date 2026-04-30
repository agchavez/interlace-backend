"""
ViewSets para modelos principales del ciclo del camión
"""
import io
from collections import defaultdict
from datetime import date
from decimal import Decimal

from django.http import HttpResponse
from django.db.models import Count, Avg, Sum
from django.db.models.functions import Length
from django.utils import timezone
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django_filters.rest_framework import DjangoFilterBackend

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.truck_cycle.models.catalogs import TruckModel, ProductCatalogModel
from apps.truck_cycle.models.core import (
    PalletComplexUploadModel,
    PautaModel,
    PautaDeliveryDetailModel,
)
from apps.truck_cycle.models.operational import PautaTimestampModel, PautaAssignmentModel
from apps.truck_cycle.serializers.core_serializers import (
    PalletComplexUploadSerializer,
    PalletComplexUploadCreateSerializer,
    PautaListSerializer,
    PautaDetailSerializer,
)
from apps.truck_cycle.filters import PautaFilter


def get_user_distributor_center(request):
    """Obtener el centro de distribución seleccionado por el usuario"""
    try:
        if request.user.centro_distribucion_id:
            return request.user.centro_distribucion
        return request.user.personnel_profile.primary_distributor_center
    except Exception:
        return None


# Mapa de transiciones válidas: estado_actual -> (nuevo_estado, evento_timestamp)
STATUS_TRANSITIONS = {
    'assign_picker': {
        'from': ['PENDING_PICKING'],
        'to': 'PICKING_ASSIGNED',
        'event': None,
    },
    'start_picking': {
        'from': ['PICKING_ASSIGNED'],
        'to': 'PICKING_IN_PROGRESS',
        'event': 'T0_PICKING_START',
    },
    'complete_picking': {
        'from': ['PICKING_IN_PROGRESS'],
        'to': 'PICKING_DONE',
        'event': 'T1_PICKING_END',
    },
    'assign_yard_driver': {
        'from': ['PICKING_DONE'],
        'to': 'MOVING_TO_BAY',
        'event': 'T1A_YARD_START',
    },
    'position_at_bay': {
        'from': ['MOVING_TO_BAY'],
        'to': 'IN_BAY',
        'event': 'T1B_YARD_END',
    },
    'assign_bay': {
        'from': ['PICKING_DONE'],
        'to': 'IN_BAY',
        'event': 'T2_BAY_ASSIGNED',
    },
    'complete_loading': {
        'from': ['IN_BAY'],
        'to': 'PENDING_COUNT',
        'event': 'T4_LOADING_END',
    },
    'assign_counter': {
        'from': ['PENDING_COUNT'],
        'to': 'COUNTING',
        'event': 'T5_COUNT_START',
    },
    'complete_count': {
        'from': ['COUNTING'],
        'to': 'COUNTED',
        'event': 'T6_COUNT_END',
    },
    'move_to_parking': {
        'from': ['COUNTED', 'PENDING_CHECKOUT'],
        'to': 'MOVING_TO_PARKING',
        'event': 'T8A_YARD_RETURN_START',
    },
    'confirm_parked': {
        'from': ['MOVING_TO_PARKING'],
        'to': 'PARKED',
        'event': 'T8B_YARD_RETURN_END',
    },
    'checkout_security': {
        'from': ['PARKED'],
        'to': 'CHECKOUT_SECURITY',
        'event': 'T7_CHECKOUT_SECURITY',
    },
    'checkout_ops': {
        'from': ['PARKED', 'CHECKOUT_SECURITY'],
        'to': 'CHECKOUT_OPS',
        'event': 'T8_CHECKOUT_OPS',
    },
    'dispatch': {
        'from': ['PARKED', 'CHECKOUT_SECURITY', 'CHECKOUT_OPS'],
        'to': 'DISPATCHED',
        'event': 'T9_DISPATCH',
    },
    'arrival': {
        'from': ['DISPATCHED'],
        'to': 'IN_RELOAD_QUEUE',
        'event': 'T10_ARRIVAL',
    },
    'process_return': {
        'from': ['PENDING_RETURN', 'IN_RELOAD_QUEUE'],
        'to': 'RETURN_PROCESSED',
        'event': 'T13_RETURN_END',
    },
    'start_audit': {
        'from': ['RETURN_PROCESSED', 'COUNTED'],
        'to': 'IN_AUDIT',
        'event': 'T14_AUDIT_START',
    },
    'complete_audit': {
        'from': ['IN_AUDIT'],
        'to': 'AUDIT_COMPLETE',
        'event': 'T15_AUDIT_END',
    },
    'close': {
        'from': ['AUDIT_COMPLETE', 'DISPATCHED', 'RETURN_PROCESSED'],
        'to': 'CLOSED',
        'event': 'T16_CLOSE',
    },
}


class PalletComplexUploadViewSet(viewsets.ModelViewSet):
    """Gestión de cargas masivas de pautas"""
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_serializer_class(self):
        if self.action == 'create':
            return PalletComplexUploadCreateSerializer
        return PalletComplexUploadSerializer

    def get_queryset(self):
        dc = get_user_distributor_center(self.request)
        if dc:
            return PalletComplexUploadModel.objects.filter(distributor_center=dc)
        return PalletComplexUploadModel.objects.none()

    @action(detail=False, methods=['get'])
    def template(self, request):
        """Descargar plantilla Excel con las columnas esperadas (nivel resumen por transporte)"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pautas"

        columns = [
            ('Viaje', 12),
            ('Transporte', 18),
            ('Camión', 15),
            ('Ruta', 15),
            ('Cajas', 12),
            ('SKUs', 12),
            ('Pallets Completos', 18),
            ('Fracciones Armadas', 20),
            ('Complejidad %', 15),
        ]

        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_idx, (col_name, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Ejemplo de fila — complejidad como porcentaje (0-100). La placa
        # se resuelve desde el catálogo a partir del código del camión.
        example = [1, '4000825134', 'HN-1264', 'R-TGU-01', 775, 55, 1, 3, 7.52]
        for col_idx, val in enumerate(example, 1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            if col_idx == 9:
                cell.number_format = '0.00"%"'

        ws.freeze_panes = 'A2'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_pautas.xlsx"'
        return response

    @action(detail=False, methods=['post'])
    def preview(self, request):
        """
        Subir archivo Excel con pautas a nivel resumen (1 fila = 1 pauta).
        Columnas: Viaje, Transporte, Camión, Ruta, Cajas, SKUs,
                  Pallets Completos, Fracciones Armadas, Complejidad %
        Valida:
         - Camión debe existir en el catálogo del CD (no se crea on-the-fly).
         - (Transporte, Camión, Ruta) único por operational_date.
        """
        file = request.FILES.get('file')
        if not file:
            return Response(
                {'error': 'No se proporcionó archivo.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        dc = get_user_distributor_center(request)
        if not dc:
            return Response(
                {'error': 'Usuario sin centro de distribución asignado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
        except Exception:
            return Response(
                {'error': 'No se pudo leer el archivo Excel.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        operational_date = request.data.get('operational_date') or timezone.now().date()

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        errors = []
        warnings = []
        pautas = []
        missing_trucks = set()
        seen_keys = {}

        for row_idx, row in enumerate(rows, start=2):
            if not row or all(c is None for c in row):
                continue

            if len(row) < 5:
                errors.append(f"Fila {row_idx}: columnas insuficientes (mínimo 5).")
                continue

            viaje = str(row[0] or '').strip()
            transporte = str(row[1] or '').strip()
            camion = str(row[2] or '').strip()
            ruta = str(row[3] or '').strip() if len(row) > 3 else ''

            if not transporte:
                errors.append(f"Fila {row_idx}: Transporte es requerido.")
                continue
            if not viaje:
                errors.append(f"Fila {row_idx}: Viaje es requerido.")
                continue
            if not camion:
                errors.append(f"Fila {row_idx}: Camión es requerido.")
                continue

            try:
                cajas = int(float(row[4])) if len(row) > 4 and row[4] else 0
            except (ValueError, TypeError):
                errors.append(f"Fila {row_idx}: Cajas inválido '{row[4] if len(row) > 4 else ''}'.")
                continue

            try:
                skus = int(float(row[5])) if len(row) > 5 and row[5] else 0
            except (ValueError, TypeError):
                skus = 0

            try:
                completas = int(float(row[6])) if len(row) > 6 and row[6] else 0
            except (ValueError, TypeError):
                completas = 0

            try:
                fracciones = int(float(row[7])) if len(row) > 7 and row[7] else 0
            except (ValueError, TypeError):
                fracciones = 0

            # Complejidad ahora es porcentaje (0-100). Aceptamos también valores
            # legacy < 1 (por si alguien sube un archivo viejo) y los convertimos.
            try:
                raw_complex = float(row[8]) if len(row) > 8 and row[8] is not None and row[8] != '' else 0
                if 0 < raw_complex <= 1:
                    raw_complex = raw_complex * 100
                complejidad = round(raw_complex, 2)
            except (ValueError, TypeError):
                complejidad = 0

            # Validar existencia del camión en el catálogo del CD.
            truck = TruckModel.objects.filter(
                code=camion, distributor_center=dc, is_active=True,
            ).first()
            if not truck:
                errors.append(
                    f"Fila {row_idx}: Camión '{camion}' no existe en el catálogo del centro de distribución."
                )
                missing_trucks.add(camion)
                continue

            # Validar unicidad (transporte, camión, ruta) dentro del archivo
            dup_key = (transporte, camion, ruta)
            if dup_key in seen_keys:
                errors.append(
                    f"Fila {row_idx}: Duplicado dentro del archivo — ya aparece en fila {seen_keys[dup_key]} "
                    f"(Transporte '{transporte}', Camión '{camion}', Ruta '{ruta}')."
                )
                continue
            seen_keys[dup_key] = row_idx

            # Validar unicidad contra pautas existentes del día
            exists = PautaModel.objects.filter(
                transport_number=transporte,
                truck=truck,
                route_code=ruta,
                operational_date=operational_date,
                distributor_center=dc,
            ).exists()
            if exists:
                errors.append(
                    f"Fila {row_idx}: Ya existe una pauta para {operational_date} con "
                    f"Transporte '{transporte}', Camión '{camion}', Ruta '{ruta}'."
                )
                continue

            pautas.append({
                'trip_number': viaje,
                'transport_number': transporte,
                'truck_code': truck.code,
                'truck_plate': truck.plate,
                'route_code': ruta,
                'total_boxes': cajas,
                'total_skus': skus,
                'full_pallets': completas,
                'assembled_fractions': fracciones,
                'complexity_score': complejidad,
            })

        # Crear registro de upload en estado PREVIEW
        upload = PalletComplexUploadModel.objects.create(
            file_name=file.name,
            file=file,
            status='PREVIEW',
            errors_json={'errors': errors, 'missing_trucks': sorted(missing_trucks)},
            row_count=len(rows),
            distributor_center=dc,
            uploaded_by=request.user,
        )

        return Response({
            'upload_id': upload.id,
            'file_name': file.name,
            'row_count': len(rows),
            'pautas_count': len(pautas),
            'errors': errors,
            'warnings': warnings,
            'missing_trucks': sorted(missing_trucks),
            'pautas_preview': pautas,
        })

    @action(detail=True, methods=['post'])
    def confirm(self, request, pk=None):
        """
        Confirmar una carga y crear las pautas a partir del archivo subido.
        Valida:
         - Camión debe existir en el catálogo del CD.
         - (Transporte, Camión, Ruta) único por operational_date (vs archivo y vs BD).

        Body opcional:
         - skip_invalid (bool): si True, omite las filas con error y crea
           solo las válidas (en lugar de abortar toda la carga).
        """
        upload = self.get_object()

        if upload.status != 'PREVIEW':
            return Response(
                {'error': 'Solo se pueden confirmar cargas en estado Vista Previa.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        dc = get_user_distributor_center(request)
        if not dc:
            return Response(
                {'error': 'Usuario sin centro de distribución asignado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        operational_date = request.data.get('operational_date') or timezone.now().date()

        try:
            wb = openpyxl.load_workbook(upload.file, read_only=True)
            ws = wb.active
        except Exception:
            return Response(
                {'error': 'No se pudo leer el archivo guardado.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        errors = []
        missing_trucks = set()
        to_create = []
        seen_keys = {}

        for row_idx, row in enumerate(rows, start=2):
            if not row or all(c is None for c in row):
                continue
            if len(row) < 5:
                continue

            viaje = str(row[0] or '').strip()
            transporte = str(row[1] or '').strip()
            camion_code = str(row[2] or '').strip()
            ruta = str(row[3] or '').strip() if len(row) > 3 else ''

            if not transporte or not viaje or not camion_code:
                errors.append(f"Fila {row_idx}: Transporte, Viaje y Camión son requeridos.")
                continue

            try:
                cajas = int(float(row[4])) if len(row) > 4 and row[4] else 0
            except (ValueError, TypeError):
                errors.append(f"Fila {row_idx}: Cajas inválido.")
                continue

            try:
                skus = int(float(row[5])) if len(row) > 5 and row[5] else 0
            except (ValueError, TypeError):
                skus = 0

            try:
                completas = int(float(row[6])) if len(row) > 6 and row[6] else 0
            except (ValueError, TypeError):
                completas = 0

            try:
                fracciones = int(float(row[7])) if len(row) > 7 and row[7] else 0
            except (ValueError, TypeError):
                fracciones = 0

            try:
                raw_complex = float(row[8]) if len(row) > 8 and row[8] is not None and row[8] != '' else 0
                if 0 < raw_complex <= 1:
                    raw_complex = raw_complex * 100
                complejidad = round(raw_complex, 2)
            except (ValueError, TypeError):
                complejidad = 0

            truck = TruckModel.objects.filter(
                code=camion_code, distributor_center=dc, is_active=True,
            ).first()
            if not truck:
                errors.append(
                    f"Fila {row_idx}: Camión '{camion_code}' no existe en el catálogo del CD."
                )
                missing_trucks.add(camion_code)
                continue

            dup_key = (transporte, camion_code, ruta)
            if dup_key in seen_keys:
                errors.append(
                    f"Fila {row_idx}: Duplicado con fila {seen_keys[dup_key]} "
                    f"(Transporte '{transporte}', Camión '{camion_code}', Ruta '{ruta}')."
                )
                continue
            seen_keys[dup_key] = row_idx

            if PautaModel.objects.filter(
                transport_number=transporte,
                truck=truck,
                route_code=ruta,
                operational_date=operational_date,
                distributor_center=dc,
            ).exists():
                errors.append(
                    f"Fila {row_idx}: Ya existe una pauta para {operational_date} con "
                    f"Transporte '{transporte}', Camión '{camion_code}', Ruta '{ruta}'."
                )
                continue

            to_create.append({
                'transport_number': transporte,
                'trip_number': viaje,
                'route_code': ruta,
                'total_boxes': cajas,
                'total_skus': skus,
                'total_pallets': completas,
                'assembled_fractions': fracciones,
                'complexity_score': complejidad,
                'truck': truck,
                'is_reload': str(viaje).strip() != '1',
            })

        skip_invalid = bool(request.data.get('skip_invalid'))

        if errors and not skip_invalid:
            return Response(
                {
                    'error': 'La carga tiene errores. Corrija y vuelva a subir el archivo, o reintente omitiendo filas inválidas.',
                    'errors': errors,
                    'missing_trucks': sorted(missing_trucks),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if skip_invalid and not to_create:
            return Response(
                {
                    'error': 'Ninguna fila es válida — no hay pautas para crear.',
                    'errors': errors,
                    'missing_trucks': sorted(missing_trucks),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        created_pautas = []
        for data in to_create:
            pauta = PautaModel.objects.create(
                status='PENDING_PICKING',
                operational_date=operational_date,
                upload=upload,
                distributor_center=dc,
                **data,
            )
            created_pautas.append(pauta.id)

        upload.status = 'CONFIRMED'
        upload.save(update_fields=['status'])

        return Response({
            'message': (
                f'Se crearon {len(created_pautas)} pautas'
                + (f' (se omitieron {len(errors)} filas inválidas)' if errors else '')
                + '.'
            ),
            'upload_id': upload.id,
            'pauta_ids': created_pautas,
            'skipped_errors': errors if skip_invalid else [],
        })


class PautaViewSet(viewsets.ModelViewSet):
    """Gestión de pautas del ciclo del camión"""
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = PautaFilter

    def get_serializer_class(self):
        if self.action == 'list':
            return PautaListSerializer
        return PautaDetailSerializer

    def get_queryset(self):
        dc = get_user_distributor_center(self.request)
        if not dc:
            return PautaModel.objects.none()
        # Orden estable: día desc, dentro del día por viaje ascendente (primer
        # viaje arriba), luego por transporte y created_at. trip_number es
        # CharField; se ordena primero por longitud y luego alfabético para
        # que "1" < "2" < "9" < "10" < "99" sin necesidad de castear a int
        # (evita fallas si algún valor llegara no-numérico).
        return (
            PautaModel.objects
            .filter(distributor_center=dc)
            .select_related('truck', 'upload', 'distributor_center')
            .annotate(_trip_len=Length('trip_number'))
            .order_by('-operational_date', '_trip_len', 'trip_number', 'transport_number', 'created_at')
        )

    @action(detail=False, methods=['post'], url_path='manual-create')
    def manual_create(self, request):
        """Carga emergente: crear una sola pauta manualmente sin Excel.

        Body:
          trip_number, transport_number, truck_code (o truck_id), route_code,
          total_boxes, total_skus, total_pallets, assembled_fractions,
          complexity_score (en %), operational_date (opcional, default hoy).
        """
        dc = get_user_distributor_center(request)
        if not dc:
            return Response(
                {'error': 'Usuario sin centro de distribución asignado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        truck_id = request.data.get('truck_id')
        truck_code = (request.data.get('truck_code') or '').strip()
        truck = None
        if truck_id:
            truck = TruckModel.objects.filter(id=truck_id, distributor_center=dc, is_active=True).first()
        if not truck and truck_code:
            truck = TruckModel.objects.filter(code=truck_code, distributor_center=dc, is_active=True).first()
        if not truck:
            return Response(
                {'error': 'Camión no encontrado o no pertenece al centro de distribución.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        transporte = (request.data.get('transport_number') or '').strip()
        viaje = str(request.data.get('trip_number') or '').strip()
        ruta = (request.data.get('route_code') or '').strip()
        if not transporte or not viaje:
            return Response(
                {'error': 'Transporte y Viaje son requeridos.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        operational_date = request.data.get('operational_date') or timezone.localdate()

        if PautaModel.objects.filter(
            transport_number=transporte,
            truck=truck,
            route_code=ruta,
            operational_date=operational_date,
            distributor_center=dc,
        ).exists():
            return Response(
                {'error': f'Ya existe una pauta para {operational_date} con esos datos.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        def _int(name, default=0):
            try:
                return int(float(request.data.get(name) or 0))
            except (TypeError, ValueError):
                return default

        try:
            raw_complex = float(request.data.get('complexity_score') or 0)
            if 0 < raw_complex <= 1:
                raw_complex = raw_complex * 100
            complejidad = round(raw_complex, 2)
        except (TypeError, ValueError):
            complejidad = 0

        pauta = PautaModel.objects.create(
            status='PENDING_PICKING',
            operational_date=operational_date,
            distributor_center=dc,
            truck=truck,
            transport_number=transporte,
            trip_number=viaje,
            route_code=ruta,
            total_boxes=_int('total_boxes'),
            total_skus=_int('total_skus'),
            total_pallets=_int('total_pallets'),
            assembled_fractions=_int('assembled_fractions'),
            complexity_score=complejidad,
            is_reload=viaje != '1',
            notes=request.data.get('notes', '') or 'Carga emergente (manual)',
        )

        return Response(PautaDetailSerializer(pauta).data, status=status.HTTP_201_CREATED)

    def _do_transition(self, request, pk, action_name):
        """Ejecutar una transición de estado"""
        pauta = self.get_object()
        transition = STATUS_TRANSITIONS.get(action_name)

        if not transition:
            return Response(
                {'error': f'Acción "{action_name}" no reconocida.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if pauta.status not in transition['from']:
            return Response(
                {'error': f'No se puede ejecutar "{action_name}" desde el estado "{pauta.get_status_display()}".'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        pauta.status = transition['to']
        pauta.save(update_fields=['status'])

        # Crear timestamp si corresponde
        if transition['event']:
            PautaTimestampModel.objects.create(
                event_type=transition['event'],
                pauta=pauta,
                recorded_by=request.user,
                notes=request.data.get('notes', ''),
            )

        serializer = PautaDetailSerializer(pauta)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def assign_picker(self, request, pk=None):
        """Asignar picker a la pauta"""
        response = self._do_transition(request, pk, 'assign_picker')
        if response.status_code == 200:
            pauta = self.get_object()
            personnel_id = request.data.get('personnel_id')
            if personnel_id:
                PautaAssignmentModel.objects.create(
                    role='PICKER',
                    pauta=pauta,
                    personnel_id=personnel_id,
                    assigned_by=request.user,
                )
        return response

    @action(detail=False, methods=['get'])
    def picker_stats(self, request):
        """
        Stats del turno del picker autenticado (o cualquier personnel via ?personnel_id=).

        Devuelve:
          date
          completed_count           — pautas con picking completado hoy
          total_boxes               — cajas totales de esas pautas
          avg_picking_minutes       — promedio de minutos por pauta (T0→T1)
          boxes_per_hour            — productividad (cajas / hora de picking)
          in_progress               — pauta en progreso (si hay) con started_at
        """
        try:
            profile = request.user.personnel_profile
        except Exception:
            return Response(
                {'error': 'El usuario no tiene perfil de personal asociado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Admin puede mirar las stats de otro picker.
        personnel_id = request.query_params.get('personnel_id')
        if personnel_id and (request.user.is_superuser or request.user.is_staff):
            from apps.personnel.models.personnel import PersonnelProfile
            try:
                profile = PersonnelProfile.objects.get(pk=personnel_id)
            except PersonnelProfile.DoesNotExist:
                return Response({'error': 'Personnel no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        operational_date = request.query_params.get('operational_date') or timezone.localdate()

        # Pautas donde este profile fue picker y ya pasó de picking.
        done_statuses = [
            'PICKING_DONE', 'MOVING_TO_BAY', 'IN_BAY', 'PENDING_COUNT', 'COUNTING',
            'COUNTED', 'MOVING_TO_PARKING', 'PARKED',
            'PENDING_CHECKOUT', 'CHECKOUT_SECURITY', 'CHECKOUT_OPS',
            'DISPATCHED', 'IN_RELOAD_QUEUE', 'PENDING_RETURN', 'RETURN_PROCESSED',
            'IN_AUDIT', 'AUDIT_COMPLETE', 'CLOSED',
        ]
        pautas_done = PautaModel.objects.filter(
            operational_date=operational_date,
            assignments__role='PICKER',
            assignments__personnel=profile,
            status__in=done_statuses,
        ).distinct()

        completed = 0
        total_boxes = 0
        total_minutes = 0.0

        for p in pautas_done:
            t0 = p.timestamps.filter(event_type='T0_PICKING_START').order_by('timestamp').first()
            t1 = p.timestamps.filter(event_type='T1_PICKING_END').order_by('timestamp').first()
            if t0 and t1:
                delta = (t1.timestamp - t0.timestamp).total_seconds() / 60
                if delta > 0:
                    completed += 1
                    total_minutes += delta
                    total_boxes += p.total_boxes

        avg_minutes = (total_minutes / completed) if completed else None
        boxes_per_hour = (total_boxes / (total_minutes / 60)) if total_minutes > 0 else None

        # Pauta en progreso (si hay).
        in_progress = PautaModel.objects.filter(
            operational_date=operational_date,
            status='PICKING_IN_PROGRESS',
            assignments__role='PICKER',
            assignments__personnel=profile,
            assignments__is_active=True,
        ).distinct().first()

        in_progress_data = None
        if in_progress:
            t0 = in_progress.timestamps.filter(event_type='T0_PICKING_START').order_by('timestamp').first()
            in_progress_data = {
                'id': in_progress.id,
                'transport_number': in_progress.transport_number,
                'total_boxes': in_progress.total_boxes,
                'started_at': t0.timestamp.isoformat() if t0 else None,
            }

        return Response({
            'date': str(operational_date),
            'completed_count': completed,
            'total_boxes': total_boxes,
            'avg_picking_minutes': round(avg_minutes, 1) if avg_minutes is not None else None,
            'boxes_per_hour': round(boxes_per_hour, 1) if boxes_per_hour is not None else None,
            'in_progress': in_progress_data,
        })

    @action(detail=True, methods=['post'])
    def take_as_picker(self, request, pk=None):
        """Asigna un picker a la pauta (PENDING_PICKING → PICKING_ASSIGNED).

        El dispositivo lo opera un supervisor/checador (los pickers no tienen
        celular), por lo que se debe enviar `personnel_id` en el body con la
        persona que realmente va a armar la pauta. Si no viene, se intenta
        como fallback el perfil del usuario autenticado (para compatibilidad).

        Valida que el personnel tenga position_type ∈ {PICKER, LOADER} y
        pertenezca al mismo centro de distribución de la pauta.
        """
        from apps.personnel.models.personnel import PersonnelProfile

        pauta = self.get_object()

        personnel_id = request.data.get('personnel_id')
        if personnel_id:
            try:
                profile = PersonnelProfile.objects.get(pk=personnel_id, is_active=True)
            except PersonnelProfile.DoesNotExist:
                return Response(
                    {'error': 'Personal no encontrado o inactivo.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if profile.position_type not in ('PICKER', 'LOADER'):
                return Response(
                    {'error': f'El personal "{profile.first_name} {profile.last_name}" no es PICKER ni LOADER.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            picker_dc_id = profile.primary_distributor_center_id
            if picker_dc_id and picker_dc_id != pauta.distributor_center_id:
                return Response(
                    {'error': 'El picker seleccionado no pertenece al mismo centro de distribución.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            try:
                profile = request.user.personnel_profile
            except Exception:
                return Response(
                    {'error': 'Debe enviar personnel_id o el usuario debe tener perfil.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        if pauta.status != 'PENDING_PICKING':
            return Response(
                {'error': f'La pauta ya está en estado "{pauta.get_status_display()}" y no puede tomarse.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        pauta.status = 'PICKING_ASSIGNED'
        pauta.save(update_fields=['status'])
        PautaAssignmentModel.objects.create(
            role='PICKER',
            pauta=pauta,
            personnel=profile,
            assigned_by=request.user,
        )
        return Response(PautaDetailSerializer(pauta).data)

    @action(detail=True, methods=['post'])
    def take_as_counter(self, request, pk=None):
        """Auto-asignación del contador. Acepta IN_BAY o PENDING_COUNT.

        Si la pauta viene en IN_BAY, se emite T4_LOADING_END automáticamente
        antes del T5_COUNT_START. Esto evita depender del paso manual
        "Carga OK" (IN_BAY → PENDING_COUNT) que un supervisor tendría que
        apretar desde /truck-cycle/operations: ese delay distorsiona las
        mediciones de tiempo.
        """
        try:
            profile = request.user.personnel_profile
        except Exception:
            return Response(
                {'error': 'El usuario no tiene un perfil de personal asociado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        pauta = self.get_object()
        if pauta.status not in ('IN_BAY', 'PENDING_COUNT'):
            return Response(
                {'error': f'La pauta está en estado "{pauta.get_status_display()}" y no puede tomarse para conteo.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if pauta.status == 'IN_BAY':
            PautaTimestampModel.objects.create(
                event_type='T4_LOADING_END',
                pauta=pauta,
                recorded_by=request.user,
            )

        pauta.status = 'COUNTING'
        pauta.save(update_fields=['status'])
        PautaAssignmentModel.objects.create(
            role='COUNTER',
            pauta=pauta,
            personnel=profile,
            assigned_by=request.user,
        )
        PautaTimestampModel.objects.create(
            event_type='T5_COUNT_START',
            pauta=pauta,
            recorded_by=request.user,
        )
        return Response(PautaDetailSerializer(pauta).data)

    @action(detail=True, methods=['post'])
    def take_as_security(self, request, pk=None):
        """
        Auto-validación: el guardia autenticado toma y valida seguridad en
        una pauta (PARKED → CHECKOUT_SECURITY). Es un paso opcional.
        """
        from apps.truck_cycle.models.operational import CheckoutValidationModel
        try:
            profile = request.user.personnel_profile
        except Exception:
            return Response(
                {'error': 'El usuario no tiene un perfil de personal asociado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        pauta = self.get_object()
        if pauta.status != 'PARKED':
            return Response(
                {'error': f'La pauta está en estado "{pauta.get_status_display()}" — solo se puede validar seguridad cuando está estacionada.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        pauta.status = 'CHECKOUT_SECURITY'
        pauta.save(update_fields=['status'])
        PautaTimestampModel.objects.create(
            event_type='T7_CHECKOUT_SECURITY',
            pauta=pauta,
            recorded_by=request.user,
        )
        checkout, _ = CheckoutValidationModel.objects.get_or_create(pauta=pauta)
        checkout.security_validated = True
        checkout.security_validated_at = timezone.now()
        checkout.security_validator = profile
        exit_pass = request.data.get('exit_pass_consumables')
        if exit_pass is not None:
            checkout.exit_pass_consumables = exit_pass
        checkout.save()
        return Response(PautaDetailSerializer(pauta).data)

    @action(detail=False, methods=['get'])
    def security_stats(self, request):
        """Stats del turno del guardia autenticado (validaciones de seguridad)."""
        try:
            profile = request.user.personnel_profile
        except Exception:
            return Response({'error': 'Sin perfil.'}, status=status.HTTP_400_BAD_REQUEST)

        personnel_id = request.query_params.get('personnel_id')
        if personnel_id and (request.user.is_superuser or request.user.is_staff):
            from apps.personnel.models.personnel import PersonnelProfile
            try:
                profile = PersonnelProfile.objects.get(pk=personnel_id)
            except PersonnelProfile.DoesNotExist:
                return Response({'error': 'Personnel no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        operational_date = request.query_params.get('operational_date') or timezone.localdate()

        # Pautas validadas por este profile (pasaron de PARKED → CHECKOUT_SECURITY).
        done_statuses = [
            'CHECKOUT_SECURITY', 'CHECKOUT_OPS', 'DISPATCHED',
            'IN_RELOAD_QUEUE', 'PENDING_RETURN', 'RETURN_PROCESSED',
            'IN_AUDIT', 'AUDIT_COMPLETE', 'CLOSED',
        ]
        pautas_done = PautaModel.objects.filter(
            operational_date=operational_date,
            checkout_validation__security_validator=profile,
            status__in=done_statuses,
        ).distinct()

        completed = 0
        total_boxes = 0
        total_minutes = 0.0
        for p in pautas_done:
            t0 = p.timestamps.filter(event_type='T8B_YARD_RETURN_END').order_by('timestamp').first()
            t1 = p.timestamps.filter(event_type='T7_CHECKOUT_SECURITY').order_by('timestamp').first()
            if t0 and t1:
                delta = (t1.timestamp - t0.timestamp).total_seconds() / 60
                if delta >= 0:
                    completed += 1
                    total_minutes += delta
                    total_boxes += p.total_boxes

        avg_minutes = (total_minutes / completed) if completed else None
        pautas_per_hour = (completed / (total_minutes / 60)) if total_minutes > 0 else None

        return Response({
            'date': str(operational_date),
            'completed_count': completed,
            'total_boxes': total_boxes,
            'avg_validation_minutes': round(avg_minutes, 1) if avg_minutes is not None else None,
            'pautas_per_hour': round(pautas_per_hour, 1) if pautas_per_hour is not None else None,
        })

    @action(detail=True, methods=['post'])
    def take_as_ops(self, request, pk=None):
        """
        Auto-validación: el validador de operaciones toma y valida ops en
        una pauta (PARKED o CHECKOUT_SECURITY → CHECKOUT_OPS). Opcional.
        """
        from apps.truck_cycle.models.operational import CheckoutValidationModel
        try:
            profile = request.user.personnel_profile
        except Exception:
            return Response({'error': 'Sin perfil.'}, status=status.HTTP_400_BAD_REQUEST)
        pauta = self.get_object()
        if pauta.status not in ('PARKED', 'CHECKOUT_SECURITY'):
            return Response(
                {'error': f'La pauta está en estado "{pauta.get_status_display()}" — no se puede validar operaciones.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        pauta.status = 'CHECKOUT_OPS'
        pauta.save(update_fields=['status'])
        PautaTimestampModel.objects.create(
            event_type='T8_CHECKOUT_OPS',
            pauta=pauta,
            recorded_by=request.user,
        )
        checkout, _ = CheckoutValidationModel.objects.get_or_create(pauta=pauta)
        checkout.ops_validated = True
        checkout.ops_validated_at = timezone.now()
        checkout.ops_validator = profile
        checkout.save()
        return Response(PautaDetailSerializer(pauta).data)

    @action(detail=False, methods=['get'])
    def ops_stats(self, request):
        """Stats del turno del validador de operaciones."""
        try:
            profile = request.user.personnel_profile
        except Exception:
            return Response({'error': 'Sin perfil.'}, status=status.HTTP_400_BAD_REQUEST)

        personnel_id = request.query_params.get('personnel_id')
        if personnel_id and (request.user.is_superuser or request.user.is_staff):
            from apps.personnel.models.personnel import PersonnelProfile
            try:
                profile = PersonnelProfile.objects.get(pk=personnel_id)
            except PersonnelProfile.DoesNotExist:
                return Response({'error': 'Personnel no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        operational_date = request.query_params.get('operational_date') or timezone.localdate()

        done_statuses = [
            'CHECKOUT_OPS', 'DISPATCHED', 'IN_RELOAD_QUEUE',
            'PENDING_RETURN', 'RETURN_PROCESSED', 'IN_AUDIT', 'AUDIT_COMPLETE', 'CLOSED',
        ]
        pautas_done = PautaModel.objects.filter(
            operational_date=operational_date,
            checkout_validation__ops_validator=profile,
            status__in=done_statuses,
        ).distinct()

        completed = 0
        total_boxes = 0
        total_minutes = 0.0
        for p in pautas_done:
            # Inicio = T7 si seguridad validó, si no T8B (estacionado).
            t_security = p.timestamps.filter(event_type='T7_CHECKOUT_SECURITY').order_by('timestamp').first()
            t_parked = p.timestamps.filter(event_type='T8B_YARD_RETURN_END').order_by('timestamp').first()
            t0 = t_security or t_parked
            t1 = p.timestamps.filter(event_type='T8_CHECKOUT_OPS').order_by('timestamp').first()
            if t0 and t1:
                delta = (t1.timestamp - t0.timestamp).total_seconds() / 60
                if delta >= 0:
                    completed += 1
                    total_minutes += delta
                    total_boxes += p.total_boxes

        avg_minutes = (total_minutes / completed) if completed else None
        pautas_per_hour = (completed / (total_minutes / 60)) if total_minutes > 0 else None

        return Response({
            'date': str(operational_date),
            'completed_count': completed,
            'total_boxes': total_boxes,
            'avg_validation_minutes': round(avg_minutes, 1) if avg_minutes is not None else None,
            'pautas_per_hour': round(pautas_per_hour, 1) if pautas_per_hour is not None else None,
        })

    @action(detail=False, methods=['get'])
    def vendor_stats(self, request):
        """
        Stats del turno del chofer vendedor autenticado.
        Incluye active_trip con el timestamp T9_DISPATCH para el contador en vivo.
        """
        from django.db.models import Q
        try:
            profile = request.user.personnel_profile
        except Exception:
            return Response({'error': 'Sin perfil.'}, status=status.HTTP_400_BAD_REQUEST)

        operational_date = request.query_params.get('operational_date') or timezone.localdate()

        qs = PautaModel.objects.filter(
            operational_date=operational_date,
            is_reload=True,  # Vendedores solo ven recargas, no el primer viaje.
        ).filter(
            Q(truck__primary_driver=profile)
            | Q(assignments__role='DELIVERY_DRIVER', assignments__is_active=True, assignments__personnel=profile)
        ).distinct()

        trips_dispatched = qs.filter(
            status__in=['DISPATCHED', 'IN_RELOAD_QUEUE', 'PENDING_RETURN', 'RETURN_PROCESSED',
                        'IN_AUDIT', 'AUDIT_COMPLETE', 'CLOSED'],
        ).count()
        trips_completed = qs.filter(status='CLOSED').count()
        total_boxes = sum(p.total_boxes for p in qs.filter(status__in=[
            'DISPATCHED', 'IN_RELOAD_QUEUE', 'PENDING_RETURN', 'RETURN_PROCESSED',
            'IN_AUDIT', 'AUDIT_COMPLETE', 'CLOSED',
        ]))

        # Viaje activo: el más prioritario (PENDING_RETURN > IN_RELOAD_QUEUE > DISPATCHED).
        active_statuses_order = ['PENDING_RETURN', 'IN_RELOAD_QUEUE', 'DISPATCHED']
        active_pauta = None
        for s in active_statuses_order:
            active_pauta = qs.filter(status=s).order_by('id').first()
            if active_pauta:
                break

        active_count = qs.filter(status__in=['DISPATCHED', 'IN_RELOAD_QUEUE', 'PENDING_RETURN']).count()

        active_trip = None
        if active_pauta:
            dispatch_ts = active_pauta.timestamps.filter(event_type='T9_DISPATCH').order_by('timestamp').first()
            trip_start_ts = active_pauta.timestamps.filter(event_type='T9B_TRIP_START').order_by('timestamp').first()
            active_trip = {
                'id': active_pauta.id,
                'transport_number': active_pauta.transport_number,
                'status': active_pauta.status,
                'truck_code': active_pauta.truck.code if active_pauta.truck else None,
                'dispatched_at': dispatch_ts.timestamp.isoformat() if dispatch_ts else None,
                'trip_started_at': trip_start_ts.timestamp.isoformat() if trip_start_ts else None,
            }

        return Response({
            'date': str(operational_date),
            'trips_dispatched': trips_dispatched,
            'trips_completed': trips_completed,
            'active': active_count,
            'total_boxes': total_boxes,
            'active_trip': active_trip,
        })

    @action(detail=True, methods=['post'])
    def start_trip(self, request, pk=None):
        """
        Chofer vendedor marca inicio del viaje (ya se subió al camión y sale).
        Crea timestamp T9B_TRIP_START. Solo permitido si status=DISPATCHED y aún
        no se había marcado inicio.
        """
        pauta = self.get_object()
        if pauta.status != 'DISPATCHED':
            return Response(
                {'error': f'La pauta está en "{pauta.get_status_display()}" — no se puede iniciar viaje.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if pauta.timestamps.filter(event_type='T9B_TRIP_START').exists():
            return Response(
                {'error': 'El viaje ya fue iniciado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        PautaTimestampModel.objects.create(
            event_type='T9B_TRIP_START',
            pauta=pauta,
            recorded_by=request.user,
        )
        return Response(PautaDetailSerializer(pauta).data)

    @action(detail=True, methods=['post'])
    def take_as_yard_driver(self, request, pk=None):
        """
        Auto-asignación: el chofer de patio toma una pauta PICKING_DONE,
        la pone en MOVING_TO_BAY y arranca el cronómetro de movimiento.
        Solo aplica a cargas (trip 1, is_reload=False).
        """
        try:
            profile = request.user.personnel_profile
        except Exception:
            return Response(
                {'error': 'El usuario no tiene un perfil de personal asociado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        pauta = self.get_object()
        if pauta.is_reload:
            return Response(
                {'error': 'Las recargas no pasan por el chofer de patio.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if pauta.status != 'PICKING_DONE':
            return Response(
                {'error': f'La pauta está en "{pauta.get_status_display()}" y no puede tomarse.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        pauta.status = 'MOVING_TO_BAY'
        pauta.save(update_fields=['status'])
        PautaAssignmentModel.objects.create(
            role='YARD_DRIVER',
            pauta=pauta,
            personnel=profile,
            assigned_by=request.user,
        )
        PautaTimestampModel.objects.create(
            event_type='T1A_YARD_START',
            pauta=pauta,
            recorded_by=request.user,
        )
        return Response(PautaDetailSerializer(pauta).data)

    @action(detail=True, methods=['post'])
    def take_bay_for_return(self, request, pk=None):
        """Yard driver toma un camión desde la bahía para llevarlo al estacionamiento.

        COUNTED/PENDING_CHECKOUT → MOVING_TO_PARKING. Emite T8A_YARD_RETURN_START
        y deja la asignación de YARD_DRIVER. La bay_assignment se libera al
        confirmar estacionamiento (park_truck).
        """
        try:
            profile = request.user.personnel_profile
        except Exception:
            return Response(
                {'error': 'El usuario no tiene un perfil de personal asociado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        pauta = self.get_object()
        if pauta.status not in ('COUNTED', 'PENDING_CHECKOUT'):
            return Response(
                {'error': f'La pauta está en "{pauta.get_status_display()}" — solo se puede mover a estacionamiento desde Contado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if pauta.timestamps.filter(event_type='T8A_YARD_RETURN_START').exists():
            return Response(
                {'error': 'Ya se registró el inicio del movimiento Bahía→Estacionamiento.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        pauta.status = 'MOVING_TO_PARKING'
        pauta.save(update_fields=['status'])
        PautaAssignmentModel.objects.create(
            role='YARD_DRIVER',
            pauta=pauta,
            personnel=profile,
            assigned_by=request.user,
        )
        PautaTimestampModel.objects.create(
            event_type='T8A_YARD_RETURN_START',
            pauta=pauta,
            recorded_by=request.user,
        )
        return Response(PautaDetailSerializer(pauta).data)

    @action(detail=True, methods=['post'])
    def park_truck(self, request, pk=None):
        """Yard driver confirma que el camión quedó estacionado.

        MOVING_TO_PARKING → PARKED. Emite T8B_YARD_RETURN_END y libera la
        bay_assignment abierta.
        """
        pauta = self.get_object()
        if pauta.status != 'MOVING_TO_PARKING':
            return Response(
                {'error': f'La pauta está en "{pauta.get_status_display()}" — solo se puede confirmar estacionado mientras se mueve a estacionamiento.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if pauta.timestamps.filter(event_type='T8B_YARD_RETURN_END').exists():
            return Response(
                {'error': 'El camión ya fue estacionado (T8B registrado).'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        pauta.status = 'PARKED'
        pauta.save(update_fields=['status'])
        PautaTimestampModel.objects.create(
            event_type='T8B_YARD_RETURN_END',
            pauta=pauta,
            recorded_by=request.user,
        )
        bay_assignment = getattr(pauta, 'bay_assignment', None)
        if bay_assignment and not bay_assignment.released_at:
            bay_assignment.released_at = timezone.now()
            bay_assignment.save(update_fields=['released_at'])
        return Response(PautaDetailSerializer(pauta).data)

    @action(detail=False, methods=['get'])
    def yard_stats(self, request):
        """
        Stats del turno del chofer de patio autenticado.
        Devuelve: completed_count (pautas movidas), total_boxes,
        avg_movement_minutes (T1A→T1B), boxes_per_hour, in_progress.
        """
        try:
            profile = request.user.personnel_profile
        except Exception:
            return Response(
                {'error': 'Sin perfil.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        personnel_id = request.query_params.get('personnel_id')
        if personnel_id and (request.user.is_superuser or request.user.is_staff):
            from apps.personnel.models.personnel import PersonnelProfile
            try:
                profile = PersonnelProfile.objects.get(pk=personnel_id)
            except PersonnelProfile.DoesNotExist:
                return Response({'error': 'Personnel no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        operational_date = request.query_params.get('operational_date') or timezone.localdate()

        done_statuses = [
            'IN_BAY', 'PENDING_COUNT', 'COUNTING', 'COUNTED',
            'MOVING_TO_PARKING', 'PARKED',
            'PENDING_CHECKOUT', 'CHECKOUT_SECURITY', 'CHECKOUT_OPS',
            'DISPATCHED', 'IN_RELOAD_QUEUE', 'PENDING_RETURN', 'RETURN_PROCESSED',
            'IN_AUDIT', 'AUDIT_COMPLETE', 'CLOSED',
        ]
        pautas_done = PautaModel.objects.filter(
            operational_date=operational_date,
            assignments__role='YARD_DRIVER',
            assignments__personnel=profile,
            status__in=done_statuses,
        ).distinct()

        completed = 0
        total_boxes = 0
        total_minutes = 0.0
        for p in pautas_done:
            t0 = p.timestamps.filter(event_type='T1A_YARD_START').order_by('timestamp').first()
            t1 = p.timestamps.filter(event_type='T1B_YARD_END').order_by('timestamp').first()
            if t0 and t1:
                delta = (t1.timestamp - t0.timestamp).total_seconds() / 60
                if delta >= 0:
                    completed += 1
                    total_minutes += delta
                    total_boxes += p.total_boxes

        avg_minutes = (total_minutes / completed) if completed else None
        boxes_per_hour = (total_boxes / (total_minutes / 60)) if total_minutes > 0 else None

        # Pauta en movimiento ahora mismo por este chofer.
        in_progress = PautaModel.objects.filter(
            operational_date=operational_date,
            status='MOVING_TO_BAY',
            assignments__role='YARD_DRIVER',
            assignments__personnel=profile,
            assignments__is_active=True,
        ).distinct().first()

        in_progress_data = None
        if in_progress:
            t0 = in_progress.timestamps.filter(event_type='T1A_YARD_START').order_by('timestamp').first()
            in_progress_data = {
                'id': in_progress.id,
                'transport_number': in_progress.transport_number,
                'total_boxes': in_progress.total_boxes,
                'started_at': t0.timestamp.isoformat() if t0 else None,
            }

        return Response({
            'date': str(operational_date),
            'completed_count': completed,
            'total_boxes': total_boxes,
            'avg_movement_minutes': round(avg_minutes, 1) if avg_minutes is not None else None,
            'boxes_per_hour': round(boxes_per_hour, 1) if boxes_per_hour is not None else None,
            'in_progress': in_progress_data,
        })

    @action(detail=False, methods=['get'])
    def counter_stats(self, request):
        """
        Stats del turno del contador autenticado.
        Devuelve: completed_count, total_boxes, avg_counting_minutes,
        boxes_per_hour, in_progress.
        """
        try:
            profile = request.user.personnel_profile
        except Exception:
            return Response(
                {'error': 'El usuario no tiene perfil de personal asociado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        personnel_id = request.query_params.get('personnel_id')
        if personnel_id and (request.user.is_superuser or request.user.is_staff):
            from apps.personnel.models.personnel import PersonnelProfile
            try:
                profile = PersonnelProfile.objects.get(pk=personnel_id)
            except PersonnelProfile.DoesNotExist:
                return Response({'error': 'Personnel no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        operational_date = request.query_params.get('operational_date') or timezone.localdate()

        done_statuses = [
            'COUNTED', 'MOVING_TO_PARKING', 'PARKED',
            'PENDING_CHECKOUT', 'CHECKOUT_SECURITY', 'CHECKOUT_OPS',
            'DISPATCHED', 'IN_RELOAD_QUEUE', 'PENDING_RETURN', 'RETURN_PROCESSED',
            'IN_AUDIT', 'AUDIT_COMPLETE', 'CLOSED',
        ]
        pautas_done = PautaModel.objects.filter(
            operational_date=operational_date,
            assignments__role='COUNTER',
            assignments__personnel=profile,
            status__in=done_statuses,
        ).distinct()

        completed = 0
        total_boxes = 0
        total_minutes = 0.0

        for p in pautas_done:
            t0 = p.timestamps.filter(event_type='T5_COUNT_START').order_by('timestamp').first()
            t1 = p.timestamps.filter(event_type='T6_COUNT_END').order_by('timestamp').first()
            if t0 and t1:
                delta = (t1.timestamp - t0.timestamp).total_seconds() / 60
                if delta > 0:
                    completed += 1
                    total_minutes += delta
                    total_boxes += p.total_boxes

        avg_minutes = (total_minutes / completed) if completed else None
        boxes_per_hour = (total_boxes / (total_minutes / 60)) if total_minutes > 0 else None

        in_progress = PautaModel.objects.filter(
            operational_date=operational_date,
            status='COUNTING',
            assignments__role='COUNTER',
            assignments__personnel=profile,
            assignments__is_active=True,
        ).distinct().first()

        in_progress_data = None
        if in_progress:
            t0 = in_progress.timestamps.filter(event_type='T5_COUNT_START').order_by('timestamp').first()
            in_progress_data = {
                'id': in_progress.id,
                'transport_number': in_progress.transport_number,
                'total_boxes': in_progress.total_boxes,
                'started_at': t0.timestamp.isoformat() if t0 else None,
            }

        return Response({
            'date': str(operational_date),
            'completed_count': completed,
            'total_boxes': total_boxes,
            'avg_counting_minutes': round(avg_minutes, 1) if avg_minutes is not None else None,
            'boxes_per_hour': round(boxes_per_hour, 1) if boxes_per_hour is not None else None,
            'in_progress': in_progress_data,
        })

    @action(detail=True, methods=['post'])
    def start_picking(self, request, pk=None):
        return self._do_transition(request, pk, 'start_picking')

    @action(detail=True, methods=['post'])
    def complete_picking(self, request, pk=None):
        """
        Completa el picking. Si es una recarga ya re-ingresada con bahía asignada,
        auto-avanza a IN_BAY sin pasar por selección manual.
        """
        response = self._do_transition(request, pk, 'complete_picking')
        if response.status_code == 200:
            pauta = self.get_object()
            if (
                pauta.is_reload
                and pauta.reentered_at
                and getattr(pauta, 'bay_assignment', None)
            ):
                pauta.status = 'IN_BAY'
                pauta.save(update_fields=['status'])
                PautaTimestampModel.objects.create(
                    event_type='T2_BAY_ASSIGNED',
                    pauta=pauta,
                    recorded_by=request.user,
                )
                return Response(PautaDetailSerializer(pauta).data)
        return response

    @action(detail=True, methods=['post'])
    def assign_yard_driver(self, request, pk=None):
        """
        Cargas (trip 1): PICKING_DONE → MOVING_TO_BAY. Asigna chofer de patio e inicia timer.
        """
        pauta = self.get_object()
        if pauta.is_reload:
            return Response(
                {'error': 'Las recargas no usan flujo de chofer de patio. Use re-ingreso.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        personnel_id = request.data.get('personnel_id')
        if not personnel_id:
            return Response(
                {'error': 'Debe seleccionar un chofer de patio.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        response = self._do_transition(request, pk, 'assign_yard_driver')
        if response.status_code == 200:
            PautaAssignmentModel.objects.create(
                role='YARD_DRIVER',
                pauta=pauta,
                personnel_id=personnel_id,
                assigned_by=request.user,
            )
        return response

    @action(detail=True, methods=['post'])
    def position_at_bay(self, request, pk=None):
        """
        Cargas (trip 1): MOVING_TO_BAY → IN_BAY. Registra bahía y cierra timer del movimiento.
        """
        from apps.truck_cycle.models.operational import PautaBayAssignmentModel
        pauta = self.get_object()
        bay_id = request.data.get('bay_id')
        if not bay_id:
            return Response(
                {'error': 'Debe seleccionar la bahía donde se posicionó el camión.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        response = self._do_transition(request, pk, 'position_at_bay')
        if response.status_code == 200:
            PautaBayAssignmentModel.objects.update_or_create(
                pauta=pauta,
                defaults={
                    'bay_id': bay_id,
                    'assigned_by': request.user,
                },
            )
        return response

    @action(detail=True, methods=['post'])
    def assign_bay(self, request, pk=None):
        """Asignar andén directamente (legacy / uso manual)."""
        pauta = self.get_object()
        if pauta.is_reload and not pauta.reentered_at:
            return Response(
                {'error': 'La recarga aún no re-ingresó al CD. Debe registrarse el re-ingreso primero.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        response = self._do_transition(request, pk, 'assign_bay')
        if response.status_code == 200:
            from apps.truck_cycle.models.operational import PautaBayAssignmentModel
            bay_id = request.data.get('bay_id')
            if bay_id:
                PautaBayAssignmentModel.objects.update_or_create(
                    pauta=pauta,
                    defaults={
                        'bay_id': bay_id,
                        'assigned_by': request.user,
                    }
                )
            yard_driver_id = request.data.get('yard_driver_id')
            if yard_driver_id:
                PautaAssignmentModel.objects.create(
                    role='YARD_DRIVER',
                    pauta=pauta,
                    personnel_id=yard_driver_id,
                    assigned_by=request.user,
                )
        return response

    @action(detail=True, methods=['post'])
    def reload_reentry(self, request, pk=None):
        """
        Recargas: registra el re-ingreso del camión al CD.
        Body: { truck_id: int, bay_id: int }
        - Reasigna pauta.truck (permite que un camión traiga pauta de otro).
        - Crea/actualiza PautaBayAssignment.
        - Setea pauta.reentered_at = now.
        - Si pauta está en PICKING_DONE → auto-avanza a IN_BAY.
        - Si picking aún no completa → la bahía queda guardada y complete_picking la avanzará.
        """
        from apps.truck_cycle.models.catalogs import TruckModel
        from apps.truck_cycle.models.operational import PautaBayAssignmentModel, PautaTimestampModel

        pauta = self.get_object()
        if not pauta.is_reload:
            return Response(
                {'error': 'El re-ingreso solo aplica a recargas.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        truck_id = request.data.get('truck_id')
        bay_id = request.data.get('bay_id')
        if not truck_id or not bay_id:
            return Response(
                {'error': 'Debe seleccionar el camión que re-ingresa y la bahía.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        truck = TruckModel.objects.filter(
            pk=truck_id, distributor_center=pauta.distributor_center, is_active=True,
        ).first()
        if not truck:
            return Response(
                {'error': 'Camión no válido para este centro de distribución.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Reasigna el camión (puede traer pauta de otro camión).
        pauta.truck = truck
        pauta.reentered_at = timezone.now()
        pauta.save(update_fields=['truck', 'reentered_at'])

        PautaBayAssignmentModel.objects.update_or_create(
            pauta=pauta,
            defaults={'bay_id': bay_id, 'assigned_by': request.user},
        )
        PautaTimestampModel.objects.create(
            event_type='T10A_RELOAD_REENTRY',
            pauta=pauta,
            recorded_by=request.user,
            notes=request.data.get('notes', ''),
        )

        # Si ya está con picking listo, auto-avanzamos a IN_BAY.
        if pauta.status == 'PICKING_DONE':
            pauta.status = 'IN_BAY'
            pauta.save(update_fields=['status'])
            PautaTimestampModel.objects.create(
                event_type='T2_BAY_ASSIGNED',
                pauta=pauta,
                recorded_by=request.user,
            )

        return Response(PautaDetailSerializer(pauta).data)

    @action(detail=True, methods=['post'])
    def complete_loading(self, request, pk=None):
        """
        Fin de carga: IN_BAY → PENDING_COUNT. El camión SIGUE físicamente
        en la bahía mientras se cuenta, hace checkout, etc. La bahía se
        libera hasta el despacho (T9_DISPATCH).
        """
        return self._do_transition(request, pk, 'complete_loading')

    @action(detail=True, methods=['post'])
    def assign_counter(self, request, pk=None):
        """Asignar contador a la pauta"""
        response = self._do_transition(request, pk, 'assign_counter')
        if response.status_code == 200:
            pauta = self.get_object()
            personnel_id = request.data.get('personnel_id')
            if personnel_id:
                PautaAssignmentModel.objects.create(
                    role='COUNTER',
                    pauta=pauta,
                    personnel_id=personnel_id,
                    assigned_by=request.user,
                )
        return response

    @action(detail=True, methods=['post'])
    def complete_count(self, request, pk=None):
        return self._do_transition(request, pk, 'complete_count')

    @action(detail=True, methods=['post'])
    def checkout_security(self, request, pk=None):
        """Validación de seguridad en checkout"""
        response = self._do_transition(request, pk, 'checkout_security')
        if response.status_code == 200:
            from apps.truck_cycle.models.operational import CheckoutValidationModel
            pauta = self.get_object()
            checkout, _ = CheckoutValidationModel.objects.get_or_create(pauta=pauta)
            checkout.security_validated = True
            checkout.security_validated_at = timezone.now()
            validator_id = request.data.get('validator_id')
            if validator_id:
                checkout.security_validator_id = validator_id
            exit_pass = request.data.get('exit_pass_consumables')
            if exit_pass is not None:
                checkout.exit_pass_consumables = exit_pass
            checkout.save()
        return response

    @action(detail=True, methods=['post'])
    def checkout_ops(self, request, pk=None):
        """Validación de operaciones en checkout"""
        response = self._do_transition(request, pk, 'checkout_ops')
        if response.status_code == 200:
            from apps.truck_cycle.models.operational import CheckoutValidationModel
            pauta = self.get_object()
            checkout, _ = CheckoutValidationModel.objects.get_or_create(pauta=pauta)
            checkout.ops_validated = True
            checkout.ops_validated_at = timezone.now()
            validator_id = request.data.get('validator_id')
            if validator_id:
                checkout.ops_validator_id = validator_id
            checkout.save()
        return response

    @action(detail=True, methods=['post'], url_path='dispatch', url_name='dispatch')
    def dispatch_truck(self, request, pk=None):
        """
        Despacho (T9_DISPATCH). Libera la bahía — el camión sale físicamente.

        Requiere un chofer vendedor asignado. Si la pauta ya tiene asignación
        activa con rol DELIVERY_DRIVER (ej. auto-asignado al regresar), se usa.
        Si no, se requiere `driver_id` en el body para crear la asignación.
        """
        pauta = self.get_object()

        existing = pauta.assignments.filter(
            is_active=True, role='DELIVERY_DRIVER'
        ).order_by('-assigned_at').first()

        driver_id = request.data.get('driver_id')
        if not existing and not driver_id:
            return Response(
                {'error': 'Debe seleccionar un chofer vendedor para despachar.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if driver_id and not existing:
            from apps.personnel.models.personnel import PersonnelProfile
            try:
                driver = PersonnelProfile.objects.get(pk=driver_id, is_active=True)
            except PersonnelProfile.DoesNotExist:
                return Response(
                    {'error': 'Chofer no encontrado o inactivo.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if driver.position_type != 'DELIVERY_DRIVER':
                return Response(
                    {'error': 'El personal seleccionado no es un chofer vendedor.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        previous_status = pauta.status
        response = self._do_transition(request, pk, 'dispatch')
        if response.status_code == 200:
            pauta = self.get_object()
            if driver_id and not existing:
                PautaAssignmentModel.objects.create(
                    role='DELIVERY_DRIVER',
                    pauta=pauta,
                    personnel_id=driver_id,
                    assigned_by=request.user,
                )
            bay_assignment = getattr(pauta, 'bay_assignment', None)
            if bay_assignment and not bay_assignment.released_at:
                bay_assignment.released_at = timezone.now()
                bay_assignment.save(update_fields=['released_at'])
            if previous_status == 'PARKED':
                from apps.truck_cycle.models.operational import CheckoutValidationModel
                checkout, _ = CheckoutValidationModel.objects.get_or_create(pauta=pauta)
                if not checkout.security_validated:
                    checkout.dispatched_without_security = True
                    checkout.save(update_fields=['dispatched_without_security'])
        return response

    @action(detail=True, methods=['post'])
    def arrival(self, request, pk=None):
        return self._do_transition(request, pk, 'arrival')

    @action(detail=False, methods=['post'], url_path='public-arrival/(?P<truck_code>[^/.]+)',
            permission_classes=[permissions.AllowAny], authentication_classes=[])
    def public_arrival(self, request, truck_code=None):
        """Registro público de llegada — el chofer escanea QR en el bunker"""
        from apps.truck_cycle.models.catalogs import TruckModel
        truck = TruckModel.objects.filter(code=truck_code, is_active=True).first()
        if not truck:
            truck = TruckModel.objects.filter(plate=truck_code, is_active=True).first()
        if not truck:
            return Response({'error': 'Camión no encontrado'}, status=404)

        pauta = PautaModel.objects.filter(
            truck=truck, status='DISPATCHED'
        ).order_by('-created_at').first()
        if not pauta:
            return Response({'error': 'No hay pautas despachadas para este camión'}, status=404)

        transition = STATUS_TRANSITIONS['arrival']
        pauta.status = transition['to']
        pauta.save(update_fields=['status'])
        if transition['event']:
            from apps.truck_cycle.models.operational import PautaTimestampModel
            PautaTimestampModel.objects.create(
                pauta=pauta, event_type=transition['event']
            )
        return Response({
            'message': f'Llegada registrada para T-{pauta.transport_number}',
            'transport_number': pauta.transport_number,
            'truck_code': truck.code,
            'truck_plate': truck.plate,
        })

    @action(detail=False, methods=['get'], url_path='public-truck-status/(?P<truck_code>[^/.]+)',
            permission_classes=[permissions.AllowAny], authentication_classes=[])
    def public_truck_status(self, request, truck_code=None):
        """Status público del camión para la página de QR"""
        from apps.truck_cycle.models.catalogs import TruckModel
        truck = TruckModel.objects.filter(code=truck_code, is_active=True).first()
        if not truck:
            truck = TruckModel.objects.filter(plate=truck_code, is_active=True).first()
        if not truck:
            return Response({'error': 'Camión no encontrado'}, status=404)

        pauta = PautaModel.objects.filter(
            truck=truck
        ).exclude(status__in=['CLOSED', 'CANCELLED']).order_by('-created_at').first()

        return Response({
            'truck_code': truck.code,
            'truck_plate': truck.plate,
            'has_active_pauta': pauta is not None,
            'status': pauta.status if pauta else None,
            'status_display': pauta.get_status_display() if pauta else None,
            'transport_number': pauta.transport_number if pauta else None,
            'can_register_arrival': pauta.status == 'DISPATCHED' if pauta else False,
        })

    @action(detail=True, methods=['post'])
    def process_return(self, request, pk=None):
        return self._do_transition(request, pk, 'process_return')

    @action(detail=True, methods=['post'])
    def start_audit(self, request, pk=None):
        return self._do_transition(request, pk, 'start_audit')

    @action(detail=True, methods=['post'])
    def complete_audit(self, request, pk=None):
        return self._do_transition(request, pk, 'complete_audit')

    @action(detail=True, methods=['post'])
    def close(self, request, pk=None):
        return self._do_transition(request, pk, 'close')

    @action(detail=False, methods=['get'])
    def workstation(self, request):
        """Agrupar pautas por estado para vista de estación de trabajo"""
        qs = self.get_queryset()
        operational_date = request.query_params.get('operational_date')
        if operational_date:
            qs = qs.filter(operational_date=operational_date)

        grouped = {}
        for choice_value, choice_label in PautaModel.STATUS_CHOICES:
            pautas = qs.filter(status=choice_value)
            grouped[choice_value] = {
                'label': choice_label,
                'count': pautas.count(),
                'pautas': PautaListSerializer(pautas, many=True).data,
            }

        return Response(grouped)

    @action(detail=False, methods=['get'])
    def reload_queue(self, request):
        """Pautas en cola de recarga ordenadas por fecha de creación"""
        qs = self.get_queryset().filter(
            status='IN_RELOAD_QUEUE'
        ).order_by('created_at')
        operational_date = request.query_params.get('operational_date')
        if operational_date:
            qs = qs.filter(operational_date=operational_date)
        serializer = PautaListSerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def kpi_summary(self, request):
        """Resumen de KPIs agregados con productividad y TAR"""
        qs = self.get_queryset()
        operational_date = request.query_params.get('operational_date')
        if operational_date:
            qs = qs.filter(operational_date=operational_date)

        total_pautas = qs.count()
        dispatched = qs.filter(status='DISPATCHED').count()
        closed = qs.filter(status='CLOSED').count()
        completed_pautas = dispatched + closed
        total_boxes = qs.aggregate(total=Sum('total_boxes'))['total'] or 0

        from apps.truck_cycle.models.operational import InconsistencyModel
        inconsistency_count = InconsistencyModel.objects.filter(
            pauta__in=qs
        ).count()

        # Calcular productividad de picking (cajas/hora)
        avg_boxes_per_hour = None
        picking_pautas = qs.exclude(status='PENDING_PICKING')
        picking_rates = []
        for pauta in picking_pautas:
            t0 = pauta.timestamps.filter(event_type='T0_PICKING_START').first()
            t1 = pauta.timestamps.filter(event_type='T1_PICKING_END').first()
            if t0 and t1:
                delta_hours = (t1.timestamp - t0.timestamp).total_seconds() / 3600
                if delta_hours >= 0.05:  # al menos 3 minutos
                    picking_rates.append(pauta.total_boxes / delta_hours)
        if picking_rates:
            avg_boxes_per_hour = round(sum(picking_rates) / len(picking_rates), 1)

        # Calcular TAR promedio (minutos)
        avg_tar_minutes = None
        tar_values = []
        for pauta in qs.filter(is_reload=True):
            t_arrival = pauta.timestamps.filter(event_type='T10_ARRIVAL').first()
            t_counted = pauta.timestamps.filter(event_type='T6_COUNT_END').first()
            if t_arrival and t_counted:
                tar_min = (t_counted.timestamp - t_arrival.timestamp).total_seconds() / 60
                if tar_min > 0:
                    tar_values.append(tar_min)
        if tar_values:
            avg_tar_minutes = round(sum(tar_values) / len(tar_values), 1)

        # Precisión de conteo
        avg_count_accuracy = None
        if total_pautas > 0:
            pautas_with_issues = InconsistencyModel.objects.filter(
                pauta__in=qs, phase='VERIFICATION'
            ).values('pauta').distinct().count()
            avg_count_accuracy = round(1 - (pautas_with_issues / total_pautas), 3) if total_pautas > 0 else None

        # Error rate de picking
        avg_picking_error_rate = None
        if total_pautas > 0:
            picking_errors = InconsistencyModel.objects.filter(
                pauta__in=qs, phase='VERIFICATION'
            ).count()
            avg_picking_error_rate = round(picking_errors / total_pautas, 3)

        # Status breakdown
        pautas_by_status = list(
            qs.values('status').annotate(count=Count('id')).order_by('status')
        )

        return Response({
            'operational_date': operational_date or str(timezone.now().date()),
            'total_pautas': total_pautas,
            'completed_pautas': completed_pautas,
            'dispatched': dispatched,
            'closed': closed,
            'total_boxes': total_boxes,
            'avg_boxes_per_hour': avg_boxes_per_hour,
            'avg_count_accuracy': avg_count_accuracy,
            'avg_picking_error_rate': avg_picking_error_rate,
            'avg_tar_minutes': avg_tar_minutes,
            'inconsistency_count': inconsistency_count,
            'pautas_by_status': pautas_by_status,
        })

    @action(detail=True, methods=['get'])
    def download_pdf(self, request, pk=None):
        """Descargar PDF de una pauta individual (estilo moderno con pdfkit)"""
        pauta = self.get_object()
        try:
            from apps.truck_cycle.utils.pdf_generator import generate_pauta_pdf
            pdf_buffer = generate_pauta_pdf(pauta)
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="pauta_{pauta.transport_number}_V{pauta.trip_number}.pdf"'
            return response
        except Exception as e:
            import logging
            import traceback
            logger = logging.getLogger(__name__)
            logger.error(f"Error generando PDF para pauta {pauta.id}: {e}")
            logger.error(f"Traceback:\n{traceback.format_exc()}")
            return Response(
                {'error': 'Error al generar el documento PDF'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='export_excel')
    def export_excel(self, request):
        """Exportar pautas a Excel"""
        qs = self.filter_queryset(self.get_queryset()).select_related('truck')
        today = date.today().strftime('%Y-%m-%d')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pautas"

        # Header style
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0'),
        )

        columns = [
            ('Transporte', 15), ('Viaje', 10), ('Ruta', 12), ('Camión', 20),
            ('Cajas', 10), ('SKUs', 10), ('Pallets', 10), ('Estado', 25),
            ('Fecha Operativa', 15), ('Recarga', 10),
        ]

        for col_idx, (col_name, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = 'A2'

        alt_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        cell_alignment = Alignment(vertical='center')

        for row_idx, pauta in enumerate(qs, start=2):
            values = [
                pauta.transport_number,
                pauta.trip_number,
                pauta.route_code,
                f"{pauta.truck.code} - {pauta.truck.plate}" if pauta.truck else '-',
                pauta.total_boxes,
                pauta.total_skus,
                pauta.total_pallets,
                pauta.get_status_display(),
                str(pauta.operational_date) if pauta.operational_date else '-',
                'Sí' if pauta.is_reload else 'No',
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="pautas-{today}.xlsx"'
        return response

    @action(detail=False, methods=['get'], url_path='export_pdf')
    def export_pdf(self, request):
        """Exportar listado de pautas a PDF"""
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
        except ImportError:
            return Response(
                {'error': 'reportlab no está instalado.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        qs = self.filter_queryset(self.get_queryset()).select_related('truck')
        today = date.today().strftime('%Y-%m-%d')

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
        styles = getSampleStyleSheet()
        story = []

        title_style = ParagraphStyle('Title2', parent=styles['Title'], fontSize=16, textColor=colors.HexColor('#1976d2'))
        story.append(Paragraph("Listado de Pautas", title_style))
        meta_style = ParagraphStyle('Meta', parent=styles['Normal'], fontSize=9, textColor=colors.grey)
        story.append(Paragraph(f"Generado: {today} | Total: {qs.count()} registros", meta_style))
        story.append(Spacer(1, 12))

        data = [['Transporte', 'Viaje', 'Ruta', 'Camión', 'Cajas', 'SKUs', 'Estado', 'Fecha']]
        cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=8)

        for pauta in qs[:500]:
            truck_str = f"{pauta.truck.code}-{pauta.truck.plate}" if pauta.truck else '-'
            data.append([
                pauta.transport_number,
                pauta.trip_number,
                pauta.route_code or '-',
                Paragraph(truck_str, cell_style),
                str(pauta.total_boxes),
                str(pauta.total_skus),
                pauta.get_status_display(),
                str(pauta.operational_date) if pauta.operational_date else '-',
            ])

        col_widths = [3*cm, 2*cm, 2.5*cm, 4*cm, 2*cm, 2*cm, 5*cm, 3*cm]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (4, 0), (5, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e0e0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(t)

        doc.build(story)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="pautas-{today}.pdf"'
        return response
