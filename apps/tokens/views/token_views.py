"""
ViewSet principal para TokenRequest
"""
import io
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from django.shortcuts import get_object_or_404
from django.db.models import Q
from django.http import HttpResponse

from ..models import TokenRequest
from ..serializers import (
    TokenRequestListSerializer,
    TokenRequestDetailSerializer,
    TokenRequestCreateSerializer,
    TokenApprovalSerializer,
    TokenRejectSerializer,
    TokenValidateSerializer,
    PublicTokenSerializer,
    BulkOvertimeCreateSerializer,
)
from ..permissions import (
    CanRequestTokens,
    CanApproveTokenL1,
    CanApproveTokenL2,
    CanApproveTokenL3,
    CanValidateToken,
    IsTokenOwnerOrApprover,
)
from ..filters import TokenRequestFilter
from ..utils import generate_token_qr, TokenNotificationHelper, generate_token_pdf, generate_token_receipt, generate_receipt_html
from ..services.approval_service import ApprovalLevelService
from apps.personnel.models import PersonnelProfile


class TokenRequestViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de tokens.

    Endpoints:
    - GET /api/tokens/ - Listar tokens (con filtros)
    - POST /api/tokens/ - Crear nuevo token
    - GET /api/tokens/{id}/ - Detalle de token
    - PUT/PATCH /api/tokens/{id}/ - Actualizar token (solo borrador)
    - DELETE /api/tokens/{id}/ - Eliminar token (solo borrador)

    Actions:
    - POST /api/tokens/{id}/approve_l1/ - Aprobar nivel 1
    - POST /api/tokens/{id}/approve_l2/ - Aprobar nivel 2
    - POST /api/tokens/{id}/approve_l3/ - Aprobar nivel 3
    - POST /api/tokens/{id}/reject/ - Rechazar token
    - POST /api/tokens/{id}/cancel/ - Cancelar token
    - POST /api/tokens/validate/ - Validar token por QR (Seguridad)
    - GET /api/tokens/pending_my_approval/ - Tokens pendientes de mi aprobación
    - GET /api/tokens/my_tokens/ - Mis tokens (como beneficiario o solicitante)
    """
    queryset = TokenRequest.objects.all()
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = TokenRequestFilter
    search_fields = ['display_number', 'personnel__first_name', 'personnel__last_name', 'personnel__employee_code']
    ordering_fields = ['created_at', 'valid_from', 'valid_until', 'status']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return TokenRequestListSerializer
        if self.action == 'create':
            return TokenRequestCreateSerializer
        if self.action in ['approve_l1', 'approve_l2', 'approve_l3']:
            return TokenApprovalSerializer
        if self.action == 'reject':
            return TokenRejectSerializer
        return TokenRequestDetailSerializer

    def get_permissions(self):
        if self.action == 'create':
            return [IsAuthenticated(), CanRequestTokens()]
        if self.action == 'approve_l1':
            return [IsAuthenticated(), CanApproveTokenL1()]
        if self.action == 'approve_l2':
            return [IsAuthenticated(), CanApproveTokenL2()]
        if self.action == 'approve_l3':
            return [IsAuthenticated(), CanApproveTokenL3()]
        if self.action == 'validate':
            return [IsAuthenticated(), CanValidateToken()]
        if self.action in ['retrieve', 'update', 'partial_update', 'destroy']:
            return [IsAuthenticated(), IsTokenOwnerOrApprover()]
        return [IsAuthenticated()]

    def get_queryset(self):
        """Filtrar tokens según permisos del usuario"""
        user = self.request.user
        qs = TokenRequest.objects.select_related(
            'personnel', 'personnel__area',
            'requested_by',
            'distributor_center',
            'approved_level_1_by',
            'approved_level_2_by',
            'approved_level_3_by',
            'rejected_by',
            'validated_by',
        )

        if user.is_superuser or user.is_staff:
            return qs

        # Filtrar por centro de distribución del usuario (aplica a todos, incluso sin personnel_profile)
        user_centers = list(user.distributions_centers.values_list('id', flat=True))
        if user.centro_distribucion:
            user_centers.append(user.centro_distribucion.id)

        if not user_centers:
            return qs.none()

        return qs.filter(distributor_center_id__in=user_centers)

    def _process_created_token(self, token, request_user):
        """Auto-aprobación + QR + notificación. Reusado por create() y bulk."""
        import logging
        from django.utils import timezone

        logger = logging.getLogger(__name__)

        # Verificar auto-aprobación
        try:
            requester_profile = request_user.personnel_profile
            beneficiary_profile = token.personnel

            if beneficiary_profile and ApprovalLevelService.can_auto_approve(
                requester_profile, beneficiary_profile
            ):
                if token.requires_level_1 and not token.approved_level_1_at:
                    token.approved_level_1_by = requester_profile
                    token.approved_level_1_at = timezone.now()
                    token.approved_level_1_notes = "Auto-aprobado por jerarquía superior"

                if token.requires_level_2 and not token.approved_level_2_at:
                    token.approved_level_2_by = requester_profile
                    token.approved_level_2_at = timezone.now()
                    token.approved_level_2_notes = "Auto-aprobado por jerarquía superior"

                if token.requires_level_3 and not token.approved_level_3_at:
                    token.approved_level_3_by = requester_profile
                    token.approved_level_3_at = timezone.now()
                    token.approved_level_3_notes = "Auto-aprobado por jerarquía superior"

                token.status = TokenRequest.Status.APPROVED
                token.save()

                logger.info(
                    f"Token {token.display_number} auto-aprobado por {requester_profile.full_name} "
                    f"(jerarquía: {requester_profile.hierarchy_level})"
                )

        except PersonnelProfile.DoesNotExist:
            pass

        # Generar código QR
        try:
            qr_url = generate_token_qr(token)
            token.qr_code_url = qr_url
            token.save(update_fields=['qr_code_url'])
        except Exception as e:
            logger.error(f"Error generando QR para token {token.id}: {e}")

        # Enviar notificación al siguiente aprobador (solo si no fue auto-aprobado)
        if token.status != TokenRequest.Status.APPROVED:
            TokenNotificationHelper.notify_pending_approval(token)

    def create(self, request, *args, **kwargs):
        """Crear token, generar QR y devolver detalle completo"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        token = serializer.save(requested_by=request.user)
        self._process_created_token(token, request.user)

        detail_serializer = TokenRequestDetailSerializer(token)
        return Response(detail_serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='bulk_create_overtime')
    def bulk_create_overtime(self, request):
        """
        Crear múltiples tokens de horas extra para diferentes personas con los mismos detalles.
        POST /api/tokens/bulk_create_overtime/
        """
        from datetime import datetime
        from django.db import transaction
        from ..models import OvertimeDetail

        serializer = BulkOvertimeCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data
        personnel_ids = data['personnel_ids']
        distributor_center = data['distributor_center']
        overtime_detail_raw = data['overtime_detail']

        # Parse overtime detail once (same for all)
        overtime_data = dict(overtime_detail_raw)
        overtime_type_model_id = overtime_data.pop('overtime_type_model', None)
        reason_model_id = overtime_data.pop('reason_model', None)
        if overtime_type_model_id:
            overtime_data['overtime_type_model_id'] = overtime_type_model_id
        if reason_model_id:
            overtime_data['reason_model_id'] = reason_model_id
        if isinstance(overtime_data.get('start_time'), str):
            overtime_data['start_time'] = datetime.strptime(overtime_data['start_time'], '%H:%M').time()
        if isinstance(overtime_data.get('end_time'), str):
            overtime_data['end_time'] = datetime.strptime(overtime_data['end_time'], '%H:%M').time()
        if isinstance(overtime_data.get('overtime_date'), str):
            overtime_data['overtime_date'] = datetime.strptime(overtime_data['overtime_date'], '%Y-%m-%d').date()

        personnel_list = PersonnelProfile.objects.filter(
            id__in=personnel_ids, is_active=True
        ).select_related('area')

        personnel_map = {p.id: p for p in personnel_list}

        created = []
        failed = []

        for pid in personnel_ids:
            person = personnel_map.get(pid)
            if not person:
                failed.append({
                    'personnel_id': pid,
                    'employee_code': '',
                    'name': '',
                    'error': 'No encontrado o inactivo',
                })
                continue

            try:
                with transaction.atomic():
                    # Determine approval levels
                    requires_l1, requires_l2, requires_l3, initial_status = (
                        ApprovalLevelService.determine_approval_levels(
                            TokenRequest.TokenType.OVERTIME,
                            person.hierarchy_level,
                            False,
                        )
                    )

                    token = TokenRequest.objects.create(
                        token_type=TokenRequest.TokenType.OVERTIME,
                        personnel=person,
                        distributor_center=distributor_center,
                        valid_from=data['valid_from'],
                        valid_until=data['valid_until'],
                        requester_notes=data.get('requester_notes', ''),
                        requested_by=request.user,
                        requires_level_1=requires_l1,
                        requires_level_2=requires_l2,
                        requires_level_3=requires_l3,
                        status=initial_status,
                    )

                    OvertimeDetail.objects.create(token=token, **dict(overtime_data))
                    self._process_created_token(token, request.user)

                    created.append({
                        'id': token.id,
                        'display_number': token.display_number,
                        'personnel_name': person.full_name,
                        'personnel_code': person.employee_code,
                        'status': token.status,
                        'status_display': token.get_status_display(),
                    })
            except Exception as e:
                failed.append({
                    'personnel_id': pid,
                    'employee_code': person.employee_code,
                    'name': person.full_name,
                    'error': str(e),
                })

        return Response({
            'created': created,
            'failed': failed,
            'total_requested': len(personnel_ids),
            'total_created': len(created),
            'total_failed': len(failed),
        }, status=status.HTTP_201_CREATED if created else status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'], url_path='resolve_employee_codes')
    def resolve_employee_codes(self, request):
        """
        Resolver códigos de empleado a perfiles de personal.
        POST /api/tokens/resolve_employee_codes/
        """
        codes = request.data.get('employee_codes', [])
        if not codes or not isinstance(codes, list):
            return Response(
                {'error': 'employee_codes debe ser una lista no vacía.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Normalize codes to strings
        codes = [str(c).strip() for c in codes if str(c).strip()]

        profiles = PersonnelProfile.objects.filter(
            employee_code__in=codes, is_active=True
        ).select_related('area')

        found_codes = set()
        resolved = []
        for p in profiles:
            found_codes.add(p.employee_code)
            resolved.append({
                'id': p.id,
                'employee_code': p.employee_code,
                'full_name': p.full_name,
                'position': p.position or '',
                'area_name': p.area.name if p.area else '',
                'hierarchy_level': p.hierarchy_level,
            })

        not_found = [c for c in codes if c not in found_codes]

        return Response({
            'resolved': resolved,
            'not_found': not_found,
        })

    @action(detail=True, methods=['post'])
    def approve_l1(self, request, pk=None):
        """
        Aprobar token en nivel 1.
        Acepta multipart/form-data con firma y foto opcionales.
        """
        token = self.get_object()
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            personnel = request.user.personnel_profile
            token.approve_level_1(
                personnel,
                notes=serializer.validated_data.get('notes', ''),
                signature=serializer.validated_data.get('signature'),
                photo=serializer.validated_data.get('photo')
            )

            # Notificar aprobación
            TokenNotificationHelper.notify_token_approved(token, 1)

            # Si hay más niveles, notificar al siguiente aprobador
            if token.status in [TokenRequest.Status.PENDING_L2, TokenRequest.Status.PENDING_L3]:
                TokenNotificationHelper.notify_pending_approval(token)

            return Response(TokenRequestDetailSerializer(token).data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def approve_l2(self, request, pk=None):
        """
        Aprobar token en nivel 2.
        Acepta multipart/form-data con firma y foto opcionales.
        """
        token = self.get_object()
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            personnel = request.user.personnel_profile
            token.approve_level_2(
                personnel,
                notes=serializer.validated_data.get('notes', ''),
                signature=serializer.validated_data.get('signature'),
                photo=serializer.validated_data.get('photo')
            )

            # Notificar aprobación
            TokenNotificationHelper.notify_token_approved(token, 2)

            # Si hay más niveles, notificar al siguiente aprobador
            if token.status == TokenRequest.Status.PENDING_L3:
                TokenNotificationHelper.notify_pending_approval(token)

            return Response(TokenRequestDetailSerializer(token).data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def approve_l3(self, request, pk=None):
        """
        Aprobar token en nivel 3.
        Acepta multipart/form-data con firma y foto opcionales.
        """
        token = self.get_object()
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            personnel = request.user.personnel_profile
            token.approve_level_3(
                personnel,
                notes=serializer.validated_data.get('notes', ''),
                signature=serializer.validated_data.get('signature'),
                photo=serializer.validated_data.get('photo')
            )

            # Notificar aprobación final
            TokenNotificationHelper.notify_token_approved(token, 3)

            return Response(TokenRequestDetailSerializer(token).data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Rechazar token"""
        token = self.get_object()
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            personnel = request.user.personnel_profile
            token.reject(personnel, serializer.validated_data['reason'])

            # Notificar rechazo
            TokenNotificationHelper.notify_token_rejected(token)

            return Response(TokenRequestDetailSerializer(token).data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Cancelar token"""
        token = self.get_object()

        try:
            token.cancel()
            return Response(TokenRequestDetailSerializer(token).data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def validate(self, request):
        """
        Validar token para marcarlo como USED.
        - EXIT_PASS, PERMIT_HOUR: validados por Seguridad
        - RATE_CHANGE, SUBSTITUTION: validados por Planilla
        Los demás tipos (PERMIT_DAY, OVERTIME, SHIFT_CHANGE) se quedan en APPROVED.
        Acepta multipart/form-data con firma y foto opcionales.
        Espera: { "token_code": "uuid-del-token", "notes"?: "", "signature"?: File, "photo"?: File }
        """
        serializer = TokenValidateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        token_code = serializer.validated_data.get('token_code')
        token = get_object_or_404(TokenRequest, token_code=token_code)

        # Verificar si el tipo de token requiere validación
        if not token.requires_validation:
            return Response(
                {'error': f'Este tipo de token ({token.get_token_type_display()}) no requiere validación. Se queda en estado Aprobado.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar permisos según tipo de validación
        validation_type = token.validation_type
        user = request.user

        if validation_type == 'security':
            # Validar que el usuario tenga permiso de seguridad (portería)
            if not user.has_perm('tokens.can_validate_token'):
                return Response(
                    {'error': 'No tiene permisos para validar tokens de seguridad.'},
                    status=status.HTTP_403_FORBIDDEN
                )
        elif validation_type == 'payroll':
            # Validar que el usuario tenga permiso de planilla
            if not user.has_perm('tokens.can_validate_payroll'):
                return Response(
                    {'error': 'No tiene permisos para marcar tokens como utilizados (Planilla).'},
                    status=status.HTTP_403_FORBIDDEN
                )

        if not token.can_be_used:
            if token.status != TokenRequest.Status.APPROVED:
                return Response(
                    {'error': f'El token no está aprobado. Estado actual: {token.get_status_display()}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if not token.is_valid:
                return Response(
                    {'error': 'El token está fuera del período de vigencia'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        try:
            personnel = request.user.personnel_profile
            token.mark_as_used(
                validated_by=personnel,
                signature=serializer.validated_data.get('signature'),
                photo=serializer.validated_data.get('photo'),
                notes=serializer.validated_data.get('notes', '')
            )

            # Notificar que fue utilizado (excluyendo al usuario que validó)
            TokenNotificationHelper.notify_token_used(token, used_by_user=request.user)

            return Response(TokenRequestDetailSerializer(token).data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def pending_my_approval(self, request):
        """Obtener tokens pendientes de aprobación para el usuario actual"""
        try:
            personnel = request.user.personnel_profile
        except:
            return Response([])

        qs = self.get_queryset()

        # Filtrar según nivel de aprobación del usuario
        if personnel.can_approve_tokens_level_3():
            qs = qs.filter(
                Q(status=TokenRequest.Status.PENDING_L1) |
                Q(status=TokenRequest.Status.PENDING_L2) |
                Q(status=TokenRequest.Status.PENDING_L3)
            )
        elif personnel.can_approve_tokens_level_2():
            qs = qs.filter(
                Q(status=TokenRequest.Status.PENDING_L1) |
                Q(status=TokenRequest.Status.PENDING_L2)
            )
        elif personnel.can_approve_tokens_level_1():
            qs = qs.filter(status=TokenRequest.Status.PENDING_L1)
        else:
            return Response([])

        # Excluir tokens donde el usuario es el beneficiario (no puede aprobar su propio token)
        qs = qs.exclude(personnel=personnel)

        serializer = TokenRequestListSerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def my_tokens(self, request):
        """Obtener tokens donde el usuario es beneficiario o solicitante"""
        user = request.user
        qs = self.get_queryset()

        try:
            personnel = user.personnel_profile
            qs = qs.filter(
                Q(personnel=personnel) | Q(requested_by=user)
            )
        except:
            qs = qs.filter(requested_by=user)

        serializer = TokenRequestListSerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='by_code/(?P<code>[^/.]+)')
    def by_code(self, request, code=None):
        """
        Buscar token por display_number (TK-2026-000001) o token_code (UUID).
        Usado en la página de validación para buscar tokens.
        """
        if not code:
            return Response(
                {'error': 'Se requiere el código del token'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Intentar buscar por display_number primero (más común)
        token = TokenRequest.objects.filter(display_number__iexact=code).first()

        # Si no se encuentra, intentar por token_code (UUID)
        if not token:
            token = TokenRequest.objects.filter(token_code=code).first()

        if not token:
            return Response(
                {'error': 'Token no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Devolver datos para validación (similar a public pero con más info)
        serializer = PublicTokenSerializer(token)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def download_pdf(self, request, pk=None):
        """
        Descargar el token como documento PDF.
        GET /api/tokens/{id}/download_pdf/
        """
        token = self.get_object()

        try:
            pdf_buffer = generate_token_pdf(token)
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="token_{token.display_number}.pdf"'
            return response
        except Exception as e:
            import logging
            import traceback
            logger = logging.getLogger(__name__)
            logger.error(f"Error generando PDF para token {token.id}: {e}")
            logger.error(f"Traceback:\n{traceback.format_exc()}")
            return Response(
                {'error': 'Error al generar el documento PDF'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'])
    def download_receipt(self, request, pk=None):
        """
        Descargar recibo tipo ticket (80mm) para impresora térmica.
        Solo disponible cuando el token está APPROVED o USED.
        GET /api/tokens/{id}/download_receipt/?copy=true para copia
        GET /api/tokens/{id}/download_receipt/ para original
        """
        token = self.get_object()

        # Verificar estado del token
        allowed_states = [TokenRequest.Status.APPROVED, TokenRequest.Status.USED]
        if token.status not in allowed_states:
            return Response(
                {'error': f'El recibo solo está disponible para tokens aprobados o utilizados. Estado actual: {token.get_status_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Determinar si es copia u original
        is_copy = request.query_params.get('copy', '').lower() in ['true', '1', 'yes']

        try:
            receipt_buffer = generate_token_receipt(token, is_copy=is_copy)
            response = HttpResponse(receipt_buffer.getvalue(), content_type='application/pdf')
            suffix = '_copia' if is_copy else ''
            response['Content-Disposition'] = f'attachment; filename="recibo_{token.display_number}{suffix}.pdf"'
            return response
        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error generando recibo para token {token.id}: {e}")
            return Response(
                {'error': 'Error al generar el recibo'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'])
    def print_receipt(self, request, pk=None):
        """
        Retorna el recibo como HTML para imprimir directamente desde el navegador.
        El HTML incluye @page { size: 80mm auto } para impresoras térmicas.
        GET /api/tokens/{id}/print_receipt/?copy=true
        """
        token = self.get_object()

        allowed_states = [TokenRequest.Status.APPROVED, TokenRequest.Status.USED]
        if token.status not in allowed_states:
            return Response(
                {'error': f'El recibo solo está disponible para tokens aprobados o utilizados.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        is_copy = request.query_params.get('copy', '').lower() in ['true', '1', 'yes']

        try:
            html_content = generate_receipt_html(token, is_copy=is_copy)
            return HttpResponse(html_content, content_type='text/html; charset=utf-8')
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error generando recibo HTML para token {token.id}: {e}")
            return Response({'error': 'Error al generar el recibo'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def pending_validation(self, request):
        """
        Obtener tokens aprobados pendientes de validación (para Seguridad).
        Solo tokens APPROVED que están dentro del período de vigencia.
        GET /api/tokens/pending_validation/
        """
        from django.utils import timezone

        qs = self.get_queryset().filter(
            status=TokenRequest.Status.APPROVED,
            valid_from__lte=timezone.now(),
            valid_until__gte=timezone.now(),
        ).order_by('valid_until')  # Ordenar por los que expiran primero

        serializer = PublicTokenSerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def complete_uniform_delivery(self, request, pk=None):
        """
        Completar la entrega de uniforme con foto y firma.
        POST /api/tokens/{id}/complete_uniform_delivery/

        Expected data (multipart/form-data):
        - signature: image file (required)
        - photo_1: image file (optional)
        - photo_2: image file (optional)
        - notes: string (optional)
        """
        token = self.get_object()

        # Validar que sea un token de entrega de uniforme
        if token.token_type != TokenRequest.TokenType.UNIFORM_DELIVERY:
            return Response(
                {'error': 'Este token no es de entrega de uniforme'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar que el token esté aprobado
        if token.status != TokenRequest.Status.APPROVED:
            return Response(
                {'error': f'El token debe estar aprobado. Estado actual: {token.get_status_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar que tenga detalle de uniforme
        if not hasattr(token, 'uniform_delivery_detail'):
            return Response(
                {'error': 'Token sin detalle de uniforme'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar que no esté ya entregado
        if token.uniform_delivery_detail.is_delivered:
            return Response(
                {'error': 'Este uniforme ya fue entregado'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obtener datos
        signature = request.FILES.get('signature')
        photo_1 = request.FILES.get('photo_1')
        photo_2 = request.FILES.get('photo_2')
        notes = request.data.get('notes', '')

        # Validar firma
        if not signature:
            return Response(
                {'error': 'Se requiere la firma del beneficiario'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            personnel = request.user.personnel_profile
        except:
            return Response(
                {'error': 'Usuario sin perfil de personal'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Completar la entrega
        try:
            token.uniform_delivery_detail.mark_as_delivered(
                delivered_by=personnel,
                signature=signature,
                photo1=photo_1,
                photo2=photo_2,
                notes=notes
            )

            # Marcar el token como usado
            token.mark_as_used(personnel)

            # Notificar (excluyendo al usuario que completó la entrega)
            TokenNotificationHelper.notify_token_used(token, used_by_user=request.user)

            return Response(TokenRequestDetailSerializer(token).data)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error completando entrega de uniforme {token.id}: {e}")
            return Response(
                {'error': 'Error al completar la entrega'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    # ──────────────────────────────────────────────────────────────────────────
    # Exportación
    # ──────────────────────────────────────────────────────────────────────────

    TOKEN_TYPE_LABELS = {
        'PERMIT_HOUR': 'Permiso por Hora',
        'PERMIT_DAY': 'Permiso por Día',
        'EXIT_PASS': 'Pase de Salida',
        'UNIFORM_DELIVERY': 'Entrega de Uniforme',
        'SUBSTITUTION': 'Sustitución',
        'RATE_CHANGE': 'Cambio de Tasa',
        'OVERTIME': 'Horas Extra',
        'SHIFT_CHANGE': 'Cambio de Turno',
    }

    STATUS_LABELS_EXPORT = {
        'DRAFT': 'Borrador',
        'PENDING_L1': 'Pendiente',
        'PENDING_L2': 'Pendiente',
        'PENDING_L3': 'Pendiente',
        'APPROVED': 'Abierto',
        'USED': 'Finalizado',
        'EXPIRED': 'Vencido',
        'CANCELLED': 'Cerrado',
        'REJECTED': 'Cerrado',
    }

    @action(detail=False, methods=['get'], url_path='export_excel')
    def export_excel(self, request):
        """Exporta el listado de tokens filtrado a Excel."""
        queryset = self.filter_queryset(self.get_queryset())

        HDR_FILL = '1976D2'
        EX_FILL = 'F8F9FA'

        def _hdr(ws, row, col, value, width=None):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(bold=True, color='FFFFFF', size=10)
            cell.fill = PatternFill(start_color=HDR_FILL, end_color=HDR_FILL, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'),
            )
            if width:
                ws.column_dimensions[get_column_letter(col)].width = width
            return cell

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Tokens'
        ws.row_dimensions[1].height = 32

        COLS = [
            ('N° Token', 16),
            ('Tipo', 22),
            ('Beneficiario', 32),
            ('Código Emp.', 14),
            ('Centro', 22),
            ('Solicitante', 28),
            ('F. Creación', 14),
            ('F. Válido', 14),
            ('Estado', 14),
            ('Observaciones', 40),
        ]
        for i, (label, w) in enumerate(COLS, start=1):
            _hdr(ws, 1, i, label, w)

        ws.freeze_panes = 'A2'

        for row_idx, token in enumerate(queryset, start=2):
            row_bg = 'FFFFFF' if row_idx % 2 == 0 else EX_FILL

            beneficiary = ''
            if token.personnel:
                beneficiary = token.personnel.full_name
            elif hasattr(token, 'exit_pass_detail') and token.exit_pass_detail and token.exit_pass_detail.external_person:
                beneficiary = token.exit_pass_detail.external_person.name

            requester = ''
            if token.requested_by:
                name = f"{token.requested_by.first_name} {token.requested_by.last_name}".strip()
                requester = name or token.requested_by.email

            values = [
                token.display_number or '',
                self.TOKEN_TYPE_LABELS.get(token.token_type, token.token_type),
                beneficiary,
                token.personnel.employee_code if token.personnel else '',
                token.distributor_center.name if token.distributor_center else '',
                requester,
                token.created_at.strftime('%d/%m/%Y') if token.created_at else '',
                token.valid_until.strftime('%d/%m/%Y') if token.valid_until else '',
                self.STATUS_LABELS_EXPORT.get(token.status, token.status),
                token.requester_notes or '',
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type='solid')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'),
                )
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        today = date.today().strftime('%Y-%m-%d')
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="tokens-{today}.xlsx"'
        return response

    @action(detail=False, methods=['get'], url_path='export_pdf')
    def export_pdf(self, request):
        """Exporta el listado de tokens filtrado a PDF."""
        import os
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
        except ImportError:
            return Response(
                {'detail': 'Módulo reportlab no instalado. Ejecute: pip install reportlab'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        from django.conf import settings

        queryset = self.filter_queryset(self.get_queryset())

        STATUS_COLORS = {
            'DRAFT':      colors.HexColor('#64748b'),
            'PENDING_L1': colors.HexColor('#d97706'),
            'PENDING_L2': colors.HexColor('#d97706'),
            'PENDING_L3': colors.HexColor('#d97706'),
            'APPROVED':   colors.HexColor('#059669'),
            'USED':       colors.HexColor('#2563eb'),
            'EXPIRED':    colors.HexColor('#64748b'),
            'CANCELLED':  colors.HexColor('#dc2626'),
            'REJECTED':   colors.HexColor('#dc2626'),
        }

        BRAND_BLUE = colors.HexColor('#1976d2')
        DARK = colors.HexColor('#1a1a2e')
        LIGHT_GRAY = colors.HexColor('#f5f5f5')

        base_styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'title', parent=base_styles['Normal'],
            fontSize=13, fontName='Helvetica-Bold',
            textColor=DARK, spaceAfter=3,
        )
        meta_style = ParagraphStyle(
            'meta', parent=base_styles['Normal'],
            fontSize=8, textColor=colors.HexColor('#757575'), spaceAfter=8,
        )
        cell_style = ParagraphStyle(
            'cell', parent=base_styles['Normal'],
            fontSize=7.5, leading=9, textColor=DARK,
        )

        logo_path = os.path.join(settings.STATIC_ROOT or '', 'images', 'logo.png')
        logo_img = None
        if os.path.exists(logo_path):
            try:
                logo_img = RLImage(logo_path, height=28, width=80)
            except Exception:
                logo_img = None

        frontend_url = getattr(settings, 'FRONTEND_URL', 'http://localhost:3000')

        headers = [
            'N° Token', 'Tipo', 'Beneficiario', 'Cód. Emp.',
            'Centro', 'Solicitante',
            'F. Creación', 'F. Válido', 'Estado', 'Observaciones',
        ]

        table_data = [headers]

        for token in queryset:
            st = token.status

            token_url = f'{frontend_url}/tokens/{token.id}'
            id_cell = Paragraph(
                f'<link href="{token_url}" color="#1976d2">{token.display_number or token.id}</link>',
                cell_style,
            )

            beneficiary = '—'
            if token.personnel:
                beneficiary = token.personnel.full_name
            elif hasattr(token, 'exit_pass_detail') and token.exit_pass_detail and token.exit_pass_detail.external_person:
                beneficiary = token.exit_pass_detail.external_person.name

            requester = '—'
            if token.requested_by:
                name = f"{token.requested_by.first_name} {token.requested_by.last_name}".strip()
                requester = name or token.requested_by.email

            table_data.append([
                id_cell,
                Paragraph(self.TOKEN_TYPE_LABELS.get(token.token_type, token.token_type), cell_style),
                Paragraph(beneficiary, cell_style),
                Paragraph(token.personnel.employee_code if token.personnel else '—', cell_style),
                Paragraph(token.distributor_center.name if token.distributor_center else '—', cell_style),
                Paragraph(requester, cell_style),
                token.created_at.strftime('%d/%m/%Y') if token.created_at else '—',
                token.valid_until.strftime('%d/%m/%Y') if token.valid_until else '—',
                Paragraph(self.STATUS_LABELS_EXPORT.get(st, st), ParagraphStyle(
                    f'st_{st}', parent=cell_style,
                    fontName='Helvetica-Bold',
                    textColor=STATUS_COLORS.get(st, DARK),
                )),
                Paragraph(token.requester_notes or '—', cell_style),
            ])

        col_widths = [2*cm, 2.5*cm, 4*cm, 1.7*cm, 3*cm, 3.5*cm, 2*cm, 2*cm, 2*cm, 5*cm]

        table = Table(table_data, colWidths=col_widths, repeatRows=1)

        ts = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), BRAND_BLUE),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7.5),
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (3, 1), (3, -1), 'CENTER'),
            ('ALIGN', (6, 1), (7, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#e0e0e0')),
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, BRAND_BLUE),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ])
        table.setStyle(ts)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=1*cm, rightMargin=1*cm,
            topMargin=1.5*cm, bottomMargin=1.5*cm,
        )

        today_str = date.today().strftime('%d/%m/%Y')
        total = len(table_data) - 1

        title_para = Paragraph('Reporte de Tokens', title_style)
        meta_para = Paragraph(f'Generado: {today_str} &nbsp;&nbsp; Total: {total} registro(s)', meta_style)

        if logo_img:
            hdr_table = Table([[[title_para, meta_para], logo_img]], colWidths=[24.4*cm, 3*cm])
            hdr_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))
            story = [hdr_table, Spacer(1, 0.4*cm), table]
        else:
            story = [title_para, meta_para, Spacer(1, 0.3*cm), table]

        try:
            doc.build(story)
        except Exception as e:
            return Response(
                {'detail': f'Error al generar el PDF: {e}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        buffer.seek(0)
        today_file = date.today().strftime('%Y-%m-%d')
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="tokens-{today_file}.pdf"'
        return response
