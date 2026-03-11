"""
Vistas para certificaciones y entrenamientos
"""
import io
from datetime import date, timedelta, datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.db import transaction
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import OrderingFilter, SearchFilter
from django.db.models import Q, Count
from django.http import HttpResponse

from ..models.certification import Certification, CertificationType
from ..models.personnel import PersonnelProfile
from ..serializers.certification_serializers import (
    CertificationSerializer,
    CertificationListSerializer,
    CertificationTypeSerializer
)
from ..filters import CertificationFilter
from ..permissions import IsSupervisorOrAbove


class CertificationTypeViewSet(viewsets.ModelViewSet):
    """
    ViewSet completo para tipos de certificación (CRUD)
    """
    queryset = CertificationType.objects.all()
    serializer_class = CertificationTypeSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['name', 'code', 'description']
    ordering_fields = ['name', 'code', 'created_at']
    ordering = ['name']

    def get_queryset(self):
        queryset = super().get_queryset()
        is_active = self.request.query_params.get('is_active', None)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        return queryset

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save()


class CertificationViewSet(viewsets.ModelViewSet):
    """
    ViewSet para certificaciones y entrenamientos del personal
    """
    queryset = Certification.objects.select_related(
        'personnel',
        'certification_type',
        'created_by',
        'completed_by',
    )
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = CertificationFilter
    ordering_fields = ['issue_date', 'expiration_date', 'created_at', 'status']
    ordering = ['-created_at']
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'list':
            return CertificationListSerializer
        return CertificationSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user

        if user.is_superuser or user.is_staff:
            return queryset

        try:
            user_personnel = user.personnel_profile
        except PersonnelProfile.DoesNotExist:
            return queryset.none()

        if user_personnel.hierarchy_level == PersonnelProfile.CD_MANAGER:
            return queryset.filter(
                personnel__primary_distributor_center=user_personnel.primary_distributor_center
            )

        if user_personnel.hierarchy_level == PersonnelProfile.AREA_MANAGER:
            return queryset.filter(personnel__area=user_personnel.area)

        if user_personnel.hierarchy_level == PersonnelProfile.SUPERVISOR:
            subordinates_ids = [p.id for p in user_personnel.get_all_subordinates()]
            return queryset.filter(
                Q(personnel=user_personnel) |
                Q(personnel__id__in=subordinates_ids)
            )

        return queryset.filter(personnel=user_personnel)

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    # ──────────────────────────────────────────────────────────────────────────
    # Acciones de flujo de progreso
    # ──────────────────────────────────────────────────────────────────────────

    @action(detail=True, methods=['post'])
    def mark_in_progress(self, request, pk=None):
        """
        Marcar certificación como en progreso
        POST /api/personnel/certifications/{id}/mark_in_progress/
        """
        certification = self.get_object()
        if certification.status == Certification.STATUS_COMPLETED:
            return Response(
                {'detail': 'No se puede modificar una certificación ya completada'},
                status=status.HTTP_400_BAD_REQUEST
            )
        certification.status = Certification.STATUS_IN_PROGRESS
        certification.save()
        serializer = self.get_serializer(certification)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """
        Completar certificación con firma del participante
        POST /api/personnel/certifications/{id}/complete/
        Body: multipart/form-data { signature: file, notes: string (opcional) }
        """
        certification = self.get_object()
        if certification.status == Certification.STATUS_COMPLETED:
            return Response(
                {'detail': 'Esta certificación ya fue completada'},
                status=status.HTTP_400_BAD_REQUEST
            )

        signature_file = request.FILES.get('signature')
        if not signature_file:
            return Response(
                {'detail': 'Se requiere la firma del participante'},
                status=status.HTTP_400_BAD_REQUEST
            )

        notes = request.data.get('notes', '')
        document_file = request.FILES.get('certificate_document')

        certification.status = Certification.STATUS_COMPLETED
        certification.signature = signature_file
        certification.completion_notes = notes
        certification.completed_at = timezone.now()
        certification.completed_by = request.user
        if document_file:
            certification.certificate_document = document_file
        certification.save()

        serializer = self.get_serializer(certification)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def mark_not_completed(self, request, pk=None):
        """
        Marcar certificación como no completada
        POST /api/personnel/certifications/{id}/mark_not_completed/
        Body: { reason: string }
        """
        certification = self.get_object()
        if certification.status == Certification.STATUS_COMPLETED:
            return Response(
                {'detail': 'No se puede modificar una certificación ya completada'},
                status=status.HTTP_400_BAD_REQUEST
            )

        reason = request.data.get('reason', '').strip()
        if not reason:
            return Response(
                {'detail': 'Debe proporcionar el motivo de no completado'},
                status=status.HTTP_400_BAD_REQUEST
            )

        certification.status = Certification.STATUS_NOT_COMPLETED
        certification.non_completion_reason = reason
        certification.save()

        serializer = self.get_serializer(certification)
        return Response(serializer.data)

    # ──────────────────────────────────────────────────────────────────────────
    # Acciones existentes
    # ──────────────────────────────────────────────────────────────────────────

    @action(detail=False, methods=['get'])
    def expiring_soon(self, request):
        """Certificaciones que vencen pronto"""
        days = int(request.query_params.get('days', 30))
        threshold = date.today() + timedelta(days=days)
        queryset = self.get_queryset().filter(
            expiration_date__lte=threshold,
            expiration_date__gte=date.today(),
            is_valid=True
        )
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def expired(self, request):
        """Certificaciones vencidas"""
        queryset = self.get_queryset().filter(
            expiration_date__lt=date.today(),
            is_valid=True
        )
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def revoke(self, request, pk=None):
        """Revocar una certificación"""
        certification = self.get_object()
        reason = request.data.get('reason', '')
        if not reason:
            return Response(
                {'detail': 'Debe proporcionar un motivo de revocación'},
                status=status.HTTP_400_BAD_REQUEST
            )
        certification.revoked = True
        certification.is_valid = False
        certification.revocation_reason = reason
        certification.revocation_date = date.today()
        certification.save()
        serializer = self.get_serializer(certification)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Estadísticas de certificaciones"""
        queryset = self.get_queryset()
        by_type = queryset.filter(is_valid=True).values(
            'certification_type__name'
        ).annotate(count=Count('id'))

        by_status = queryset.values('status').annotate(count=Count('id'))

        data = {
            'by_type': list(by_type),
            'by_status': {item['status']: item['count'] for item in by_status},
            'total': queryset.count()
        }
        return Response(data)

    # ──────────────────────────────────────────────────────────────────────────
    # Carga masiva
    # ──────────────────────────────────────────────────────────────────────────

    @action(detail=False, methods=['get'], url_path='bulk_upload_template')
    def bulk_upload_template(self, request):
        """
        Descarga plantilla Excel para carga masiva de certificaciones
        GET /api/personnel/certifications/bulk_upload_template/
        """
        wb = openpyxl.Workbook()

        # ── Estilos ───────────────────────────────────────────────────────────
        HDR_FILL = '1976D2'   # azul primario
        HDR_OPT  = '1565C0'   # azul (campos opcionales)
        EX_FILL  = 'FFEBEE'   # fondo filas de ejemplo

        def _hdr(ws, row, col, value, fill_hex, width=None):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(bold=True, color='FFFFFF', size=10)
            cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            if width:
                ws.column_dimensions[get_column_letter(col)].width = width
            return cell

        def _ex(ws, row, col, value):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill = PatternFill(start_color=EX_FILL, end_color=EX_FILL, fill_type='solid')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            return cell

        # ── Hoja 1: Datos ─────────────────────────────────────────────────────
        ws = wb.active
        ws.title = 'Datos'
        ws.row_dimensions[1].height = 35

        COLS = [
            ('Codigo_Empleado*',     HDR_FILL, 22),
            ('Codigo_Certificacion*', HDR_FILL, 26),
            ('Fecha_Inicio',         HDR_OPT,  18),
            ('Fecha_Fin',            HDR_OPT,  18),
            ('Instructor_Autoridad', HDR_OPT,  28),
            ('Numero_Certificacion', HDR_OPT,  24),
            ('Notas',                HDR_OPT,  30),
        ]

        for i, (label, color, w) in enumerate(COLS, start=1):
            _hdr(ws, 1, i, label, color, w)

        # Fila de ejemplo (marcada claramente para que el usuario la borre)
        example = [
            'EMP001 ← BORRAR ESTA FILA',
            'VER_HOJA_TIPOS',
            '15/03/2026',
            '15/03/2027',
            'Instituto Nacional de Seguridad',
            'CERT-2026-001',
            'Entrenamiento anual obligatorio',
        ]
        for i, val in enumerate(example, start=1):
            _ex(ws, 2, i, val)

        ws.freeze_panes = 'A2'

        # ── Hoja 2: Tipos de Certificación disponibles ────────────────────────
        ws_tipos = wb.create_sheet('Tipos_Certificacion')
        ws_tipos.row_dimensions[1].height = 30

        tipos_cols = [
            ('Codigo', '1976D2', 22),
            ('Nombre', '1565C0', 42),
            ('Descripcion', '2E7D32', 50),
            ('Valido (dias)', '6A1B9A', 16),
        ]
        for i, (label, color, w) in enumerate(tipos_cols, start=1):
            _hdr(ws_tipos, 1, i, label, color, w)

        cert_types = CertificationType.objects.filter(is_active=True).order_by('code')
        TIPO_ROW_FILL = 'F8F9FA'
        for row_idx, ct in enumerate(cert_types, start=2):
            row_bg = 'FFFFFF' if row_idx % 2 == 0 else TIPO_ROW_FILL
            for col_idx, value in enumerate([
                ct.code,
                ct.name,
                ct.description or '—',
                ct.validity_period_days or '—',
            ], start=1):
                cell = ws_tipos.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type='solid')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                # Código en negrita
                if col_idx == 1:
                    cell.font = Font(bold=True, size=10)

        ws_tipos.freeze_panes = 'A2'

        if not cert_types.exists():
            no_data = ws_tipos.cell(row=2, column=1, value='No hay tipos de certificación activos en el sistema.')
            no_data.font = Font(italic=True, color='999999')

        # ── Hoja 3: Instrucciones ─────────────────────────────────────────────
        ws2 = wb.create_sheet('Instrucciones')
        ws2.column_dimensions['A'].width = 28
        ws2.column_dimensions['B'].width = 60

        instrucciones = [
            ('Campo', 'Descripción'),
            ('Codigo_Empleado*', 'Código del empleado (requerido). Debe existir en el sistema.'),
            ('Codigo_Certificacion*', 'Código del tipo (requerido). Ver hoja "Tipos_Certificacion" para los códigos válidos.'),
            ('Fecha_Inicio', 'Fecha de inicio/emisión. Formato: DD/MM/YYYY (opcional).'),
            ('Fecha_Fin', 'Fecha de vencimiento. Formato: DD/MM/YYYY (opcional).'),
            ('Instructor_Autoridad', 'Nombre del instructor o autoridad emisora (opcional).'),
            ('Numero_Certificacion', 'Número de registro del certificado (opcional).'),
            ('Notas', 'Notas adicionales (opcional).'),
            ('', ''),
            ('NOTA', 'Los campos con * son obligatorios.'),
            ('NOTA', 'Borra la fila 2 (es solo un ejemplo).'),
            ('NOTA', 'Usa los códigos exactos de la hoja "Tipos_Certificacion".'),
            ('NOTA', 'Las certificaciones se crean en estado PENDIENTE.'),
            ('NOTA', 'Máximo 500 filas por archivo.'),
        ]

        header_fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
        for row_idx, (campo, desc) in enumerate(instrucciones, start=1):
            c1 = ws2.cell(row=row_idx, column=1, value=campo)
            c2 = ws2.cell(row=row_idx, column=2, value=desc)
            if row_idx == 1:
                for c in [c1, c2]:
                    c.font = Font(bold=True, color='FFFFFF')
                    c.fill = header_fill
            c1.alignment = Alignment(vertical='center', wrap_text=True)
            c2.alignment = Alignment(vertical='center', wrap_text=True)

        # ── Respuesta ─────────────────────────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_certificaciones.xlsx"'
        return response

    @action(detail=False, methods=['post'], url_path='bulk_upload_preview')
    def bulk_upload_preview(self, request):
        """
        Previsualiza y valida un archivo Excel de carga masiva de certificaciones
        POST /api/personnel/certifications/bulk_upload_preview/
        Body: multipart/form-data { file: xlsx }
        """
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'Se requiere un archivo Excel'}, status=status.HTTP_400_BAD_REQUEST)

        if not file.name.endswith(('.xlsx', '.xls')):
            return Response({'detail': 'El archivo debe ser .xlsx o .xls'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        except Exception:
            return Response({'detail': 'No se pudo leer el archivo Excel'}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        if not rows:
            return Response({'detail': 'El archivo no contiene datos'}, status=status.HTTP_400_BAD_REQUEST)

        if len(rows) > 500:
            return Response({'detail': 'El archivo supera el máximo de 500 filas'}, status=status.HTTP_400_BAD_REQUEST)

        # Precargar datos para validación
        employee_codes = set(
            PersonnelProfile.objects.values_list('employee_code', flat=True)
        )
        _cert_types = CertificationType.objects.filter(is_active=True)
        cert_type_codes = {ct.code: ct.id for ct in _cert_types}
        cert_type_names = {ct.code: ct.name for ct in _cert_types}

        valid_rows = []
        error_rows = []

        for row_num, row in enumerate(rows, start=2):
            # Saltar filas completamente vacías
            if not any(v for v in row):
                continue

            employee_code = str(row[0]).strip() if row[0] is not None else ''
            cert_code     = str(row[1]).strip() if row[1] is not None else ''
            fecha_inicio  = row[2]
            fecha_fin     = row[3]
            instructor    = str(row[4]).strip() if row[4] is not None else ''
            num_cert      = str(row[5]).strip() if row[5] is not None else ''
            notas         = str(row[6]).strip() if row[6] is not None else ''

            errores = []

            # Validar campos requeridos
            if not employee_code:
                errores.append({'campo': 'Codigo_Empleado', 'mensaje': 'Campo requerido'})
            elif employee_code not in employee_codes:
                errores.append({'campo': 'Codigo_Empleado', 'mensaje': f'No existe empleado con código "{employee_code}"'})

            if not cert_code:
                errores.append({'campo': 'Codigo_Certificacion', 'mensaje': 'Campo requerido'})
            elif cert_code not in cert_type_codes:
                errores.append({'campo': 'Codigo_Certificacion', 'mensaje': f'No existe tipo de certificación con código "{cert_code}"'})

            # Parsear fechas
            parsed_inicio = None
            if fecha_inicio:
                parsed_inicio = self._parse_date(fecha_inicio)
                if parsed_inicio is None:
                    errores.append({'campo': 'Fecha_Inicio', 'mensaje': 'Formato inválido. Use DD/MM/YYYY'})

            parsed_fin = None
            if fecha_fin:
                parsed_fin = self._parse_date(fecha_fin)
                if parsed_fin is None:
                    errores.append({'campo': 'Fecha_Fin', 'mensaje': 'Formato inválido. Use DD/MM/YYYY'})

            if parsed_inicio and parsed_fin and parsed_fin <= parsed_inicio:
                errores.append({'campo': 'Fecha_Fin', 'mensaje': 'La fecha de fin debe ser posterior a la fecha de inicio'})

            fila_data = {
                'fila': row_num,
                'employee_code': employee_code,
                'certification_type_code': cert_code,
                'certification_type_name': cert_type_names.get(cert_code, ''),
                'issue_date': parsed_inicio.strftime('%Y-%m-%d') if parsed_inicio else None,
                'expiration_date': parsed_fin.strftime('%Y-%m-%d') if parsed_fin else None,
                'issuing_authority': instructor,
                'certification_number': num_cert,
                'notes': notas,
            }

            if errores:
                error_rows.append({'fila': row_num, 'datos': fila_data, 'errores': errores})
            else:
                valid_rows.append(fila_data)

        return Response({
            'total': len(valid_rows) + len(error_rows),
            'valid': len(valid_rows),
            'errors': len(error_rows),
            'valid_rows': valid_rows,
            'error_rows': error_rows,
        })

    @action(detail=False, methods=['post'], url_path='bulk_upload_confirm')
    def bulk_upload_confirm(self, request):
        """
        Crea certificaciones en masa a partir de filas validadas
        POST /api/personnel/certifications/bulk_upload_confirm/
        Body: JSON { rows: [...valid_rows] }
        """
        rows = request.data.get('rows', [])
        if not rows:
            return Response({'detail': 'No se proporcionaron filas para crear'}, status=status.HTTP_400_BAD_REQUEST)

        initial_status_value = request.data.get('initial_status', Certification.STATUS_PENDING)
        if initial_status_value not in (Certification.STATUS_PENDING, Certification.STATUS_IN_PROGRESS):
            return Response({'detail': 'Estado inicial inválido'}, status=status.HTTP_400_BAD_REQUEST)

        # Precargar datos
        personnel_map = {
            p.employee_code: p
            for p in PersonnelProfile.objects.filter(
                employee_code__in=[r['employee_code'] for r in rows]
            )
        }
        cert_type_map = {
            ct.code: ct
            for ct in CertificationType.objects.filter(
                code__in=[r['certification_type_code'] for r in rows]
            )
        }

        created = []
        errors = []

        try:
            with transaction.atomic():
                for row in rows:
                    employee_code = row.get('employee_code', '')
                    cert_code = row.get('certification_type_code', '')

                    personnel = personnel_map.get(employee_code)
                    cert_type = cert_type_map.get(cert_code)

                    if not personnel or not cert_type:
                        errors.append({
                            'employee_code': employee_code,
                            'error': 'Empleado o tipo de certificación no encontrado'
                        })
                        continue

                    issue_date_raw = row.get('issue_date') or None
                    expiration_date_raw = row.get('expiration_date') or None

                    issue_date = date.fromisoformat(issue_date_raw) if issue_date_raw else None
                    expiration_date = date.fromisoformat(expiration_date_raw) if expiration_date_raw else None

                    # Si no hay fecha de vencimiento pero el tipo tiene período, calcularla
                    if not expiration_date and issue_date and cert_type.validity_period_days:
                        expiration_date = issue_date + timedelta(days=cert_type.validity_period_days)

                    cert = Certification.objects.create(
                        personnel=personnel,
                        certification_type=cert_type,
                        status=initial_status_value,
                        issue_date=issue_date or date.today(),
                        expiration_date=expiration_date or date.today(),
                        issuing_authority=row.get('issuing_authority', ''),
                        certification_number=row.get('certification_number', ''),
                        notes=row.get('notes', ''),
                        created_by=request.user,
                    )
                    created.append({
                        'id': cert.id,
                        'employee_code': employee_code,
                        'employee_name': personnel.full_name,
                        'certification_type': cert_type.name,
                        'certification_number': cert.certification_number or '',
                        'status': cert.status,
                    })

                if errors:
                    raise Exception('Errores en la carga')

        except Exception:
            if errors:
                return Response({
                    'detail': 'Algunos registros no se pudieron crear',
                    'errors': errors
                }, status=status.HTTP_400_BAD_REQUEST)
            raise

        return Response({
            'status': 'success',
            'created': len(created),
            'records': created,
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='export_pdf')
    def export_pdf(self, request):
        """
        Exporta el listado de certificaciones filtrado a PDF (tabla con firma).
        GET /api/personnel/certifications/export_pdf/
        Usa ReportLab (pip install reportlab).
        """
        import os
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
        except ImportError:
            return Response(
                {'detail': 'Módulo reportlab no instalado. Ejecute: pip install reportlab'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        from django.conf import settings

        queryset = self.filter_queryset(self.get_queryset())

        STATUS_LABELS = {
            'PENDING':       'Pendiente',
            'IN_PROGRESS':   'En Progreso',
            'COMPLETED':     'Completado',
            'NOT_COMPLETED': 'No Completó',
        }
        STATUS_COLORS = {
            'PENDING':       colors.HexColor('#757575'),
            'IN_PROGRESS':   colors.HexColor('#1565c0'),
            'COMPLETED':     colors.HexColor('#2e7d32'),
            'NOT_COMPLETED': colors.HexColor('#c62828'),
        }

        BRAND_BLUE = colors.HexColor('#1976d2')
        DARK       = colors.HexColor('#1a1a2e')
        LIGHT_GRAY = colors.HexColor('#f5f5f5')

        # ── Estilos de texto ──────────────────────────────────────────────────
        base_styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'title', parent=base_styles['Normal'],
            fontSize=13, fontName='Helvetica-Bold',
            textColor=DARK, spaceAfter=3,
        )
        meta_style = ParagraphStyle(
            'meta', parent=base_styles['Normal'],
            fontSize=8, textColor=colors.HexColor('#757575'), spaceAfter=8,
        )
        cell_style = ParagraphStyle(
            'cell', parent=base_styles['Normal'],
            fontSize=7.5, leading=9, textColor=DARK,
        )

        # ── Logo (archivo local estático) ─────────────────────────────────────
        logo_path = os.path.join(settings.STATIC_ROOT, 'images', 'logo.png')
        logo_img = None
        if os.path.exists(logo_path):
            try:
                logo_img = RLImage(logo_path, height=28, width=80)
            except Exception:
                logo_img = None

        # ── Datos de la tabla ─────────────────────────────────────────────────
        SIG_H = 35
        SIG_W = 70

        frontend_url = getattr(settings, 'FRONTEND_URL', 'http://localhost:3000')

        headers = [
            'N° Ref.', 'Código', 'Nombre', 'Tipo Certificación',
            'N° Cert', 'Instructor / Autoridad',
            'F. Inicio', 'F. Venc.', 'Estado', 'Válido', 'Firma',
        ]

        table_data = [headers]

        for cert in queryset:
            st = cert.status

            # ID con enlace clickeable
            cert_url = f'{frontend_url}/personnel/certifications/{cert.id}'
            id_cell = Paragraph(
                f'<link href="{cert_url}" color="#1976d2">{cert.id}</link>',
                cell_style,
            )

            # Firma: lee desde el storage backend de Django (funciona con Azure, S3, local, etc.)
            sig_cell = '—'
            if cert.signature:
                try:
                    sig_bytes = io.BytesIO(cert.signature.read())
                    sig_img = RLImage(sig_bytes, width=SIG_W, height=SIG_H, kind='proportional')
                    sig_img.hAlign = 'CENTER'
                    sig_cell = sig_img
                except Exception:
                    pass

            table_data.append([
                id_cell,
                Paragraph(cert.personnel.employee_code if cert.personnel else '—', cell_style),
                Paragraph(cert.personnel.full_name if cert.personnel else '—', cell_style),
                Paragraph(cert.certification_type.name if cert.certification_type else '—', cell_style),
                Paragraph(cert.certification_number or '—', cell_style),
                Paragraph(cert.issuing_authority or '—', cell_style),
                cert.issue_date.strftime('%d/%m/%Y') if cert.issue_date else '—',
                cert.expiration_date.strftime('%d/%m/%Y') if cert.expiration_date else '—',
                Paragraph(STATUS_LABELS.get(st, st), ParagraphStyle(
                    f'st_{st}', parent=cell_style,
                    fontName='Helvetica-Bold',
                    textColor=STATUS_COLORS.get(st, DARK),
                )),
                'Sí' if cert.is_valid else 'No',
                sig_cell,
            ])

        # ── Anchos de columna (A4 landscape ≈ 27.7 cm útiles) ────────────────
        # 11 columnas: ID, Código, Nombre, Tipo, N°Cert, Instructor, FInicio, FFin, Estado, Válido, Firma
        col_widths = [1.3*cm, 1.7*cm, 4*cm, 4*cm, 1.9*cm, 3.5*cm, 1.9*cm, 1.9*cm, 2.3*cm, 1.3*cm, 3.4*cm]

        table = Table(table_data, colWidths=col_widths, repeatRows=1)

        ts = TableStyle([
            # Cabecera
            ('BACKGROUND',    (0, 0), (-1, 0), BRAND_BLUE),
            ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
            ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, 0), 8),
            ('ALIGN',         (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN',        (0, 0), (-1, 0), 'MIDDLE'),
            # Filas alternas
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
            # Cuerpo
            ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE',      (0, 1), (-1, -1), 7.5),
            ('VALIGN',        (0, 1), (-1, -1), 'MIDDLE'),
            ('ALIGN',         (0, 1), (1, -1), 'CENTER'),  # ID + código centrados
            ('ALIGN',         (6, 1), (7, -1), 'CENTER'),  # fechas centradas
            ('ALIGN',         (9, 1), (10, -1), 'CENTER'), # válido + firma centrados
            # Bordes
            ('GRID',          (0, 0), (-1, -1), 0.3, colors.HexColor('#e0e0e0')),
            ('LINEBELOW',     (0, 0), (-1, 0), 1.5, BRAND_BLUE),
            # Padding
            ('TOPPADDING',    (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING',   (0, 0), (-1, -1), 4),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
        ])
        table.setStyle(ts)

        # ── Construir documento ───────────────────────────────────────────────
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=1*cm, rightMargin=1*cm,
            topMargin=1.5*cm, bottomMargin=1.5*cm,
        )

        today_str = date.today().strftime('%d/%m/%Y')
        total     = len(table_data) - 1

        title_para = Paragraph('Reporte de Certificaciones y Entrenamientos', title_style)
        meta_para  = Paragraph(f'Generado: {today_str} &nbsp;&nbsp; Total: {total} registro(s)', meta_style)

        if logo_img:
            # Título a la izquierda, logo a la derecha
            hdr_table = Table([[[title_para, meta_para], logo_img]], colWidths=[24.4*cm, 3*cm])
            hdr_table.setStyle(TableStyle([
                ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN',         (1, 0), (1, 0),   'RIGHT'),
                ('LEFTPADDING',   (0, 0), (-1, -1), 0),
                ('RIGHTPADDING',  (0, 0), (-1, -1), 0),
                ('TOPPADDING',    (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))
            story = [hdr_table, Spacer(1, 0.4*cm), table]
        else:
            story = [title_para, meta_para, Spacer(1, 0.3*cm), table]

        try:
            doc.build(story)
        except Exception as e:
            return Response(
                {'detail': f'Error al generar el PDF: {e}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        buffer.seek(0)
        today_file = date.today().strftime('%Y-%m-%d')
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="certificaciones-{today_file}.pdf"'
        return response

    @action(detail=False, methods=['get'], url_path='export_excel')
    def export_excel(self, request):
        """
        Exporta el listado de certificaciones filtrado a Excel.
        GET /api/personnel/certifications/export_excel/
        Respeta los mismos filtros que el listado principal.
        """
        queryset = self.filter_queryset(self.get_queryset())

        HDR_FILL = '1976D2'
        EX_FILL  = 'F8F9FA'

        def _hdr(ws, row, col, value, width=None):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(bold=True, color='FFFFFF', size=10)
            cell.fill = PatternFill(start_color=HDR_FILL, end_color=HDR_FILL, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'),
            )
            if width:
                ws.column_dimensions[get_column_letter(col)].width = width
            return cell

        STATUS_LABELS = {
            'PENDING': 'Pendiente',
            'IN_PROGRESS': 'En Progreso',
            'COMPLETED': 'Completado',
            'NOT_COMPLETED': 'No Completó',
        }

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Certificaciones'
        ws.row_dimensions[1].height = 32

        COLS = [
            ('Código',                16),
            ('Nombre',                32),
            ('Tipo Certificación',    32),
            ('N° Cert',               20),
            ('Instructor / Autoridad',30),
            ('Fecha Inicio',          16),
            ('Fecha Fin',             16),
            ('Estado',                18),
            ('Válido',                10),
            ('Notas',                 36),
        ]
        for i, (label, w) in enumerate(COLS, start=1):
            _hdr(ws, 1, i, label, w)

        ws.freeze_panes = 'A2'

        for row_idx, cert in enumerate(queryset, start=2):
            row_bg = 'FFFFFF' if row_idx % 2 == 0 else EX_FILL
            values = [
                cert.personnel.employee_code if cert.personnel else '',
                cert.personnel.full_name if cert.personnel else '',
                cert.certification_type.name if cert.certification_type else '',
                cert.certification_number or '',
                cert.issuing_authority or '',
                cert.issue_date.strftime('%d/%m/%Y') if cert.issue_date else '',
                cert.expiration_date.strftime('%d/%m/%Y') if cert.expiration_date else '',
                STATUS_LABELS.get(cert.status, cert.status),
                'Sí' if cert.is_valid else 'No',
                cert.notes or '',
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type='solid')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'),
                )
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        today = date.today().strftime('%Y-%m-%d')
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="certificaciones-{today}.xlsx"'
        return response

    # ──────────────────────────────────────────────────────────────────────────
    # Utilidades privadas
    # ──────────────────────────────────────────────────────────────────────────

    @staticmethod
    def _parse_date(value):
        """Parsea una fecha en varios formatos. Retorna un objeto date o None."""
        if isinstance(value, (date, datetime)):
            return value if isinstance(value, date) else value.date()
        if not value:
            return None
        value = str(value).strip()
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        return None
