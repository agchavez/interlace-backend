
# Rest_framework
from rest_framework import mixins, viewsets, permissions
from rest_framework.decorators import permission_classes, action
from rest_framework.permissions import IsAuthenticated, IsAdminUser, IsAuthenticatedOrReadOnly, BasePermission
from rest_framework.response import Response

# Models
from apps.user.models import UserModel, DetailGroup
from django.contrib.auth.models import Group, Permission
from apps.maintenance.models.distributor_center import DistributorCenter

from django.contrib.auth import get_user_model
User = get_user_model()

# Log de administrador
from django.contrib.admin.models import LogEntry

# Serializers
from apps.user.serializers import (UserSerializer,
                                   LogEntrySerializer, DetailGroupSerializer,
                                   BulkUploadRowSerializer)

# Utilities
import io
import unicodedata
from django.db import transaction
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Personnel models (para carga masiva)
from apps.personnel.models import PersonnelProfile, Area

# filters
from django_filters.rest_framework import DjangoFilterBackend
import django_filters
from rest_framework.filters import SearchFilter, OrderingFilter


class CustomAccessPermission(BasePermission):
    """
    Clase personalizada de permisos que verifica si el usuario tiene los permisos adecuados.
    """

    def has_permission(self, request, view):
        # usuario esta activo
        if not request.user.is_active:
            return False
        if request.method in permissions.SAFE_METHODS:
            # Si el método es seguro (GET, HEAD, OPTIONS), permitir el acceso a todos
            return True

        # Obtener los permisos requeridos para la acción específica (crear, actualizar, eliminar, etc.)
        required_permissions = view.get_required_permissions(request.method)

        # Verificar si el usuario tiene todos los permisos requeridos
        return request.user.has_perms(required_permissions)


class UserFilter(django_filters.FilterSet):
    class Meta:
        model = UserModel
        fields = {
            'is_active': ['exact'],
            'is_staff': ['exact'],
            'is_superuser': ['exact'],
            'groups': ['exact'],
        }

# ViewSets by UserModel
class UserViewSet(mixins.CreateModelMixin,
                  mixins.RetrieveModelMixin,
                  mixins.UpdateModelMixin,
                  mixins.DestroyModelMixin,
                  mixins.ListModelMixin,
                  viewsets.GenericViewSet):
    # Omitir los superusuarios
    queryset = UserModel.objects.filter(is_superuser=False)
    serializer_class = UserSerializer
    filter_backends = (SearchFilter, DjangoFilterBackend)
    search_fields = ('username', 'email', 'first_name', 'last_name')
    filterset_class = UserFilter 
    #permission_classes = [CustomAccessPermission]
    # Mapeo de métodos HTTP a los permisos requeridos
    PERMISSION_MAPPING = {
        'GET': ['user.view_usermodel'],
        'POST': ['user.add_usermodel'],
        'PUT': ['user.change_usermodel'],
        'PATCH': ['user.change_usermodel'],
        'DELETE': ['user.delete_usermodel']
    }

    def get_required_permissions(self, http_method):
        return self.PERMISSION_MAPPING.get(http_method, [])

    # listar usuarios
    def list(self, request, *args, **kwargs):
        query = self.filter_queryset(self.get_queryset())
        return super().list(request, *args, **kwargs)

    # actualizar mi perfil
    @action(methods=['put'], detail=False, permission_classes=[IsAuthenticated], url_path='update-profile')
    def update_profile(self, request):
        user = request.user
        serializer = UserSerializer(user, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @action(methods=['post'], detail=False, permission_classes=[IsAuthenticated], url_path='generate-username')
    def generate_username(self, request):
        """
        Genera sugerencias de username basado en nombre y apellido
        POST /api/users/generate-username/
        Body: {"first_name": "Juan", "last_name": "Perez"}
        """
        first_name = request.data.get('first_name', '').strip()
        last_name = request.data.get('last_name', '').strip()

        if not first_name or not last_name:
            return Response({
                'error': 'Se requieren first_name y last_name'
            }, status=400)

        # Normalizar nombres (eliminar acentos, convertir a minúsculas)
        import unicodedata
        def normalize(text):
            text = unicodedata.normalize('NFD', text)
            text = text.encode('ascii', 'ignore').decode('utf-8')
            return text.lower().replace(' ', '')

        first = normalize(first_name)
        last = normalize(last_name)

        # Generar múltiples sugerencias
        suggestions = []
        patterns = [
            f"{first}.{last}",                    # juan.perez
            f"{first[0]}{last}",                  # jperez
            f"{first}{last[0]}",                  # juanp
            f"{first}_{last}",                    # juan_perez
            f"{last}.{first}",                    # perez.juan
            f"{first[:3]}{last[:3]}",             # juaper
        ]

        # Verificar disponibilidad y agregar números si es necesario
        for pattern in patterns:
            username = pattern
            counter = 1

            # Verificar disponibilidad
            while UserModel.objects.filter(username=username).exists():
                username = f"{pattern}{counter}"
                counter += 1
                if counter > 99:  # Límite de seguridad
                    break

            if counter <= 99:  # Solo agregar si encontramos uno disponible
                suggestions.append({
                    'username': username,
                    'available': True
                })

            if len(suggestions) >= 5:  # Máximo 5 sugerencias
                break

        return Response({
            'suggestions': suggestions,
            'first_name': first_name,
            'last_name': last_name
        })

    @action(methods=['post'], detail=False, permission_classes=[IsAuthenticated], url_path='check-username')
    def check_username(self, request):
        """
        Verifica si un username está disponible
        POST /api/users/check-username/
        Body: {"username": "jperez"}
        """
        username = request.data.get('username', '').strip()

        if not username:
            return Response({
                'error': 'Se requiere username'
            }, status=400)

        # Validar formato (solo letras, números, puntos, guiones bajos)
        import re
        if not re.match(r'^[a-zA-Z0-9._-]+$', username):
            return Response({
                'available': False,
                'error': 'El username solo puede contener letras, números, puntos, guiones bajos y guiones'
            })

        # Verificar longitud
        if len(username) < 3:
            return Response({
                'available': False,
                'error': 'El username debe tener al menos 3 caracteres'
            })

        if len(username) > 150:
            return Response({
                'available': False,
                'error': 'El username no puede exceder 150 caracteres'
            })

        # Verificar disponibilidad
        available = not UserModel.objects.filter(username=username).exists()

        return Response({
            'username': username,
            'available': available,
            'message': 'Usuario disponible' if available else 'Usuario no disponible'
        })

    # ──────────────────────────────────────────────────────────────
    # CARGA MASIVA — Constantes de mapeo
    # ──────────────────────────────────────────────────────────────
    _TIPO_REGISTRO_MAP = {
        'SOLO_PERSONAL': 'SOLO_PERSONAL',
        'CON_USUARIO': 'CON_USUARIO',
    }
    _GENERO_MAP = {'M': 'M', 'F': 'F'}
    _ESTADO_CIVIL_MAP = {
        'SOLTERO': 'SINGLE', 'CASADA': 'MARRIED', 'CASADO': 'MARRIED',
        'DIVORCIADO': 'DIVORCED', 'DIVORCIADA': 'DIVORCED',
        'VIUDO': 'WIDOWED', 'VIUDA': 'WIDOWED',
        'UNION_LIBRE': 'UNION',
    }
    _CONTRATO_MAP = {'PERMANENTE': 'PERMANENT', 'TEMPORAL': 'TEMPORARY', 'CONTRATO': 'CONTRACT'}
    _AREA_MAP = {
        'OPERACIONES': 'OPERATIONS', 'ADMINISTRACION': 'ADMINISTRATION',
        'PEOPLE': 'PEOPLE', 'SEGURIDAD': 'SECURITY', 'DELIVERY': 'DELIVERY',
    }
    _JERARQUIA_MAP = {
        'OPERATIVO': 'OPERATIVE', 'SUPERVISOR': 'SUPERVISOR',
        'JEFE_AREA': 'AREA_MANAGER', 'GERENTE_CD': 'CD_MANAGER',
    }
    _POSICION_MAP = {
        'PICKER': 'PICKER', 'CONTADOR': 'COUNTER', 'OPM': 'OPM',
        'CONDUCTOR_PATIO': 'YARD_DRIVER', 'CARGADOR': 'LOADER',
        'AYUDANTE_ALMACEN': 'WAREHOUSE_ASSISTANT', 'GUARDIA': 'SECURITY_GUARD',
        'CONDUCTOR_DELIVERY': 'DELIVERY_DRIVER', 'ADMINISTRATIVO': 'ADMINISTRATIVE',
        'OTRO': 'OTHER',
    }
    # Tallas: texto libre (máximo 10 caracteres) — sin validación de valores fijos

    # ──────────────────────────────────────────────────────────────
    # CARGA MASIVA — Plantilla
    # ──────────────────────────────────────────────────────────────
    @action(methods=['get'], detail=False, permission_classes=[IsAuthenticated],
            url_path='bulk-upload-template')
    def bulk_upload_template(self, request):
        """
        Descarga la plantilla Excel para carga masiva de personal.
        GET /api/users/bulk-upload-template/
        La plantilla incluye TODOS los campos del PersonnelProfile y diferencia
        entre SOLO_PERSONAL (sin acceso al sistema) y CON_USUARIO (con login).
        """
        if not request.user.has_perm('user.add_usermodel'):
            return Response({'detail': 'No tienes permiso para realizar esta acción.'}, status=403)

        wb = openpyxl.Workbook()

        # ── Estilos base ──────────────────────────────────────────
        def _hdr(ws, row, col, value, fill_hex, width=None):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(bold=True, color='FFFFFF', size=10)
            cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            if width:
                ws.column_dimensions[get_column_letter(col)].width = width
            return cell

        def _ex(ws, row, col, value):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            return cell

        # ── Hoja 1: Datos ─────────────────────────────────────────
        ws = wb.active
        ws.title = 'Datos'
        ws.row_dimensions[1].height = 40
        ws.row_dimensions[2].height = 18

        # Colores por sección
        COL_TIPO   = '1565C0'  # azul oscuro
        COL_PERS   = '1976D2'  # azul
        COL_LAB    = '2E7D32'  # verde oscuro
        COL_EPP    = '6A1B9A'  # morado
        COL_SYS    = 'B71C1C'  # rojo (solo CON_USUARIO)

        # (encabezado, color, ancho)
        COLS = [
            # Sección tipo
            ('Tipo_Registro*',       COL_TIPO, 18),
            # Datos personales
            ('Nombres*',             COL_PERS, 20),
            ('Apellidos*',           COL_PERS, 20),
            ('Codigo_Empleado*',     COL_PERS, 18),
            ('Num_Identidad',        COL_PERS, 18),
            ('Fecha_Nacimiento*',    COL_PERS, 18),
            ('Genero*',              COL_PERS, 12),
            ('Estado_Civil',         COL_PERS, 16),
            ('Telefono*',            COL_PERS, 16),
            ('Email_Contacto',       COL_PERS, 28),
            ('Direccion',            COL_PERS, 28),
            ('Ciudad',               COL_PERS, 16),
            # Datos laborales
            ('Fecha_Ingreso*',       COL_LAB, 16),
            ('Tipo_Contrato*',       COL_LAB, 16),
            ('Area*',                COL_LAB, 16),
            ('Nivel_Jerarquico*',    COL_LAB, 18),
            ('Puesto*',              COL_LAB, 24),
            ('Tipo_Posicion*',       COL_LAB, 20),
            # Tallas EPP
            ('Talla_Camisa',         COL_EPP, 14),
            ('Talla_Pantalon',       COL_EPP, 14),
            ('Talla_Zapatos',        COL_EPP, 14),
            ('Talla_Guantes',        COL_EPP, 14),
            ('Talla_Casco',          COL_EPP, 13),
            # Acceso al sistema (solo CON_USUARIO)
            ('Email_Sistema',        COL_SYS, 28),
            ('Username_Sistema',     COL_SYS, 20),
            ('Contrasena_Sistema',   COL_SYS, 20),
            ('Grupo_Sistema',        COL_SYS, 18),
        ]

        for col_idx, (label, color, width) in enumerate(COLS, start=1):
            _hdr(ws, 1, col_idx, label, color, width)

        # Fila de ejemplo 1 — SOLO_PERSONAL (operativo sin acceso)
        ex1 = [
            'SOLO_PERSONAL',
            'María', 'López García', 'OP-2201', '0801199900001',
            '15/03/1995', 'F', 'CASADA', '9999-1234', 'maria.contacto@gmail.com',
            'Col. Kennedy, Casa 45', 'Tegucigalpa',
            '01/06/2022', 'PERMANENTE', 'OPERACIONES', 'OPERATIVO',
            'Picker de Turno Mañana', 'PICKER',
            'M', '32', '8', 'M', 'M',
            '', '', '', '',  # sin acceso al sistema
        ]
        for col_idx, val in enumerate(ex1, start=1):
            _ex(ws, 2, col_idx, val)

        # Fila de ejemplo 2 — CON_USUARIO (supervisor con acceso)
        ex2 = [
            'CON_USUARIO',
            'Carlos', 'Ramírez Mejía', 'SUP-0045', '0801198800002',
            '20/07/1988', 'M', 'CASADO', '9888-5678', 'carlos.ramirez@empresa.com',
            'Res. Los Pinos, Apto 3B', 'San Pedro Sula',
            '15/01/2019', 'PERMANENTE', 'OPERACIONES', 'SUPERVISOR',
            'Supervisor de Turno Noche', 'AYUDANTE_ALMACEN',
            'L', '34', '9', 'L', 'L',
            'carlos.ramirez@empresa.com', 'carlos.ramirez', 'Sup2024!', 'SUPERVISOR',
        ]
        ws.row_dimensions[3].height = 18
        for col_idx, val in enumerate(ex2, start=1):
            cell = ws.cell(row=3, column=col_idx, value=val)
            cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            cell.alignment = Alignment(vertical='center', wrap_text=True)

        ws.freeze_panes = 'A2'

        # ── Data Validations (listas desplegables) ───────────────
        MAX_ROW = 502  # filas 4..502 = 499 filas disponibles (fila 1=header, 2-3=ejemplos)

        def _add_list_validation(ws, col_letter, options_str, allow_blank=True):
            """Agrega una validación de lista desplegable a una columna."""
            dv = DataValidation(
                type='list',
                formula1=f'"{options_str}"',
                allow_blank=allow_blank,
                showErrorMessage=True,
                errorTitle='Valor no válido',
                error=f'Seleccione un valor de la lista.',
                showInputMessage=True,
            )
            dv.add(f'{col_letter}4:{col_letter}{MAX_ROW}')
            ws.add_data_validation(dv)

        # Tipo_Registro (col A = 1)
        _add_list_validation(ws, 'A', 'SOLO_PERSONAL,CON_USUARIO', allow_blank=False)

        # Género (col G = 7)
        _add_list_validation(ws, 'G', 'M,F', allow_blank=False)

        # Estado_Civil (col H = 8)
        _add_list_validation(ws, 'H', 'SOLTERO,SOLTERA,CASADO,CASADA,DIVORCIADO,DIVORCIADA,VIUDO,VIUDA,UNION_LIBRE')

        # Tipo_Contrato (col N = 14)
        _add_list_validation(ws, 'N', 'PERMANENTE,TEMPORAL,CONTRATO', allow_blank=False)

        # Área (col O = 15)
        _add_list_validation(ws, 'O', 'OPERACIONES,ADMINISTRACION,PEOPLE,SEGURIDAD,DELIVERY', allow_blank=False)

        # Nivel_Jerárquico (col P = 16)
        _add_list_validation(ws, 'P', 'OPERATIVO,SUPERVISOR,JEFE_AREA,GERENTE_CD', allow_blank=False)

        # Tipo_Posición (col R = 18)
        _add_list_validation(ws, 'R',
            'PICKER,CONTADOR,OPM,CONDUCTOR_PATIO,CARGADOR,AYUDANTE_ALMACEN,GUARDIA,CONDUCTOR_DELIVERY,ADMINISTRATIVO,OTRO',
            allow_blank=False)

        # Grupo_Sistema (col AA = 27) — dinámico desde la BD
        group_names = list(Group.objects.values_list('name', flat=True).order_by('name'))
        if group_names:
            group_str = ','.join(group_names)
            # Si la lista cabe en 255 chars usar fórmula directa, sino referencia a hoja oculta
            if len(group_str) <= 250:
                _add_list_validation(ws, 'AA', group_str)
            else:
                # Crear hoja auxiliar oculta con los grupos
                ws_groups = wb.create_sheet(title='_grupos')
                for i, name in enumerate(group_names, start=1):
                    ws_groups.cell(row=i, column=1, value=name)
                dv = DataValidation(
                    type='list',
                    formula1=f"='_grupos'!$A$1:$A${len(group_names)}",
                    allow_blank=True,
                    showErrorMessage=True,
                    errorTitle='Grupo no válido',
                    error='Seleccione un grupo de la lista.',
                )
                dv.add(f'AA4:AA{MAX_ROW}')
                ws.add_data_validation(dv)
                ws_groups.sheet_state = 'hidden'

        # ── Hoja 2: Valores válidos (referencia) ──────────────────
        ws_ref = wb.create_sheet(title='Valores Válidos')

        def _ref_section(ws, start_row, title, rows):
            """Escribe una sección de referencia con título y filas de valores."""
            t_cell = ws.cell(row=start_row, column=1, value=title)
            t_cell.font = Font(bold=True, size=11, color='FFFFFF')
            t_cell.fill = PatternFill(start_color='37474F', end_color='37474F', fill_type='solid')
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
            t_cell.alignment = Alignment(horizontal='center')
            ws.row_dimensions[start_row].height = 20
            r = start_row + 1
            hdr_fill = PatternFill(start_color='B0BEC5', end_color='B0BEC5', fill_type='solid')
            for h, col in [('Columna', 1), ('Valor en Excel', 2), ('Descripción', 3)]:
                c = ws.cell(row=r, column=col, value=h)
                c.font = Font(bold=True, size=9)
                c.fill = hdr_fill
            r += 1
            for row_data in rows:
                alt_fill = PatternFill(start_color='ECEFF1', end_color='ECEFF1', fill_type='solid') if (r % 2 == 0) else None
                for col, val in enumerate(row_data, start=1):
                    c = ws.cell(row=r, column=col, value=val)
                    c.font = Font(size=9)
                    if alt_fill:
                        c.fill = alt_fill
                    c.border = Border(bottom=Side(style='hair'))
                r += 1
            return r + 1

        ws_ref.column_dimensions['A'].width = 22
        ws_ref.column_dimensions['B'].width = 22
        ws_ref.column_dimensions['C'].width = 42

        row = 1
        row = _ref_section(ws_ref, row, '🔷 TIPO DE REGISTRO', [
            ('Tipo_Registro*', 'SOLO_PERSONAL', 'Crea solo el perfil. No tiene acceso al sistema. Dejar vacías las columnas de acceso.'),
            ('Tipo_Registro*', 'CON_USUARIO',   'Crea perfil + usuario del sistema con login. Completar Email_Sistema, Contrasena_Sistema.'),
        ])
        row = _ref_section(ws_ref, row, '👤 GÉNERO', [
            ('Genero*', 'M', 'Masculino'),
            ('Genero*', 'F', 'Femenino'),
        ])
        row = _ref_section(ws_ref, row, '💍 ESTADO CIVIL', [
            ('Estado_Civil', 'SOLTERO / SOLTERA', 'Soltero/a'),
            ('Estado_Civil', 'CASADO / CASADA',   'Casado/a'),
            ('Estado_Civil', 'DIVORCIADO / DIVORCIADA', 'Divorciado/a'),
            ('Estado_Civil', 'VIUDO / VIUDA',      'Viudo/a'),
            ('Estado_Civil', 'UNION_LIBRE',        'Unión libre'),
        ])
        row = _ref_section(ws_ref, row, '📋 TIPO DE CONTRATO', [
            ('Tipo_Contrato*', 'PERMANENTE', 'Contrato permanente'),
            ('Tipo_Contrato*', 'TEMPORAL',   'Contrato temporal'),
            ('Tipo_Contrato*', 'CONTRATO',   'Por contrato'),
        ])
        row = _ref_section(ws_ref, row, '🏢 ÁREA', [
            ('Area*', 'OPERACIONES',   'Área de Operaciones (picking, conteo, etc.)'),
            ('Area*', 'ADMINISTRACION','Área Administrativa'),
            ('Area*', 'PEOPLE',        'People / Recursos Humanos'),
            ('Area*', 'SEGURIDAD',     'Área de Seguridad / Portería'),
            ('Area*', 'DELIVERY',      'Delivery / Despachos'),
        ])
        row = _ref_section(ws_ref, row, '🏅 NIVEL JERÁRQUICO', [
            ('Nivel_Jerarquico*', 'OPERATIVO',   'Operativo (Picker, Contador, OPM, etc.)'),
            ('Nivel_Jerarquico*', 'SUPERVISOR',  'Supervisor de área o turno'),
            ('Nivel_Jerarquico*', 'JEFE_AREA',   'Jefe de Área'),
            ('Nivel_Jerarquico*', 'GERENTE_CD',  'Gerente del Centro de Distribución'),
        ])
        row = _ref_section(ws_ref, row, '🔧 TIPO DE POSICIÓN', [
            ('Tipo_Posicion*', 'PICKER',             'Picker'),
            ('Tipo_Posicion*', 'CONTADOR',           'Contador'),
            ('Tipo_Posicion*', 'OPM',                'Operador de Montacargas'),
            ('Tipo_Posicion*', 'CONDUCTOR_PATIO',    'Conductor de Patio'),
            ('Tipo_Posicion*', 'CARGADOR',           'Cargador'),
            ('Tipo_Posicion*', 'AYUDANTE_ALMACEN',   'Ayudante de Almacén'),
            ('Tipo_Posicion*', 'GUARDIA',            'Guardia de Seguridad'),
            ('Tipo_Posicion*', 'CONDUCTOR_DELIVERY', 'Conductor de Delivery'),
            ('Tipo_Posicion*', 'ADMINISTRATIVO',     'Administrativo'),
            ('Tipo_Posicion*', 'OTRO',               'Otro'),
        ])
        row = _ref_section(ws_ref, row, '👕 TALLAS EPP', [
            ('Talla_Camisa',   'Texto libre (máx. 10 caracteres)', 'Ej: XS, S, M, L, XL, XXL, XXXL'),
            ('Talla_Pantalon', 'Texto libre (máx. 10 caracteres)', 'Ej: 28, 30, 32, 34, 36'),
            ('Talla_Zapatos',  'Texto libre (máx. 10 caracteres)', 'Ej: 7, 8, 9, 10, 11'),
            ('Talla_Guantes',  'Texto libre (máx. 10 caracteres)', 'Ej: S, M, L, XL'),
            ('Talla_Casco',    'Texto libre (máx. 10 caracteres)', 'Ej: S, M, L'),
        ])
        # Grupos del sistema (dinámico)
        group_rows = [
            ('Grupo_Sistema', g.name, f'Grupo: {g.name}')
            for g in Group.objects.all().order_by('name')
        ]
        if not group_rows:
            group_rows = [('Grupo_Sistema', '(sin grupos)', 'No hay grupos configurados en el sistema')]
        row = _ref_section(ws_ref, row, '👥 GRUPOS DEL SISTEMA', group_rows)
        _ref_section(ws_ref, row, '📅 FORMATO DE FECHAS', [
            ('Fecha_Nacimiento* / Fecha_Ingreso*', 'DD/MM/YYYY', 'Ejemplo: 15/03/1995  ó  01/06/2022'),
        ])

        # ── Hoja 3: Instrucciones ─────────────────────────────────
        ws_ins = wb.create_sheet(title='Instrucciones')
        bold_blue = Font(bold=True, size=12, color='1976D2')
        bold_blk  = Font(bold=True, size=10)
        norm      = Font(size=10)

        lines = [
            ('INSTRUCCIONES — CARGA MASIVA DE PERSONAL', bold_blue),
            ('', None),
            ('PASO 1: Revisa la hoja "Valores Válidos" para conocer las opciones de cada columna.', bold_blk),
            ('PASO 2: Completa la hoja "Datos" a partir de la fila 4 (las filas 2 y 3 son ejemplos).', bold_blk),
            ('PASO 3: Sube el archivo y revisa quién se registrará antes de confirmar.', bold_blk),
            ('', None),
            ('TIPOS DE REGISTRO', bold_blk),
            ('  SOLO_PERSONAL → Solo crea el perfil del empleado. No tiene login al sistema.', norm),
            ('                   Típico para operativos: pickers, contadores, cargadores, etc.', norm),
            ('  CON_USUARIO   → Crea perfil + usuario del sistema con acceso a la plataforma.', norm),
            ('                   Requiere: Email_Sistema, Contrasena_Sistema (mín. 8 caracteres).', norm),
            ('', None),
            ('COLUMNAS OBLIGATORIAS (para todos)', bold_blk),
            ('  Tipo_Registro*, Nombres*, Apellidos*, Codigo_Empleado*, Fecha_Nacimiento*,', norm),
            ('  Genero*, Telefono*, Fecha_Ingreso*, Tipo_Contrato*, Area*, Nivel_Jerarquico*,', norm),
            ('  Puesto*, Tipo_Posicion*', norm),
            ('', None),
            ('COLUMNAS OBLIGATORIAS adicionales solo para CON_USUARIO', bold_blk),
            ('  Email_Sistema*, Contrasena_Sistema* (mínimo 8 caracteres)', norm),
            ('', None),
            ('COLUMNAS OPCIONALES (para todos)', bold_blk),
            ('  Num_Identidad, Estado_Civil, Email_Contacto, Direccion, Ciudad,', norm),
            ('  Talla_Camisa, Talla_Pantalon, Talla_Zapatos, Talla_Guantes, Talla_Casco', norm),
            ('', None),
            ('COLUMNAS OPCIONALES solo para CON_USUARIO', bold_blk),
            ('  Username_Sistema → Si se deja vacío se genera automáticamente (ej: carlos.ramirez)', norm),
            ('  Grupo_Sistema    → Nombre exacto del grupo del sistema (ej: SUPERVISOR)', norm),
            ('', None),
            ('REGLAS GENERALES', bold_blk),
            ('  • Máximo 500 registros por archivo', norm),
            ('  • No modificar los nombres de las columnas (fila 1)', norm),
            ('  • Los Codigo_Empleado deben ser únicos en el sistema', norm),
            ('  • Para CON_USUARIO, el Email_Sistema debe ser único en el sistema', norm),
            ('  • Fechas en formato DD/MM/YYYY (ej: 15/03/1995)', norm),
            ('  • El archivo debe ser .xlsx o .xls', norm),
            ('  • Podrás revisar quién se registrará y quién no antes de confirmar', norm),
        ]

        ws_ins.column_dimensions['A'].width = 90
        for row_idx, (text, font) in enumerate(lines, start=1):
            cell = ws_ins.cell(row=row_idx, column=1, value=text)
            if font:
                cell.font = font
            ws_ins.row_dimensions[row_idx].height = 16

        # ── Serializar ────────────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_carga_masiva_personal.xlsx"'
        return response

    # ──────────────────────────────────────────────────────────────
    # CARGA MASIVA — Preview (valida sin crear)
    # ──────────────────────────────────────────────────────────────
    @action(methods=['post'], detail=False, permission_classes=[IsAuthenticated],
            url_path='bulk-upload-preview')
    def bulk_upload_preview(self, request):
        """
        Parsea y valida el archivo Excel completo (PersonnelProfile + opcional usuario).
        Retorna valid_rows y error_rows. NO crea ningún registro.
        POST /api/users/bulk-upload-preview/
        Body (multipart): file=<xlsx>, centro_distribucion=<id>
        """
        if not request.user.has_perm('user.add_usermodel'):
            return Response({'detail': 'No tienes permiso para realizar esta acción.'}, status=403)

        file = request.FILES.get('file')
        centro_id = request.data.get('centro_distribucion')

        if not file:
            return Response({'detail': 'Se requiere el archivo Excel.'}, status=400)
        if not centro_id:
            return Response({'detail': 'Se requiere el centro de distribución.'}, status=400)

        try:
            centro = DistributorCenter.objects.get(pk=centro_id)
        except DistributorCenter.DoesNotExist:
            return Response({'detail': 'El centro de distribución no existe.'}, status=400)

        filename = file.name.lower()
        if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
            return Response({'detail': 'Solo se aceptan archivos .xlsx o .xls.'}, status=400)

        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        except Exception:
            return Response({'detail': 'El archivo no es un Excel válido.'}, status=400)

        ws = wb.active
        # Omitir fila 1 (encabezados) y empezar desde fila 2
        all_rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        # Filtrar filas vacías y filas de ejemplo (fila 2 = SOLO_PERSONAL, fila 3 = CON_USUARIO ejemplo)
        def _str(v): return str(v).strip() if v is not None else ''
        data_rows = [r for r in all_rows if any(
            _str(c) not in ('', 'SOLO_PERSONAL', 'CON_USUARIO') or
            any(_str(r[i]) for i in range(1, min(len(r), 5)))
            for c in [r[0] if r else None]
        )]
        # Filtrado simple: al menos la columna Nombres (índice 1) tiene valor
        data_rows = [r for r in all_rows if len(r) > 1 and _str(r[1]) != '']

        if not data_rows:
            return Response({'detail': 'El archivo no contiene datos (recuerda llenar desde la fila 4).'}, status=400)
        if len(data_rows) > 500:
            return Response({'detail': f'El archivo contiene {len(data_rows)} filas. El máximo permitido es 500.'}, status=400)

        valid_rows, error_rows = [], []
        seen_codigos, seen_emails_sys, seen_usernames = {}, {}, {}

        def _parse_date(raw):
            """Parsea fechas en formato DD/MM/YYYY o datetime de Excel."""
            from datetime import date, datetime
            if raw is None: return None
            if isinstance(raw, (date, datetime)):
                return raw.strftime('%Y-%m-%d')
            s = str(raw).strip()
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                try:
                    return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
                except ValueError:
                    pass
            return None

        for row_idx, row in enumerate(data_rows, start=2):
            def _c(idx): return _str(row[idx]) if len(row) > idx else ''

            tipo            = _c(0).upper()
            first_name      = _c(1)
            last_name       = _c(2)
            employee_code   = _c(3)
            personal_id     = _c(4) or None
            birth_date_raw  = row[5] if len(row) > 5 else None
            genero_raw      = _c(6).upper()
            estado_civil    = _c(7).upper()
            telefono        = _c(8)
            email_contacto  = _c(9)
            direccion       = _c(10)
            ciudad          = _c(11)
            hire_date_raw   = row[12] if len(row) > 12 else None
            contrato_raw    = _c(13).upper()
            area_raw        = _c(14).upper()
            jerarquia_raw   = _c(15).upper()
            puesto          = _c(16)
            posicion_raw    = _c(17).upper()
            talla_camisa    = _c(18) or None
            talla_pantalon  = _c(19) or None
            talla_zapatos   = _c(20) or None
            talla_guantes   = _c(21) or None
            talla_casco     = _c(22) or None
            email_sistema   = _c(23).lower() or None
            username_sys    = _c(24) or None
            password_sys    = _c(25) or None
            grupo_sys       = _c(26) or None

            errs = []

            # Tipo
            if tipo not in ('SOLO_PERSONAL', 'CON_USUARIO'):
                errs.append({'campo': 'Tipo_Registro', 'mensaje': 'Debe ser SOLO_PERSONAL o CON_USUARIO.'})

            # Campos obligatorios comunes
            if not first_name:  errs.append({'campo': 'Nombres', 'mensaje': 'Obligatorio.'})
            if not last_name:   errs.append({'campo': 'Apellidos', 'mensaje': 'Obligatorio.'})
            if not employee_code: errs.append({'campo': 'Codigo_Empleado', 'mensaje': 'Obligatorio.'})
            elif PersonnelProfile.objects.filter(employee_code=employee_code).exists():
                errs.append({'campo': 'Codigo_Empleado', 'mensaje': 'Ya existe en el sistema.'})
            elif employee_code in seen_codigos:
                errs.append({'campo': 'Codigo_Empleado', 'mensaje': f'Duplicado en fila {seen_codigos[employee_code]}.'})

            birth_date = _parse_date(birth_date_raw)
            if not birth_date: errs.append({'campo': 'Fecha_Nacimiento', 'mensaje': 'Obligatorio. Formato: DD/MM/YYYY.'})

            genero = self._GENERO_MAP.get(genero_raw)
            if not genero:
                errs.append({'campo': 'Genero', 'mensaje': 'Debe ser M o F.'})

            if not telefono: errs.append({'campo': 'Telefono', 'mensaje': 'Obligatorio.'})

            hire_date = _parse_date(hire_date_raw)
            if not hire_date: errs.append({'campo': 'Fecha_Ingreso', 'mensaje': 'Obligatorio. Formato: DD/MM/YYYY.'})

            contrato = self._CONTRATO_MAP.get(contrato_raw)
            if not contrato:
                errs.append({'campo': 'Tipo_Contrato', 'mensaje': 'Valores: PERMANENTE / TEMPORAL / CONTRATO.'})

            area_code = self._AREA_MAP.get(area_raw)
            if not area_code:
                errs.append({'campo': 'Area', 'mensaje': 'Valores: OPERACIONES / ADMINISTRACION / PEOPLE / SEGURIDAD / DELIVERY.'})

            jerarquia = self._JERARQUIA_MAP.get(jerarquia_raw)
            if not jerarquia:
                errs.append({'campo': 'Nivel_Jerarquico', 'mensaje': 'Valores: OPERATIVO / SUPERVISOR / JEFE_AREA / GERENTE_CD.'})

            if not puesto: errs.append({'campo': 'Puesto', 'mensaje': 'Obligatorio.'})

            posicion = self._POSICION_MAP.get(posicion_raw)
            if not posicion:
                errs.append({'campo': 'Tipo_Posicion', 'mensaje': 'Ver hoja "Valores Válidos" para opciones.'})

            # Tallas (opcionales, texto libre, máximo 10 caracteres)
            for campo, valor, nombre in [
                ('Talla_Camisa', talla_camisa, 'camisa'),
                ('Talla_Pantalon', talla_pantalon, 'pantalón'),
                ('Talla_Zapatos', talla_zapatos, 'zapatos'),
                ('Talla_Guantes', talla_guantes, 'guantes'),
                ('Talla_Casco', talla_casco, 'casco'),
            ]:
                if valor and len(valor) > 10:
                    errs.append({'campo': campo, 'mensaje': f'Máximo 10 caracteres para talla de {nombre}.'})

            # Campos adicionales para CON_USUARIO
            if tipo == 'CON_USUARIO':
                if not email_sistema:
                    errs.append({'campo': 'Email_Sistema', 'mensaje': 'Obligatorio para CON_USUARIO.'})
                else:
                    if UserModel.objects.filter(email__iexact=email_sistema).exists():
                        errs.append({'campo': 'Email_Sistema', 'mensaje': 'Email ya registrado en el sistema.'})
                    elif email_sistema in seen_emails_sys:
                        errs.append({'campo': 'Email_Sistema', 'mensaje': f'Duplicado en fila {seen_emails_sys[email_sistema]}.'})

                if not password_sys or len(password_sys) < 8:
                    errs.append({'campo': 'Contrasena_Sistema', 'mensaje': 'Obligatorio para CON_USUARIO (mín. 8 caracteres).'})

                if username_sys:
                    if UserModel.objects.filter(username__iexact=username_sys).exists():
                        errs.append({'campo': 'Username_Sistema', 'mensaje': 'Username ya existe.'})
                    elif username_sys.lower() in seen_usernames:
                        errs.append({'campo': 'Username_Sistema', 'mensaje': f'Duplicado en fila {seen_usernames[username_sys.lower()]}.'})

            # Registrar vistos (solo si no hay error ya en ese campo)
            if employee_code and not any(e['campo'] == 'Codigo_Empleado' for e in errs):
                seen_codigos[employee_code] = row_idx
            if email_sistema and not any(e['campo'] == 'Email_Sistema' for e in errs):
                seen_emails_sys[email_sistema] = row_idx
            if username_sys and not any(e['campo'] == 'Username_Sistema' for e in errs):
                seen_usernames[username_sys.lower()] = row_idx

            # Auto-username
            auto_username = username_sys
            if tipo == 'CON_USUARIO' and not username_sys and not errs:
                auto_username = self._generate_username_from_name(first_name, last_name)

            row_summary = {
                'tipo': tipo,
                'first_name': first_name,
                'last_name': last_name,
                'employee_code': employee_code,
                'personal_id': personal_id,
                'birth_date': birth_date,
                'gender': genero,
                'marital_status': self._ESTADO_CIVIL_MAP.get(estado_civil, ''),
                'phone': telefono,
                'email': email_contacto,
                'address': direccion,
                'city': ciudad,
                'hire_date': hire_date,
                'contract_type': contrato,
                'area': area_code,
                'hierarchy_level': jerarquia,
                'position': puesto,
                'position_type': posicion,
                'shirt_size': talla_camisa,
                'pants_size': talla_pantalon,
                'shoe_size': talla_zapatos,
                'glove_size': talla_guantes,
                'helmet_size': talla_casco,
                # solo CON_USUARIO
                'email_sistema': email_sistema,
                'username': auto_username,
                'grupo_sistema': grupo_sys,
            }

            if errs:
                error_rows.append({'fila': row_idx, 'datos': row_summary, 'errores': errs})
            else:
                valid_rows.append({'fila': row_idx, **row_summary})

        return Response({
            'centro_distribucion': centro_id,
            'centro_distribucion_name': centro.name,
            'total_filas': len(data_rows),
            'filas_validas': len(valid_rows),
            'filas_con_error': len(error_rows),
            'valid_rows': valid_rows,
            'error_rows': error_rows,
        }, status=200)

    # ──────────────────────────────────────────────────────────────
    # CARGA MASIVA — Confirm (crea registros confirmados)
    # ──────────────────────────────────────────────────────────────
    @action(methods=['post'], detail=False, permission_classes=[IsAuthenticated],
            url_path='bulk-upload-confirm')
    def bulk_upload_confirm(self, request):
        """
        Crea PersonnelProfile (y opcionalmente UserModel) para las filas confirmadas.
        POST /api/users/bulk-upload-confirm/
        Body (JSON): { centro_distribucion: id, rows: [...] }
        """
        if not request.user.has_perm('user.add_usermodel'):
            return Response({'detail': 'No tienes permiso para realizar esta acción.'}, status=403)

        centro_id = request.data.get('centro_distribucion')
        rows = request.data.get('rows', [])

        if not centro_id:
            return Response({'detail': 'Se requiere el centro de distribución.'}, status=400)
        if not rows:
            return Response({'detail': 'No hay filas para procesar.'}, status=400)
        if len(rows) > 500:
            return Response({'detail': 'El máximo permitido es 500 registros.'}, status=400)

        try:
            centro = DistributorCenter.objects.get(pk=centro_id)
        except DistributorCenter.DoesNotExist:
            return Response({'detail': 'El centro de distribución no existe.'}, status=400)

        created = []

        try:
            with transaction.atomic():
              for row in rows:
                tipo = str(row.get('tipo', 'SOLO_PERSONAL')).upper()

                # Obtener el área
                area_code = row.get('area')
                try:
                    area_obj = Area.objects.get(code=area_code)
                except Area.DoesNotExist:
                    raise ValueError(f"Área '{area_code}' no encontrada.")

                sistema_user = None
                if tipo == 'CON_USUARIO':
                    email_sis  = str(row.get('email_sistema', '')).strip().lower()
                    username   = str(row.get('username') or '').strip()
                    password   = str(row.get('password') or '').strip()
                    grupo_name = str(row.get('grupo_sistema') or '').strip()

                    if not username:
                        username = self._generate_username_from_name(
                            row.get('first_name', ''), row.get('last_name', '')
                        )

                    sistema_user = UserModel(
                        first_name=row.get('first_name', ''),
                        last_name=row.get('last_name', ''),
                        email=email_sis,
                        username=username,
                        centro_distribucion=centro,
                        is_active=True,
                    )
                    sistema_user.set_password(password)
                    sistema_user.save()
                    sistema_user.distributions_centers.add(centro)

                    if grupo_name:
                        try:
                            sistema_user.groups.add(Group.objects.get(name__iexact=grupo_name))
                        except Group.DoesNotExist:
                            pass

                profile = PersonnelProfile(
                    user=sistema_user,
                    employee_code=row.get('employee_code', ''),
                    first_name=row.get('first_name', '').upper(),
                    last_name=row.get('last_name', '').upper(),
                    email=row.get('email', ''),
                    personal_id=row.get('personal_id') or None,
                    birth_date=row.get('birth_date'),
                    gender=row.get('gender', 'M'),
                    marital_status=row.get('marital_status', ''),
                    phone=row.get('phone', ''),
                    personal_email=row.get('email', ''),
                    address=row.get('address', ''),
                    city=row.get('city', ''),
                    area=area_obj,
                    hierarchy_level=row.get('hierarchy_level', 'OPERATIVE'),
                    position=row.get('position', ''),
                    position_type=row.get('position_type', 'OTHER'),
                    hire_date=row.get('hire_date'),
                    contract_type=row.get('contract_type', 'PERMANENT'),
                    shirt_size=row.get('shirt_size') or '',
                    pants_size=row.get('pants_size') or '',
                    shoe_size=row.get('shoe_size') or '',
                    glove_size=row.get('glove_size') or '',
                    helmet_size=row.get('helmet_size') or '',
                    primary_distributor_center=centro,
                    is_active=True,
                    created_by=request.user,
                )
                profile.save()
                profile.distributor_centers.add(centro)

                created.append({
                    'employee_code': profile.employee_code,
                    'full_name': profile.full_name,
                    'tipo': tipo,
                    'tiene_usuario': sistema_user is not None,
                    'username': sistema_user.username if sistema_user else None,
                })

        except ValueError as e:
            return Response({'detail': str(e)}, status=400)

        return Response({
            'status': 'success',
            'created': len(created),
            'registros': created,
        }, status=201)

    @staticmethod
    def _generate_username_from_name(first_name: str, last_name: str) -> str:
        """Genera un username único basado en nombre y apellido."""
        def normalize(text):
            text = unicodedata.normalize('NFD', text)
            text = text.encode('ascii', 'ignore').decode('utf-8')
            return text.lower().replace(' ', '')

        first = normalize(first_name)
        last = normalize(last_name)
        base = f"{first}.{last}" if first and last else first or last or 'usuario'
        username = base
        counter = 1
        while UserModel.objects.filter(username=username).exists():
            username = f"{base}{counter}"
            counter += 1
        return username





# Filter by log
class LogEntryFilter(django_filters.FilterSet):
    user = django_filters.ModelMultipleChoiceFilter(
        field_name='user__username',
        to_field_name='username',
        queryset=User.objects.all()

    )
    content_type = django_filters.CharFilter(field_name='content_type__model', lookup_expr='icontains')
    change_message = django_filters.CharFilter(field_name='change_message', lookup_expr='icontains')

    class Meta:
        model = LogEntry
        fields = ['user', 'content_type', 'action_flag', 'change_message']


# ViewSets by LogEntry
@permission_classes([IsAdminUser])
class LogEntryViewSet(mixins.RetrieveModelMixin,
                      mixins.UpdateModelMixin,
                      mixins.ListModelMixin,
                      viewsets.GenericViewSet):
    queryset = LogEntry.objects.all()
    serializer_class = LogEntrySerializer
    filter_backends = (DjangoFilterBackend, SearchFilter, OrderingFilter)
    filterset_class = LogEntryFilter
    search_fields = ('user__username', 'content_type__model', 'action_flag', 'change_message')
    ordering_fields = ('user__username', 'content_type__model', 'action_flag', 'change_message')

# ViewSets by DetailGroup
class DetailGroupViewSet(mixins.RetrieveModelMixin,
                            mixins.ListModelMixin,
                            viewsets.GenericViewSet):
        queryset = DetailGroup.objects.all()
        serializer_class = DetailGroupSerializer
        filter_backends = (SearchFilter, OrderingFilter)
        search_fields = ('group__name')
        ordering_fields = ('group__name')
