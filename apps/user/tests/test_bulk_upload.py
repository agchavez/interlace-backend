"""
Tests para la funcionalidad de Carga Masiva de Personal.

Cubre:
  - TestBulkUploadTemplate    → Descarga de plantilla Excel (3 hojas)
  - TestBulkUploadPreview     → Validación/previsualización sin crear registros
  - TestBulkUploadConfirm     → Creación real de PersonnelProfile (+ UserModel opcional)
"""
import io
from django.test import TestCase
from django.contrib.auth.models import Group, Permission
from django.contrib.contenttypes.models import ContentType
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken

import openpyxl

from apps.user.models import UserModel
from apps.maintenance.models.distributor_center import DistributorCenter
from apps.personnel.models import PersonnelProfile, Area


# ──────────────────────────────────────────────────────────────────────────────
# Constantes — cabeceras de la plantilla (27 columnas)
# ──────────────────────────────────────────────────────────────────────────────

TEMPLATE_HEADERS = [
    'Tipo_Registro*', 'Nombres*', 'Apellidos*', 'Codigo_Empleado*',
    'Num_Identidad', 'Fecha_Nacimiento*', 'Genero*', 'Estado_Civil',
    'Telefono*', 'Email_Contacto', 'Direccion', 'Ciudad',
    'Fecha_Ingreso*', 'Tipo_Contrato*', 'Area*', 'Nivel_Jerarquico*',
    'Puesto*', 'Tipo_Posicion*',
    'Talla_Camisa', 'Talla_Pantalon', 'Talla_Zapatos', 'Talla_Guantes', 'Talla_Casco',
    'Email_Sistema', 'Username_Sistema', 'Contrasena_Sistema', 'Grupo_Sistema',
]


# ──────────────────────────────────────────────────────────────────────────────
# Helpers de bajo nivel
# ──────────────────────────────────────────────────────────────────────────────

def _get_token(user: UserModel) -> str:
    """Retorna el access token JWT para un usuario."""
    return str(RefreshToken.for_user(user).access_token)


def _make_excel(rows: list, include_headers: bool = True) -> io.BytesIO:
    """
    Crea un archivo Excel en memoria con las 27 columnas de la plantilla.

    rows: lista de listas con valores en el orden de TEMPLATE_HEADERS.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    if include_headers:
        ws.append(TEMPLATE_HEADERS)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = 'test.xlsx'
    return buf


def _solo_personal_row(
    tipo='SOLO_PERSONAL',
    nombres='Maria', apellidos='Lopez', codigo='OP-001',
    identidad='', nacimiento='15/03/1995', genero='F',
    civil='SOLTERA', telefono='9999-1234', email='',
    direccion='', ciudad='',
    ingreso='01/06/2022', contrato='PERMANENTE', area='OPERACIONES',
    jerarquia='OPERATIVO', puesto='Picker', posicion='PICKER',
    camisa='M', pantalon='32', zapatos='8', guantes='M', casco='M',
    email_sys='', username_sys='', password_sys='', grupo_sys='',
):
    """Retorna una fila lista para insertar en Excel (27 columnas)."""
    return [
        tipo, nombres, apellidos, codigo, identidad, nacimiento, genero,
        civil, telefono, email, direccion, ciudad,
        ingreso, contrato, area, jerarquia, puesto, posicion,
        camisa, pantalon, zapatos, guantes, casco,
        email_sys, username_sys, password_sys, grupo_sys,
    ]


def _con_usuario_row(
    nombres='Carlos', apellidos='Ramirez', codigo='SUP-001',
    ingreso='15/01/2019', area='OPERACIONES', jerarquia='SUPERVISOR',
    puesto='Supervisor Turno', posicion='AYUDANTE_ALMACEN',
    email_sys='carlos@empresa.com', password_sys='SecurePass1!',
    username_sys='', grupo_sys='',
):
    """Fila de tipo CON_USUARIO con valores por defecto razonables."""
    return [
        'CON_USUARIO', nombres, apellidos, codigo,
        '', '20/07/1988', 'M', 'CASADO', '9888-5678', '',
        '', '',
        ingreso, 'PERMANENTE', area, jerarquia, puesto, posicion,
        'L', '34', '9', 'L', 'L',
        email_sys, username_sys, password_sys, grupo_sys,
    ]


def _make_user(email, username, password='TestPass123!', with_add_perm=False, **kwargs):
    """Crea un UserModel de prueba, opcionalmente con permiso add_usermodel."""
    user = UserModel.objects.create(
        email=email,
        username=username,
        first_name='Test',
        last_name='User',
        is_active=True,
        **kwargs
    )
    user.set_password(password)
    user.save()
    if with_add_perm:
        ct = ContentType.objects.get_for_model(UserModel)
        perm = Permission.objects.get(codename='add_usermodel', content_type=ct)
        user.user_permissions.add(perm)
    return user


def _confirm_row(
    tipo='SOLO_PERSONAL',
    employee_code='TEST-0001',
    first_name='Maria',
    last_name='Lopez',
    email_contacto='',
    area='OPERATIONS',
    email_sistema=None,
    password=None,
    username=None,
    grupo_sistema=None,
):
    """Construye un dict listo para enviar al endpoint bulk-upload-confirm."""
    row = {
        'tipo': tipo,
        'first_name': first_name,
        'last_name': last_name,
        'employee_code': employee_code,
        'personal_id': None,
        'birth_date': '1995-03-15',
        'gender': 'F',
        'marital_status': 'SINGLE',
        'phone': '9999-1234',
        'email': email_contacto,
        'address': 'Dirección de prueba',
        'city': 'Tegucigalpa',
        'hire_date': '2022-06-01',
        'contract_type': 'PERMANENT',
        'area': area,
        'hierarchy_level': 'OPERATIVE',
        'position': 'Picker de prueba',
        'position_type': 'PICKER',
        'shirt_size': 'M',
        'pants_size': '32',
        'shoe_size': '8',
        'glove_size': 'M',
        'helmet_size': 'M',
        'email_sistema': email_sistema,
        'username': username,
        'grupo_sistema': grupo_sistema,
    }
    if password:
        row['password'] = password
    return row


# ──────────────────────────────────────────────────────────────────────────────
# Base TestCase
# ──────────────────────────────────────────────────────────────────────────────

class BulkUploadBaseTestCase(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.centro = DistributorCenter.objects.create(name='Centro Test')

        # Área requerida por las filas de prueba
        self.area_ops, _ = Area.objects.get_or_create(
            code='OPERATIONS',
            defaults={'name': 'Operaciones'},
        )

        # Usuario con permiso add_usermodel
        self.admin = _make_user('admin@test.com', 'admintest', with_add_perm=True)
        self.admin_token = _get_token(self.admin)

        # Usuario sin permiso
        self.regular = _make_user('regular@test.com', 'regulartest')
        self.regular_token = _get_token(self.regular)

    def auth_admin(self):
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {self.admin_token}')

    def auth_regular(self):
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {self.regular_token}')

    def no_auth(self):
        self.client.credentials()


# ──────────────────────────────────────────────────────────────────────────────
# 1. TestBulkUploadTemplate
# ──────────────────────────────────────────────────────────────────────────────

class TestBulkUploadTemplate(BulkUploadBaseTestCase):

    URL = '/api/users/bulk-upload-template/'

    def test_descarga_plantilla_ok(self):
        """Usuario con permiso descarga el archivo .xlsx correctamente."""
        self.auth_admin()
        response = self.client.get(self.URL)
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            response['Content-Type'],
        )
        self.assertIn('plantilla_carga_masiva_personal.xlsx', response['Content-Disposition'])

    def test_descarga_plantilla_contiene_columnas_nuevas(self):
        """El archivo descargado contiene las columnas de la nueva plantilla."""
        self.auth_admin()
        response = self.client.get(self.URL)
        self.assertEqual(response.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb['Datos']
        headers = [ws.cell(row=1, column=i).value for i in range(1, 28)]
        self.assertIn('Tipo_Registro*', headers)
        self.assertIn('Nombres*', headers)
        self.assertIn('Apellidos*', headers)
        self.assertIn('Codigo_Empleado*', headers)
        self.assertIn('Fecha_Nacimiento*', headers)
        self.assertIn('Genero*', headers)
        self.assertIn('Area*', headers)
        self.assertIn('Email_Sistema', headers)
        self.assertIn('Contrasena_Sistema', headers)

    def test_descarga_plantilla_contiene_filas_de_ejemplo(self):
        """La hoja Datos tiene al menos 2 filas de ejemplo."""
        self.auth_admin()
        response = self.client.get(self.URL)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb['Datos']
        # Filas 2 y 3 son los ejemplos SOLO_PERSONAL y CON_USUARIO
        self.assertEqual(ws.cell(row=2, column=1).value, 'SOLO_PERSONAL')
        self.assertEqual(ws.cell(row=3, column=1).value, 'CON_USUARIO')

    def test_descarga_plantilla_contiene_hoja_valores_validos(self):
        """El archivo contiene la hoja 'Valores Válidos' de referencia."""
        self.auth_admin()
        response = self.client.get(self.URL)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        self.assertIn('Valores Válidos', wb.sheetnames)

    def test_descarga_plantilla_contiene_hoja_instrucciones(self):
        """El archivo contiene la hoja 'Instrucciones'."""
        self.auth_admin()
        response = self.client.get(self.URL)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        self.assertIn('Instrucciones', wb.sheetnames)

    def test_descarga_plantilla_sin_autenticacion(self):
        """Sin JWT retorna 401."""
        self.no_auth()
        response = self.client.get(self.URL)
        self.assertEqual(response.status_code, 401)

    def test_descarga_plantilla_sin_permiso(self):
        """Usuario sin permiso add_usermodel retorna 403."""
        self.auth_regular()
        response = self.client.get(self.URL)
        self.assertEqual(response.status_code, 403)


# ──────────────────────────────────────────────────────────────────────────────
# 2. TestBulkUploadPreview
# ──────────────────────────────────────────────────────────────────────────────

class TestBulkUploadPreview(BulkUploadBaseTestCase):

    URL = '/api/users/bulk-upload-preview/'

    def _post(self, excel_rows, centro_id=None):
        buf = _make_excel(excel_rows)
        return self.client.post(
            self.URL,
            data={
                'file': buf,
                'centro_distribucion': centro_id or self.centro.pk,
            },
            format='multipart',
        )

    # ── Casos felices ────────────────────────────────────────────────────────

    def test_preview_solo_personal_valido(self):
        """Filas SOLO_PERSONAL válidas aparecen en valid_rows sin crear registros."""
        self.auth_admin()
        rows = [
            _solo_personal_row(nombres='Ana', apellidos='Lopez', codigo='OP-001'),
            _solo_personal_row(nombres='Luis', apellidos='Gomez', codigo='OP-002'),
        ]
        response = self._post(rows)
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['filas_validas'], 2)
        self.assertEqual(data['filas_con_error'], 0)
        # No se creó ningún perfil
        self.assertFalse(PersonnelProfile.objects.filter(employee_code='OP-001').exists())

    def test_preview_con_usuario_valido(self):
        """Filas CON_USUARIO válidas aparecen en valid_rows con username."""
        self.auth_admin()
        rows = [_con_usuario_row(codigo='SUP-001')]
        response = self._post(rows)
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['filas_validas'], 1)
        self.assertEqual(data['filas_con_error'], 0)

    def test_preview_username_auto_generado_cuando_vacio(self):
        """CON_USUARIO sin Username_Sistema → valid_row incluye username auto-generado."""
        self.auth_admin()
        rows = [_con_usuario_row(codigo='SUP-010', username_sys='')]
        response = self._post(rows)
        data = response.json()
        self.assertEqual(data['filas_validas'], 1)
        generated = data['valid_rows'][0]['username']
        self.assertTrue(len(generated) > 0)

    def test_preview_contiene_datos_completos_en_valid_rows(self):
        """valid_rows incluye todos los campos del PersonnelProfile."""
        self.auth_admin()
        rows = [_solo_personal_row(codigo='OP-010')]
        response = self._post(rows)
        data = response.json()
        valid = data['valid_rows'][0]
        for campo in ['tipo', 'first_name', 'last_name', 'employee_code',
                      'gender', 'area', 'hierarchy_level', 'position', 'position_type',
                      'hire_date', 'contract_type']:
            self.assertIn(campo, valid)

    # ── Validaciones de campos obligatorios ─────────────────────────────────

    def test_preview_tipo_invalido(self):
        """Tipo_Registro diferente a SOLO_PERSONAL/CON_USUARIO → error."""
        self.auth_admin()
        row = _solo_personal_row(tipo='INVALIDO', codigo='OP-020')
        response = self._post([row])
        data = response.json()
        self.assertEqual(data['filas_con_error'], 1)
        campos = [e['campo'] for e in data['error_rows'][0]['errores']]
        self.assertIn('Tipo_Registro', campos)

    def test_preview_genero_invalido(self):
        """Genero diferente a M/F → error."""
        self.auth_admin()
        row = _solo_personal_row(genero='OTRO', codigo='OP-021')
        response = self._post([row])
        data = response.json()
        self.assertEqual(data['filas_con_error'], 1)
        campos = [e['campo'] for e in data['error_rows'][0]['errores']]
        self.assertIn('Genero', campos)

    def test_preview_area_invalida(self):
        """Area con valor no reconocido → error."""
        self.auth_admin()
        row = _solo_personal_row(area='DESCONOCIDA', codigo='OP-022')
        response = self._post([row])
        data = response.json()
        self.assertEqual(data['filas_con_error'], 1)
        campos = [e['campo'] for e in data['error_rows'][0]['errores']]
        self.assertIn('Area', campos)

    def test_preview_fecha_nacimiento_invalida(self):
        """Fecha_Nacimiento en formato incorrecto → error."""
        self.auth_admin()
        row = _solo_personal_row(nacimiento='1995-03-15-MALO', codigo='OP-023')
        response = self._post([row])
        data = response.json()
        self.assertEqual(data['filas_con_error'], 1)

    # ── Duplicados ──────────────────────────────────────────────────────────

    def test_preview_codigo_duplicado_en_bd(self):
        """Codigo_Empleado ya existente en PersonnelProfile → error."""
        self.auth_admin()
        # Crear un perfil con ese código en BD
        PersonnelProfile.objects.create(
            employee_code='DUP-001',
            first_name='Existente',
            last_name='Perfil',
            birth_date='1990-01-01',
            gender='M',
            phone='9999-0000',
            hire_date='2020-01-01',
            contract_type='PERMANENT',
            area=self.area_ops,
            hierarchy_level='OPERATIVE',
            position='Picker',
            position_type='PICKER',
            primary_distributor_center=self.centro,
            created_by=self.admin,
        )
        row = _solo_personal_row(codigo='DUP-001')
        response = self._post([row])
        data = response.json()
        self.assertEqual(data['filas_con_error'], 1)
        campos = [e['campo'] for e in data['error_rows'][0]['errores']]
        self.assertIn('Codigo_Empleado', campos)

    def test_preview_codigo_duplicado_dentro_del_archivo(self):
        """Mismo Codigo_Empleado en dos filas → error en la segunda."""
        self.auth_admin()
        rows = [
            _solo_personal_row(nombres='Pedro', apellidos='Ruiz', codigo='DUP-X'),
            _solo_personal_row(nombres='Maria', apellidos='Soto', codigo='DUP-X'),
        ]
        response = self._post(rows)
        data = response.json()
        self.assertEqual(data['filas_con_error'], 1)
        self.assertEqual(data['filas_validas'], 1)

    def test_preview_email_sistema_duplicado_en_bd(self):
        """Email_Sistema de CON_USUARIO ya registrado en UserModel → error."""
        self.auth_admin()
        _make_user('existente@empresa.com', 'existente_sis')
        row = _con_usuario_row(codigo='SUP-020', email_sys='existente@empresa.com')
        response = self._post([row])
        data = response.json()
        self.assertEqual(data['filas_con_error'], 1)
        campos = [e['campo'] for e in data['error_rows'][0]['errores']]
        self.assertIn('Email_Sistema', campos)

    def test_preview_email_sistema_duplicado_dentro_del_archivo(self):
        """Mismo Email_Sistema en dos filas CON_USUARIO → error en la segunda."""
        self.auth_admin()
        rows = [
            _con_usuario_row(nombres='A', apellidos='B', codigo='SUP-031', email_sys='mismo@empresa.com'),
            _con_usuario_row(nombres='C', apellidos='D', codigo='SUP-032', email_sys='mismo@empresa.com'),
        ]
        response = self._post(rows)
        data = response.json()
        self.assertEqual(data['filas_con_error'], 1)
        self.assertEqual(data['filas_validas'], 1)

    def test_preview_username_duplicado_en_bd(self):
        """Username_Sistema ya registrado en UserModel → error."""
        self.auth_admin()
        _make_user('otro@test.com', 'username_tomado')
        row = _con_usuario_row(codigo='SUP-033', email_sys='otro2@empresa.com',
                               username_sys='username_tomado')
        response = self._post([row])
        data = response.json()
        self.assertEqual(data['filas_con_error'], 1)
        campos = [e['campo'] for e in data['error_rows'][0]['errores']]
        self.assertIn('Username_Sistema', campos)

    def test_preview_contrasena_corta_con_usuario(self):
        """Contrasena_Sistema menor a 8 caracteres para CON_USUARIO → error."""
        self.auth_admin()
        row = _con_usuario_row(codigo='SUP-040', password_sys='abc')
        response = self._post([row])
        data = response.json()
        self.assertEqual(data['filas_con_error'], 1)
        campos = [e['campo'] for e in data['error_rows'][0]['errores']]
        self.assertIn('Contrasena_Sistema', campos)

    def test_preview_contrasena_requerida_para_con_usuario(self):
        """CON_USUARIO sin Contrasena_Sistema → error."""
        self.auth_admin()
        row = _con_usuario_row(codigo='SUP-041', password_sys='')
        response = self._post([row])
        data = response.json()
        self.assertEqual(data['filas_con_error'], 1)

    # ── Validaciones de request ──────────────────────────────────────────────

    def test_preview_sin_archivo(self):
        """Sin archivo en el body retorna 400."""
        self.auth_admin()
        response = self.client.post(
            self.URL,
            data={'centro_distribucion': self.centro.pk},
            format='multipart',
        )
        self.assertEqual(response.status_code, 400)

    def test_preview_formato_invalido(self):
        """Archivo no-Excel retorna 400."""
        self.auth_admin()
        buf = io.BytesIO(b'esto no es un excel')
        buf.name = 'datos.txt'
        response = self.client.post(
            self.URL,
            data={'file': buf, 'centro_distribucion': self.centro.pk},
            format='multipart',
        )
        self.assertEqual(response.status_code, 400)

    def test_preview_sin_centro_distribucion(self):
        """Sin centro_distribucion retorna 400."""
        self.auth_admin()
        buf = _make_excel([])
        response = self.client.post(
            self.URL,
            data={'file': buf},
            format='multipart',
        )
        self.assertEqual(response.status_code, 400)

    def test_preview_mas_de_500_filas(self):
        """Archivo con más de 500 filas retorna 400 con mención del límite."""
        self.auth_admin()
        rows = [
            _solo_personal_row(nombres=f'User{i}', codigo=f'OP-{i:04d}')
            for i in range(501)
        ]
        response = self._post(rows)
        self.assertEqual(response.status_code, 400)
        self.assertIn('500', response.json()['detail'])

    def test_preview_sin_autenticacion(self):
        """Sin JWT retorna 401."""
        self.no_auth()
        response = self._post([])
        self.assertEqual(response.status_code, 401)

    def test_preview_sin_permiso(self):
        """Usuario sin add_usermodel retorna 403."""
        self.auth_regular()
        response = self._post([])
        self.assertEqual(response.status_code, 403)

    def test_preview_incluye_nombre_centro_en_respuesta(self):
        """La respuesta incluye centro_distribucion_name."""
        self.auth_admin()
        rows = [_solo_personal_row(codigo='OP-099')]
        response = self._post(rows)
        data = response.json()
        self.assertIn('centro_distribucion_name', data)
        self.assertEqual(data['centro_distribucion_name'], 'Centro Test')


# ──────────────────────────────────────────────────────────────────────────────
# 3. TestBulkUploadConfirm
# ──────────────────────────────────────────────────────────────────────────────

class TestBulkUploadConfirm(BulkUploadBaseTestCase):

    URL = '/api/users/bulk-upload-confirm/'

    def _post(self, rows, centro_id=None):
        return self.client.post(
            self.URL,
            data={
                'centro_distribucion': centro_id or self.centro.pk,
                'rows': rows,
            },
            format='json',
        )

    def _rows(self, n=3, tipo='SOLO_PERSONAL'):
        return [
            _confirm_row(
                tipo=tipo,
                employee_code=f'TEST-{i:04d}',
                first_name=f'Empleado{i}',
                last_name='Test',
                email_contacto=f'contacto{i}@test.com',
                email_sistema=f'sistema{i}@empresa.com' if tipo == 'CON_USUARIO' else None,
                password='SecurePass1!' if tipo == 'CON_USUARIO' else None,
            )
            for i in range(n)
        ]

    # ── Creación de PersonnelProfile ─────────────────────────────────────────

    def test_carga_exitosa_solo_personal(self):
        """Filas SOLO_PERSONAL crean PersonnelProfile sin UserModel."""
        self.auth_admin()
        rows = self._rows(3, tipo='SOLO_PERSONAL')
        response = self._post(rows)
        self.assertEqual(response.status_code, 201)
        data = response.json()
        self.assertEqual(data['status'], 'success')
        self.assertEqual(data['created'], 3)
        # Los 3 perfiles existen en BD
        for row in rows:
            self.assertTrue(PersonnelProfile.objects.filter(
                employee_code=row['employee_code']
            ).exists())
        # Ningún UserModel nuevo creado para estos registros
        for row in rows:
            self.assertFalse(UserModel.objects.filter(email=row.get('email_sistema', '')).exists())

    def test_carga_exitosa_con_usuario(self):
        """Filas CON_USUARIO crean PersonnelProfile + UserModel."""
        self.auth_admin()
        rows = self._rows(2, tipo='CON_USUARIO')
        response = self._post(rows)
        self.assertEqual(response.status_code, 201)
        data = response.json()
        self.assertEqual(data['created'], 2)
        for row in rows:
            profile = PersonnelProfile.objects.get(employee_code=row['employee_code'])
            self.assertIsNotNone(profile.user)
            self.assertTrue(UserModel.objects.filter(email=row['email_sistema']).exists())

    def test_registros_respuesta_incluye_campos_correctos(self):
        """La respuesta incluye employee_code, full_name, tipo, tiene_usuario, username."""
        self.auth_admin()
        rows = [_confirm_row(employee_code='OP-R01')]
        response = self._post(rows)
        reg = response.json()['registros'][0]
        for campo in ['employee_code', 'full_name', 'tipo', 'tiene_usuario', 'username']:
            self.assertIn(campo, reg)
        self.assertFalse(reg['tiene_usuario'])
        self.assertIsNone(reg['username'])

    # ── Centro de distribución ───────────────────────────────────────────────

    def test_centro_distribucion_asignado_al_perfil(self):
        """PersonnelProfile tiene primary_distributor_center correcto."""
        self.auth_admin()
        rows = [_confirm_row(employee_code='OP-DC01')]
        self._post(rows)
        profile = PersonnelProfile.objects.get(employee_code='OP-DC01')
        self.assertEqual(profile.primary_distributor_center, self.centro)
        self.assertIn(self.centro, profile.distributor_centers.all())

    def test_centro_distribucion_asignado_al_usuario(self):
        """Para CON_USUARIO, UserModel tiene centro_distribucion asignado."""
        self.auth_admin()
        rows = [_confirm_row(
            tipo='CON_USUARIO', employee_code='SUP-DC01',
            email_sistema='sup.dc@empresa.com', password='SecurePass1!',
        )]
        self._post(rows)
        user = UserModel.objects.get(email='sup.dc@empresa.com')
        self.assertEqual(user.centro_distribucion, self.centro)
        self.assertIn(self.centro, user.distributions_centers.all())

    # ── Grupos ──────────────────────────────────────────────────────────────

    def test_grupo_asignado_correctamente(self):
        """CON_USUARIO con grupo_sistema válido → usuario queda en ese grupo."""
        self.auth_admin()
        grupo = Group.objects.create(name='OPERADOR_TEST')
        rows = [_confirm_row(
            tipo='CON_USUARIO', employee_code='SUP-GR01',
            email_sistema='grp.user@empresa.com', password='SecurePass1!',
            grupo_sistema='OPERADOR_TEST',
        )]
        self._post(rows)
        user = UserModel.objects.get(email='grp.user@empresa.com')
        self.assertIn(grupo, user.groups.all())

    def test_grupo_inexistente_no_rompe_creacion(self):
        """Grupo no existente → usuario se crea sin grupo, sin error."""
        self.auth_admin()
        rows = [_confirm_row(
            tipo='CON_USUARIO', employee_code='SUP-GR02',
            email_sistema='nogrp@empresa.com', password='SecurePass1!',
            grupo_sistema='GRUPO_QUE_NO_EXISTE',
        )]
        response = self._post(rows)
        self.assertEqual(response.status_code, 201)
        user = UserModel.objects.get(email='nogrp@empresa.com')
        self.assertEqual(user.groups.count(), 0)

    # ── Username auto-generado ───────────────────────────────────────────────

    def test_username_auto_generado_si_vacio(self):
        """CON_USUARIO con username vacío → UserModel tiene username generado."""
        self.auth_admin()
        rows = [_confirm_row(
            tipo='CON_USUARIO', employee_code='SUP-UN01',
            first_name='Rosa', last_name='Medina',
            email_sistema='rosa.medina@empresa.com', password='SecurePass1!',
            username=None,
        )]
        response = self._post(rows)
        self.assertEqual(response.status_code, 201)
        user = UserModel.objects.get(email='rosa.medina@empresa.com')
        self.assertNotEqual(user.username, '')

    # ── Contraseña hasheada ──────────────────────────────────────────────────

    def test_contrasena_almacenada_con_hash(self):
        """La contraseña del UserModel no se almacena en texto plano."""
        self.auth_admin()
        rows = [_confirm_row(
            tipo='CON_USUARIO', employee_code='SUP-PW01',
            email_sistema='pw.test@empresa.com', password='PlainText1!',
        )]
        self._post(rows)
        user = UserModel.objects.get(email='pw.test@empresa.com')
        self.assertNotEqual(user.password, 'PlainText1!')
        self.assertTrue(user.check_password('PlainText1!'))

    # ── Transacción atómica ──────────────────────────────────────────────────

    def test_transaccion_atomica_rollback_en_area_invalida(self):
        """Si una fila tiene área inválida, ningún registro se crea (rollback)."""
        self.auth_admin()
        rows = [
            _confirm_row(employee_code='OK-001', area='OPERATIONS'),
            _confirm_row(employee_code='OK-002', area='AREA_INVALIDA_QUE_NO_EXISTE'),
        ]
        response = self._post(rows)
        # La transacción falla (400 por ValueError)
        self.assertNotEqual(response.status_code, 201)
        # El primer perfil tampoco se creó (rollback)
        self.assertFalse(PersonnelProfile.objects.filter(employee_code='OK-001').exists())

    # ── Validaciones de request ──────────────────────────────────────────────

    def test_carga_sin_centro_distribucion(self):
        """Sin centro_distribucion retorna 400."""
        self.auth_admin()
        response = self.client.post(
            self.URL,
            data={'rows': self._rows(1)},
            format='json',
        )
        self.assertEqual(response.status_code, 400)

    def test_carga_sin_rows(self):
        """Body con rows vacío retorna 400."""
        self.auth_admin()
        response = self.client.post(
            self.URL,
            data={'centro_distribucion': self.centro.pk, 'rows': []},
            format='json',
        )
        self.assertEqual(response.status_code, 400)

    def test_carga_mas_de_500_filas(self):
        """Más de 500 filas retorna 400."""
        self.auth_admin()
        rows = [
            _confirm_row(employee_code=f'E-{i:04d}', first_name=f'E{i}')
            for i in range(501)
        ]
        response = self._post(rows)
        self.assertEqual(response.status_code, 400)

    def test_carga_sin_autenticacion(self):
        """Sin JWT retorna 401."""
        self.no_auth()
        response = self._post(self._rows(1))
        self.assertEqual(response.status_code, 401)

    def test_carga_sin_permiso(self):
        """Usuario sin add_usermodel retorna 403."""
        self.auth_regular()
        response = self._post(self._rows(1))
        self.assertEqual(response.status_code, 403)

    def test_centro_distribucion_inexistente_retorna_400(self):
        """Centro de distribución con ID inexistente retorna 400."""
        self.auth_admin()
        response = self._post(self._rows(1), centro_id=99999)
        self.assertEqual(response.status_code, 400)
